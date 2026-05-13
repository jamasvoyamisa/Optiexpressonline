"""
╔══════════════════════════════════════════════════════════════════════╗
║  GENERADOR DE EXCEL DE REVISIÓN — Histórico de checadas             ║
║  Lee todos los XLS/XLSX de la carpeta 'Reportes de checador'        ║
║  y genera un Excel consolidado para revisión antes de importar.     ║
║                                                                      ║
║  Uso:                                                               ║
║    python3 scripts/generar_excel_historico.py                       ║
║                                                                      ║
║  Salida:                                                            ║
║    ~/Desktop/historico_checadas_revision.xlsx                       ║
╚══════════════════════════════════════════════════════════════════════╝
"""

import os
import sys
import re
from datetime import datetime, date, timedelta
from collections import defaultdict

import xlrd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Rutas ────────────────────────────────────────────────────────────────────
BASE_DIR    = os.path.expanduser("~/Desktop/trabajo/Reportes de checador")
OUTPUT_FILE = os.path.expanduser("~/Desktop/historico_checadas_revision.xlsx")

# ── Colores ──────────────────────────────────────────────────────────────────
COLOR_HEADER      = "1E3A5F"   # azul oscuro
COLOR_SOLO_ENTRADA = "FFF3CD"  # amarillo suave — solo tiene entrada, falta salida
COLOR_SOLO_SALIDA  = "FFE0E0"  # rojo suave — solo tiene salida (raro)
COLOR_COMPLETO     = "FFFFFF"  # blanco — entrada+salida OK
COLOR_ERROR        = "F8D7DA"  # rojo — no reconocido / error


# ────────────────────────────────────────────────────────────────────────────
# Helpers de conversión de tiempo
# ────────────────────────────────────────────────────────────────────────────

def fraccion_a_hhmm(frac: float) -> str | None:
    """Convierte fracción de día (0.0-1.0) a 'HH:MM'. 0.0 → None."""
    if frac is None or frac <= 0:
        return None
    total_min = round(frac * 24 * 60)
    h, m = divmod(total_min, 60)
    if h > 23:
        return None
    return f"{h:02d}:{m:02d}"


def hhmm_str_a_hhmm(s: str) -> str | None:
    """Convierte 'HH:MM' o 'HH.MM' a 'HH:MM'. '00.00'/'00:00' → None."""
    if not s:
        return None
    s = s.strip().replace('.', ':')
    if s in ('00:00', ''):
        return None
    m = re.match(r'^(\d{1,2}):(\d{2})$', s)
    if not m:
        return None
    h, mi = int(m.group(1)), int(m.group(2))
    if h == 0 and mi == 0:
        return None
    if not (0 <= h <= 23 and 0 <= mi <= 59):
        return None
    return f"{h:02d}:{mi:02d}"


def split_par(raw: str):
    """Divide 'HH:MM-HH:MM' en (entrada, salida). Devuelve (None, None) si sin registro."""
    raw = str(raw).strip()
    if not raw or raw.startswith('00.00'):
        return None, None
    partes = raw.split('-', 1)
    e = hhmm_str_a_hhmm(partes[0]) if partes else None
    s = hhmm_str_a_hhmm(partes[1]) if len(partes) > 1 else None
    return e, s


def excel_serial_to_date(serial, datemode=0) -> date | None:
    try:
        return xlrd.xldate_as_datetime(float(serial), datemode).date()
    except Exception:
        return None


# ────────────────────────────────────────────────────────────────────────────
# Parsers por formato
# ────────────────────────────────────────────────────────────────────────────

def parse_formato_a_xls(path: str) -> list[dict]:
    """
    Formato A (xls): una fila por empleado+día con columnas:
    Nº Empleado | Nombre | Sección | Fecha(serial) | Ingreso(frac) | Salida(frac) | Día | Periodo
    """
    registros = []
    try:
        wb = xlrd.open_workbook(path)
        ws = wb.sheet_by_index(0)

        # Buscar fila de encabezado
        header_row = -1
        for r in range(min(ws.nrows, 10)):
            if str(ws.cell_value(r, 0)).strip() == 'Nº de Empleado':
                header_row = r
                break
        if header_row < 0:
            return registros

        for r in range(header_row + 1, ws.nrows):
            num_raw = str(ws.cell_value(r, 0)).strip()
            if not num_raw or not num_raw.replace('.', '').isdigit():
                continue

            num_emp = str(int(float(num_raw)))
            nombre  = str(ws.cell_value(r, 1)).strip()
            fecha_s = ws.cell_value(r, 3)
            ing_s   = ws.cell_value(r, 4)
            sal_s   = ws.cell_value(r, 5)

            fecha = excel_serial_to_date(fecha_s, wb.datemode)
            if not fecha:
                continue

            entrada = fraccion_a_hhmm(ing_s) if isinstance(ing_s, (int, float)) else None
            salida  = fraccion_a_hhmm(sal_s) if isinstance(sal_s, (int, float)) else None

            if not entrada and not salida:
                continue

            registros.append({
                'numero_empleado': num_emp,
                'nombre': nombre,
                'fecha': fecha,
                'entrada': entrada,
                'salida': salida,
            })
    except Exception as e:
        print(f"  [WARN] {os.path.basename(path)}: {e}")
    return registros


def parse_formato_b_xls(path: str) -> list[dict]:
    """
    Formato B (xls): bloques por empleado, una fila por día con:
    Nº Empleado | Fecha(serial) | Ingreso-Salida(HH:MM-HH:MM) | ... | Ausente-Regreso | ...
    """
    registros = []
    try:
        wb = xlrd.open_workbook(path)
        ws = wb.sheet_by_index(0)

        nombre_actual = ""

        for r in range(ws.nrows):
            col0 = str(ws.cell_value(r, 0)).strip()

            # Detectar fila de nombre de empleado
            if col0 in ('Nombre Completo', 'Nombre completo', ''):
                nombre_raw = str(ws.cell_value(r, 2)).strip()
                if nombre_raw and nombre_raw not in ('Nombre Completo', ''):
                    nombre_actual = nombre_raw
                continue

            # Saltar encabezados
            if col0 in ('Nº de Empleado', 'Sección'):
                continue

            # Fila de datos: col0 = número de empleado
            if not col0.replace('.', '').isdigit():
                continue

            num_emp = str(int(float(col0)))
            fecha_s = ws.cell_value(r, 1)
            ing_sal = str(ws.cell_value(r, 2)).strip()

            fecha = excel_serial_to_date(fecha_s, wb.datemode)
            if not fecha:
                continue

            entrada, salida = split_par(ing_sal)

            if not entrada and not salida:
                continue

            registros.append({
                'numero_empleado': num_emp,
                'nombre': nombre_actual,
                'fecha': fecha,
                'entrada': entrada,
                'salida': salida,
            })
    except Exception as e:
        print(f"  [WARN] {os.path.basename(path)}: {e}")
    return registros


def parse_formato_a_xlsx(path: str) -> list[dict]:
    """Formato A en .xlsx — misma estructura que A xls."""
    registros = []
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb.active

        header_row_idx = -1
        rows = list(ws.iter_rows(values_only=True))

        for i, row in enumerate(rows):
            if row and str(row[0] or '').strip() == 'Nº de Empleado':
                header_row_idx = i
                break
        if header_row_idx < 0:
            wb.close()
            return registros

        for row in rows[header_row_idx + 1:]:
            if not row or row[0] is None:
                continue
            num_raw = str(row[0]).strip()
            if not num_raw.replace('.', '').isdigit():
                continue

            num_emp = str(int(float(num_raw)))
            nombre  = str(row[1] or '').strip()

            fecha_s = row[3]
            if isinstance(fecha_s, (datetime, date)):
                fecha = fecha_s.date() if isinstance(fecha_s, datetime) else fecha_s
            else:
                fecha = excel_serial_to_date(fecha_s)
            if not fecha:
                continue

            ing_s = row[4]
            sal_s = row[5]
            entrada = fraccion_a_hhmm(ing_s) if isinstance(ing_s, (int, float)) else hhmm_str_a_hhmm(str(ing_s))
            salida  = fraccion_a_hhmm(sal_s) if isinstance(sal_s, (int, float)) else hhmm_str_a_hhmm(str(sal_s))

            if not entrada and not salida:
                continue

            registros.append({
                'numero_empleado': num_emp,
                'nombre': nombre,
                'fecha': fecha,
                'entrada': entrada,
                'salida': salida,
            })
        wb.close()
    except Exception as e:
        print(f"  [WARN] {os.path.basename(path)}: {e}")
    return registros


def parse_formato_b_xlsx(path: str) -> list[dict]:
    """Formato B en .xlsx."""
    registros = []
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        nombre_actual = ""

        for row in ws.iter_rows(values_only=True):
            if not row:
                continue
            col0 = str(row[0] or '').strip()

            if col0 in ('Nombre Completo', 'Nombre completo', ''):
                nombre_raw = str(row[2] or '').strip()
                if nombre_raw:
                    nombre_actual = nombre_raw
                continue

            if col0 in ('Nº de Empleado', 'Sección'):
                continue

            if not col0.replace('.', '').isdigit():
                continue

            num_emp = str(int(float(col0)))
            fecha_s = row[1]
            if isinstance(fecha_s, (datetime, date)):
                fecha = fecha_s.date() if isinstance(fecha_s, datetime) else fecha_s
            else:
                fecha = excel_serial_to_date(fecha_s)
            if not fecha:
                continue

            ing_sal = str(row[2] or '').strip()
            entrada, salida = split_par(ing_sal)

            if not entrada and not salida:
                continue

            registros.append({
                'numero_empleado': num_emp,
                'nombre': nombre_actual,
                'fecha': fecha,
                'entrada': entrada,
                'salida': salida,
            })
        wb.close()
    except Exception as e:
        print(f"  [WARN] {os.path.basename(path)}: {e}")
    return registros


# ────────────────────────────────────────────────────────────────────────────
# Detectar formato de un archivo
# ────────────────────────────────────────────────────────────────────────────

def detectar_formato(path: str, ext: str) -> str:
    """Devuelve 'A' o 'B'."""
    try:
        if ext == 'xls':
            wb = xlrd.open_workbook(path)
            ws = wb.sheet_by_index(0)
            for r in range(min(ws.nrows, 10)):
                if str(ws.cell_value(r, 0)).strip() == 'Nº de Empleado':
                    row = [str(ws.cell_value(r, c)) for c in range(ws.ncols)]
                    return 'B' if (len(row) > 2 and 'Ingreso - Salida' in row[2]) else 'A'
        else:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
            ws = wb.active
            for row in ws.iter_rows(max_row=10, values_only=True):
                if row and str(row[0] or '').strip() == 'Nº de Empleado':
                    cols = [str(v or '') for v in row]
                    wb.close()
                    return 'B' if (len(cols) > 2 and 'Ingreso - Salida' in cols[2]) else 'A'
            wb.close()
    except Exception:
        pass
    return 'A'


def parsear_archivo(path: str) -> list[dict]:
    ext = path.lower().rsplit('.', 1)[-1]
    fmt = detectar_formato(path, ext)
    if ext == 'xls':
        return parse_formato_b_xls(path) if fmt == 'B' else parse_formato_a_xls(path)
    else:
        return parse_formato_b_xlsx(path) if fmt == 'B' else parse_formato_a_xlsx(path)


# ────────────────────────────────────────────────────────────────────────────
# Escanear carpeta
# ────────────────────────────────────────────────────────────────────────────

def obtener_quincena(path: str) -> str:
    """Extrae el nombre de la quincena (subcarpeta de año) de la ruta."""
    rel = os.path.relpath(path, BASE_DIR)
    partes = rel.split(os.sep)
    if len(partes) >= 2:
        return f"{partes[0]}/{partes[1]}"
    return partes[0] if partes else ""


def escanear_carpeta() -> list[dict]:
    todos = []
    archivos = []
    for root, dirs, files in os.walk(BASE_DIR):
        for f in sorted(files):
            if f.startswith('.') or f.endswith('.pdf'):
                continue
            path = os.path.join(root, f)
            ext = f.lower().rsplit('.', 1)[-1]
            if ext in ('xls', 'xlsx'):
                archivos.append(path)

    total = len(archivos)
    print(f"Procesando {total} archivos...")

    for i, path in enumerate(archivos, 1):
        rel = os.path.relpath(path, BASE_DIR)
        quincena = obtener_quincena(path)
        departamento = os.path.splitext(os.path.basename(path))[0]
        print(f"  [{i:3d}/{total}] {rel}", end='\r')

        registros = parsear_archivo(path)
        for r in registros:
            r['quincena'] = quincena
            r['departamento'] = departamento
        todos.extend(registros)

    print(f"\nTotal de registros encontrados: {len(todos):,}")
    return todos


# ────────────────────────────────────────────────────────────────────────────
# Generar Excel de revisión
# ────────────────────────────────────────────────────────────────────────────

def generar_excel(registros: list[dict]):
    print(f"Generando Excel: {OUTPUT_FILE}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Checadas Históricas"

    # ── Encabezados ───────────────────────────────────────────────────────────
    headers = [
        "# Emp", "Nombre", "Fecha", "Día", "Entrada", "Salida",
        "Estado", "Departamento", "Quincena",
    ]
    header_fill = PatternFill("solid", fgColor=COLOR_HEADER)
    header_font = Font(color="FFFFFF", bold=True, size=10)
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # ── Ordenar por empleado → fecha ──────────────────────────────────────────
    registros_ord = sorted(
        registros,
        key=lambda r: (r['numero_empleado'].zfill(6), r['fecha'])
    )

    # ── Contadores ───────────────────────────────────────────────────────────
    n_completo = n_solo_entrada = n_solo_salida = 0

    dias_semana = {0:'Lun',1:'Mar',2:'Mié',3:'Jue',4:'Vie',5:'Sáb',6:'Dom'}

    fills = {
        'COMPLETO':      PatternFill("solid", fgColor=COLOR_COMPLETO),
        'SOLO ENTRADA':  PatternFill("solid", fgColor=COLOR_SOLO_ENTRADA),
        'SOLO SALIDA':   PatternFill("solid", fgColor=COLOR_SOLO_SALIDA),
    }

    for row_idx, r in enumerate(registros_ord, 2):
        entrada = r['entrada']
        salida  = r['salida']

        if entrada and salida:
            estado = 'COMPLETO'
            n_completo += 1
        elif entrada:
            estado = 'SOLO ENTRADA'
            n_solo_entrada += 1
        else:
            estado = 'SOLO SALIDA'
            n_solo_salida += 1

        fill = fills[estado]
        dia = dias_semana.get(r['fecha'].weekday(), '')

        valores = [
            r['numero_empleado'], r['nombre'], r['fecha'], dia,
            entrada or '', salida or '', estado,
            r['departamento'], r['quincena'],
        ]
        for col_idx, val in enumerate(valores, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.font = Font(size=9)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if col_idx == 3 and isinstance(val, date):
                cell.number_format = 'DD/MM/YYYY'
                cell.alignment = Alignment(horizontal="center")
            if col_idx in (5, 6, 4, 7):
                cell.alignment = Alignment(horizontal="center")

        ws.row_dimensions[row_idx].height = 15

    # ── Anchos de columna ─────────────────────────────────────────────────────
    anchos = [8, 32, 12, 5, 8, 8, 14, 20, 28]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    # ── Hoja de resumen ───────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Resumen")
    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 15

    resumen_data = [
        ("RESUMEN DE IMPORTACIÓN", ""),
        ("", ""),
        ("Total registros encontrados",   len(registros_ord)),
        ("Con entrada Y salida",          n_completo),
        ("Solo entrada (falta salida)",   n_solo_entrada),
        ("Solo salida (sin entrada)",     n_solo_salida),
        ("", ""),
        ("Empleados únicos",
         len({r['numero_empleado'] for r in registros_ord})),
        ("Rango de fechas",
         f"{min(r['fecha'] for r in registros_ord).strftime('%d/%m/%Y')} → {max(r['fecha'] for r in registros_ord).strftime('%d/%m/%Y')}"),
        ("", ""),
        ("Leyenda de colores", ""),
        ("Blanco = Entrada y Salida OK", ""),
        ("Amarillo = Solo tiene Entrada", ""),
        ("Rojo claro = Solo tiene Salida", ""),
    ]

    for row_idx, (label, val) in enumerate(resumen_data, 1):
        c1 = ws2.cell(row=row_idx, column=1, value=label)
        c2 = ws2.cell(row=row_idx, column=2, value=val)
        if row_idx == 1:
            c1.font = Font(bold=True, size=12, color=COLOR_HEADER)
        elif label and not label.startswith("Leyenda") and not label.startswith(("Blanco","Amarillo","Rojo")):
            c1.font = Font(bold=True, size=10)
            c2.font = Font(bold=True, size=10, color="1E3A5F")

        # Colores de leyenda
        if 'Blanco' in label:
            c1.fill = PatternFill("solid", fgColor=COLOR_COMPLETO)
        elif 'Amarillo' in label:
            c1.fill = PatternFill("solid", fgColor=COLOR_SOLO_ENTRADA)
        elif 'Rojo' in label:
            c1.fill = PatternFill("solid", fgColor=COLOR_SOLO_SALIDA)

    wb.save(OUTPUT_FILE)
    print(f"\n✓ Excel generado: {OUTPUT_FILE}")
    print(f"\nResumen:")
    print(f"  Completos (entrada+salida): {n_completo:,}")
    print(f"  Solo entrada:               {n_solo_entrada:,}")
    print(f"  Solo salida:                {n_solo_salida:,}")
    print(f"  TOTAL registros:            {len(registros_ord):,}")


# ────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    if not os.path.isdir(BASE_DIR):
        print(f"Error: no se encontró la carpeta:\n  {BASE_DIR}")
        sys.exit(1)

    registros = escanear_carpeta()
    if not registros:
        print("No se encontraron registros. Verifica la ruta.")
        sys.exit(1)

    generar_excel(registros)
    print(f"\nAbre el archivo en Excel y revisa antes de importar.")
    print(f"Cuando estés listo, corre: python3 scripts/importar_desde_excel_historico.py")
