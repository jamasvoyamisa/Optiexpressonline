from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File, Form, Request
from sqlalchemy.orm import Session
from typing import Dict, List, Optional, Tuple
from decimal import Decimal, InvalidOperation
import logging
from app.core.database import get_db
from app.core.config import settings
from app.core.security import get_current_user
from app.core.deps import get_current_empleado_with_rol, require_superuser, require_superuser_or_rh, require_superuser_download
from app.modules.audit.middleware import _client_ip
from app.modules.audit.service import ActividadService
from . import schemas, service, models
from .regimen_fiscal_sat import REGIMENES_FISCALES_SAT

logger = logging.getLogger(__name__)
router = APIRouter(prefix=f"{settings.API_V1_PREFIX}/personal", tags=["Personal"])

# Misma definición que GET /importar/plantilla (orden = columnas en importar/xlsx).
EMPLEADOS_IMPORT_XLSX_COLUMNAS: List[Tuple[str, str, int]] = [
    ("numero_empleado", "No. Empleado *", 14),
    ("nombre", "Nombre *", 18),
    ("apellido_paterno", "Ap. Paterno *", 16),
    ("apellido_materno", "Ap. Materno *", 16),
    ("departamento", "Departamento", 22),
    ("fecha_ingreso", "Fecha Ingreso * (DD/MM/AAAA)", 18),
    ("fecha_nacimiento", "Fecha Nacimiento (DD/MM/AAAA)", 18),
    ("email", "Email", 26),
    ("telefono", "Teléfono", 14),
    ("curp", "CURP", 20),
    ("rfc", "RFC", 15),
    ("nss", "NSS", 13),
    ("direccion", "Dirección", 30),
    ("colonia", "Colonia", 18),
    ("cp", "C.P.", 8),
    ("ciudad", "Ciudad", 16),
    ("contacto_emergencia", "Contacto Emergencia", 22),
    ("telefono_emergencia", "Tel. Emergencia", 16),
    (
        "vacaciones_vigentes",
        "Saldo LFT neto (días reales; por gozar − adeudo; puede ser negativo)",
        36,
    ),
    ("password", "Contraseña", 14),
]


def _depto_to_response(depto: models.Departamento) -> dict:
    data = {
        "id": depto.id, "nombre": depto.nombre, "empresa_id": depto.empresa_id,
        "jefe_id": depto.jefe_id, "padre_id": getattr(depto, "padre_id", None),
        "tipo": getattr(depto, "tipo", None),
        "activo": depto.activo,
        "created_at": depto.created_at, "updated_at": depto.updated_at,
        "empresa": depto.empresa,
        "jefe_nombre": depto.jefe_nombre,
        "padre_nombre": getattr(depto, "padre_nombre", None),
        "encargados_ids": getattr(depto, "encargados_ids", None) or [],
        "encargados_nombres": getattr(depto, "encargados_nombres", None) or [],
    }
    return data


# ========== RUTAS DE EMPRESAS ==========

@router.post("/empresas", response_model=schemas.EmpresaResponse, status_code=status.HTTP_201_CREATED)
def create_empresa(empresa: schemas.EmpresaCreate, db: Session = Depends(get_db), _admin: dict = Depends(require_superuser)):
    existing = db.query(models.Empresa).filter(models.Empresa.nombre == empresa.nombre).first()
    if existing:
        raise HTTPException(status_code=400, detail="Ya existe una empresa con ese nombre")
    try:
        return service.PersonalService.create_empresa(db, empresa)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/empresas", response_model=List[schemas.EmpresaResponse])
def get_empresas(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    activo: Optional[bool] = None,
    db: Session = Depends(get_db),
    _current: dict = Depends(get_current_user),
):
    return service.PersonalService.get_empresas(db, skip=skip, limit=limit, activo=activo)


@router.get("/empresas/{empresa_id}", response_model=schemas.EmpresaResponse)
def get_empresa(empresa_id: int, db: Session = Depends(get_db), _current: dict = Depends(get_current_user)):
    emp = service.PersonalService.get_empresa(db, empresa_id)
    if not emp:
        raise HTTPException(status_code=404, detail="Empresa no encontrada")
    return emp


@router.put("/empresas/{empresa_id}", response_model=schemas.EmpresaResponse)
def update_empresa(empresa_id: int, empresa: schemas.EmpresaUpdate, db: Session = Depends(get_db), _admin: dict = Depends(require_superuser)):
    try:
        emp = service.PersonalService.update_empresa(db, empresa_id, empresa)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    if not emp:
        raise HTTPException(status_code=404, detail="Empresa no encontrada")
    return emp


@router.delete("/empresas/{empresa_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_empresa(empresa_id: int, db: Session = Depends(get_db), _admin: dict = Depends(require_superuser)):
    if not service.PersonalService.delete_empresa(db, empresa_id):
        raise HTTPException(status_code=404, detail="Empresa no encontrada")


@router.get("/regimenes-fiscales-sat")
def get_regimenes_fiscales_sat(_current: dict = Depends(get_current_user)):
    """Catálogo c_RegimenFiscal (SAT) para alta/edición de empresas."""
    return REGIMENES_FISCALES_SAT


# ========== RUTAS DE DEPARTAMENTOS ==========

@router.post("/departamentos", response_model=schemas.DepartamentoResponse, status_code=status.HTTP_201_CREATED)
def create_departamento(depto: schemas.DepartamentoCreate, db: Session = Depends(get_db), _admin: dict = Depends(require_superuser_or_rh)):
    empresa = service.PersonalService.get_empresa(db, depto.empresa_id)
    if not empresa:
        raise HTTPException(status_code=400, detail="La empresa especificada no existe")
    if depto.jefe_id:
        jefe = service.PersonalService.get_empleado(db, depto.jefe_id)
        if not jefe:
            raise HTTPException(status_code=400, detail="El jefe especificado no existe")
    try:
        db_depto = service.PersonalService.create_departamento(db, depto)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    loaded = service.PersonalService.get_departamento(db, db_depto.id)
    return _depto_to_response(loaded)


@router.get("/departamentos", response_model=List[schemas.DepartamentoResponse])
def get_departamentos(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    empresa_id: Optional[int] = None,
    activo: Optional[bool] = None,
    db: Session = Depends(get_db),
    _current: dict = Depends(get_current_user),
):
    deptos = service.PersonalService.get_departamentos(db, skip=skip, limit=limit, empresa_id=empresa_id, activo=activo)
    return [_depto_to_response(d) for d in deptos]


@router.get("/departamentos/{depto_id}", response_model=schemas.DepartamentoResponse)
def get_departamento(depto_id: int, db: Session = Depends(get_db), _current: dict = Depends(get_current_user)):
    depto = service.PersonalService.get_departamento(db, depto_id)
    if not depto:
        raise HTTPException(status_code=404, detail="Departamento no encontrado")
    return _depto_to_response(depto)


@router.put("/departamentos/{depto_id}", response_model=schemas.DepartamentoResponse)
def update_departamento(depto_id: int, depto: schemas.DepartamentoUpdate, db: Session = Depends(get_db), _admin: dict = Depends(require_superuser_or_rh)):
    try:
        updated = service.PersonalService.update_departamento(db, depto_id, depto)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    if not updated:
        raise HTTPException(status_code=404, detail="Departamento no encontrado")
    loaded = service.PersonalService.get_departamento(db, depto_id)
    return _depto_to_response(loaded)


@router.delete("/departamentos/{depto_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_departamento(depto_id: int, db: Session = Depends(get_db), _admin: dict = Depends(require_superuser_or_rh)):
    if not service.PersonalService.delete_departamento(db, depto_id):
        raise HTTPException(status_code=404, detail="Departamento no encontrado")


# ========== RUTAS DE PUESTOS ==========

def _puesto_to_response(p):
    return schemas.PuestoResponse(**service.PersonalService._puesto_to_response(p))


@router.get("/puestos", response_model=List[schemas.PuestoResponse])
def get_puestos(
    activo: Optional[bool] = Query(None, description="true=activos, false=inactivos, omitir=todos"),
    empresa_id: Optional[int] = Query(None, description="Filtrar por empresa"),
    departamento_id: Optional[int] = Query(None, description="Filtrar por departamento"),
    db: Session = Depends(get_db),
    _current: dict = Depends(get_current_user),
):
    """Lista de puestos por empresa y departamento. Los puestos reservados son globales (sin empresa/depto)."""
    service.PersonalService.ensure_puestos_reservados(db)
    puestos = service.PersonalService.get_puestos(db, activo=activo, empresa_id=empresa_id, departamento_id=departamento_id)
    return [_puesto_to_response(p) for p in puestos]


@router.post("/puestos", response_model=schemas.PuestoResponse, status_code=status.HTTP_201_CREATED)
def create_puesto(data: schemas.PuestoCreate, db: Session = Depends(get_db), _admin: dict = Depends(require_superuser_or_rh)):
    """Crear puesto por empresa y departamento. No se pueden crear puestos reservados del sistema."""
    try:
        p = service.PersonalService.create_puesto(db, data)
        p = service.PersonalService.get_puesto(db, p.id)
        return _puesto_to_response(p)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.put("/puestos/{puesto_id}", response_model=schemas.PuestoResponse)
def update_puesto(
    puesto_id: int,
    data: schemas.PuestoUpdate,
    current_extra: dict = Depends(require_superuser_or_rh),
    db: Session = Depends(get_db)
):
    """Actualizar puesto. Los puestos reservados solo los edita el Administrador."""
    p = service.PersonalService.get_puesto(db, puesto_id)
    if not p:
        raise HTTPException(status_code=404, detail="Puesto no encontrado")
    if service.PersonalService._nombre_reservado(p.nombre) and not current_extra.get("is_superuser"):
        raise HTTPException(
            status_code=403,
            detail="Solo el Administrador puede editar los puestos Director, Gerente General, RH, Gerente y Supervisor."
        )
    try:
        updated = service.PersonalService.update_puesto(db, puesto_id, data)
        if not updated:
            raise HTTPException(status_code=404, detail="Puesto no encontrado")
        updated = service.PersonalService.get_puesto(db, puesto_id)
        return _puesto_to_response(updated)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.delete("/puestos/{puesto_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_puesto(puesto_id: int, db: Session = Depends(get_db), _admin: dict = Depends(require_superuser_or_rh)):
    """Eliminar puesto. No se pueden eliminar puestos reservados ni puestos con empleados asignados."""
    try:
        if not service.PersonalService.delete_puesto(db, puesto_id):
            raise HTTPException(status_code=404, detail="Puesto no encontrado")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


# ========== RUTAS DE ROLES ==========

@router.post("/roles", response_model=schemas.RolResponse, status_code=status.HTTP_201_CREATED)
def create_rol(rol: schemas.RolCreate, db: Session = Depends(get_db), _admin: dict = Depends(require_superuser)):
    """Crear nuevo rol"""
    # Verificar si ya existe un rol con ese nombre
    existing = db.query(models.Rol).filter(models.Rol.nombre == rol.nombre).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Ya existe un rol con ese nombre"
        )
    return service.PersonalService.create_rol(db, rol)


@router.get("/roles", response_model=List[schemas.RolResponse])
def get_roles(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    activo: Optional[bool] = None,
    db: Session = Depends(get_db),
    _current: dict = Depends(get_current_user),
):
    """Listar roles"""
    return service.PersonalService.get_roles(db, skip=skip, limit=limit, activo=activo)


@router.get("/roles/{rol_id}", response_model=schemas.RolResponse)
def get_rol(rol_id: int, db: Session = Depends(get_db), _current: dict = Depends(get_current_user)):
    """Obtener rol por ID"""
    db_rol = service.PersonalService.get_rol(db, rol_id)
    if not db_rol:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Rol no encontrado"
        )
    return db_rol


@router.put("/roles/{rol_id}", response_model=schemas.RolResponse)
def update_rol(rol_id: int, rol: schemas.RolUpdate, db: Session = Depends(get_db), _admin: dict = Depends(require_superuser)):
    """Actualizar rol"""
    db_rol = service.PersonalService.update_rol(db, rol_id, rol)
    if not db_rol:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Rol no encontrado"
        )
    return db_rol


@router.delete("/roles/{rol_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_rol(rol_id: int, db: Session = Depends(get_db), _admin: dict = Depends(require_superuser)):
    """Eliminar rol (desactivar)"""
    success = service.PersonalService.delete_rol(db, rol_id)
    if not success:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Rol no encontrado"
        )


# ========== RUTAS DE EMPLEADOS ==========

@router.get("/empleados/next-numero")
def next_numero_empleado(empresa_id: int, db: Session = Depends(get_db), _current: dict = Depends(get_current_user)):
    """Siguiente número de empleado = último numérico registrado en la empresa + 1.
    Solo considera valores totalmente numéricos; el formato (ceros a la izquierda) se
    mantiene acorde al ancho máximo existente."""
    digit_strs = [
        e.numero_empleado.strip()
        for e in db.query(models.Empleado.numero_empleado)
        .filter(models.Empleado.empresa_id == empresa_id)
        .all()
        if e.numero_empleado and str(e.numero_empleado).strip().isdigit()
    ]
    if not digit_strs:
        return {"numero_empleado": "1"}
    numeros = [int(s) for s in digit_strs]
    siguiente = max(numeros) + 1
    sig_str = str(siguiente)
    min_width = max(len(s) for s in digit_strs)
    width = max(min_width, len(sig_str))
    return {"numero_empleado": sig_str.zfill(width)}


@router.get("/empleados/check-username")
def check_username(username: str, exclude_id: int = None, db: Session = Depends(get_db), _current: dict = Depends(get_current_user)):
    """Verificar si un username está disponible. Devuelve {available, suggested}."""
    available = service.PersonalService.check_username_available(db, username, exclude_id)
    return {"available": available, "username": username}


@router.get("/empleados/suggest-username")
def suggest_username(nombre: str, apellido_paterno: str, exclude_id: int = None, db: Session = Depends(get_db), _current: dict = Depends(get_current_user)):
    """Generar un username único a partir de nombre y apellido paterno."""
    username = service.PersonalService.suggest_username(db, nombre, apellido_paterno, exclude_id)
    return {"username": username}


@router.post("/empleados", response_model=schemas.EmpleadoResponse, status_code=status.HTTP_201_CREATED)
def create_empleado(
    empleado: schemas.EmpleadoCreate,
    current_extra: dict = Depends(get_current_empleado_with_rol),
    db: Session = Depends(get_db)
):
    """Crear nuevo empleado y usuario del sistema. Datos personales y laborales son obligatorios."""
    # Solo Administrador puede asignar puestos reservados
    if empleado.puesto_id:
        service.PersonalService.ensure_puestos_reservados(db)
        puesto = service.PersonalService.get_puesto(db, empleado.puesto_id)
        if puesto and service.PersonalService._nombre_reservado(puesto.nombre):
            if not current_extra.get("is_superuser"):
                raise HTTPException(
                    status_code=403,
                    detail="Solo el Administrador puede asignar los puestos Director, Gerente General, RH, Gerente y Supervisor."
                )
    # Validar datos personales obligatorios
    if not (empleado.numero_empleado and empleado.numero_empleado.strip()):
        raise HTTPException(status_code=400, detail="Número de empleado es obligatorio")
    if not (empleado.nombre and empleado.nombre.strip()):
        raise HTTPException(status_code=400, detail="Nombre es obligatorio")
    if not (empleado.apellido_paterno and str(empleado.apellido_paterno or "").strip()):
        raise HTTPException(status_code=400, detail="Apellido paterno es obligatorio")
    if not (empleado.apellido_materno and str(empleado.apellido_materno or "").strip()):
        raise HTTPException(status_code=400, detail="Apellido materno es obligatorio")
    if not empleado.fecha_nacimiento:
        raise HTTPException(status_code=400, detail="Fecha de nacimiento es obligatoria")
    # Validar datos laborales obligatorios
    if not empleado.empresa_id:
        raise HTTPException(status_code=400, detail="Empresa es obligatoria")
    if not empleado.departamento_id:
        raise HTTPException(status_code=400, detail="Departamento es obligatorio")
    if not empleado.puesto_id:
        raise HTTPException(status_code=400, detail="Puesto es obligatorio")
    if not empleado.fecha_ingreso:
        raise HTTPException(status_code=400, detail="Fecha de ingreso es obligatoria")

    # Verificar si ya existe un empleado con ese número en la misma empresa
    existing = db.query(models.Empleado).filter(
        models.Empleado.numero_empleado == empleado.numero_empleado,
        models.Empleado.empresa_id == empleado.empresa_id
    ).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Ya existe un empleado con ese número en esta empresa"
        )

    # Verificar si el rol existe (si se proporciona)
    if empleado.rol_id:
        rol = service.PersonalService.get_rol(db, empleado.rol_id)
        if not rol:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="El rol especificado no existe"
            )
    
    # Verificar si el jefe existe (si se proporciona)
    if empleado.jefe_id:
        jefe = service.PersonalService.get_empleado(db, empleado.jefe_id)
        if not jefe:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="El jefe especificado no existe"
            )
    
    db_empleado = service.PersonalService.create_empleado(db, empleado)

    if empleado.registrar_en_checador and empleado.dispositivo_ids:
        try:
            from app.modules.asistencia import models as asist_models
            nombre_completo = f"{empleado.nombre} {empleado.apellido_paterno or ''} {empleado.apellido_materno or ''}".strip()
            for did in empleado.dispositivo_ids:
                try:
                    # Crear la cola directamente usando el pin_checador del empleado recién creado
                    pendiente = asist_models.UsuarioPendienteDispositivo(
                        dispositivo_id=int(did),
                        numero_empleado=db_empleado.numero_empleado,
                        pin_checador=db_empleado.pin_checador,
                        nombre=nombre_completo,
                        enviado=False,
                    )
                    db.add(pendiente)
                    db.commit()
                except Exception as e:
                    logger.warning(f"Fallo enqueue en dispositivo {did}: {e}")
        except Exception as e:
            logger.warning(f"Empleado creado pero fallo enqueue en checadores: {e}")

    return db_empleado


@router.get("/empleados", response_model=List[schemas.EmpleadoResponse])
def get_empleados(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=5000),
    estado: Optional[str] = None,
    rol_id: Optional[int] = None,
    jefe_id: Optional[int] = None,
    departamento_id: Optional[int] = None,
    search: Optional[str] = None,
    exento_incidencias: Optional[bool] = None,
    incluir_exentos: bool = Query(
        False,
        description="Si true, incluye usuarios especiales (exento incidencias). Por defecto el listado solo muestra empleados operativos.",
    ),
    db: Session = Depends(get_db),
    _current: dict = Depends(get_current_user),
):
    """Listar empleados con filtros. Por defecto excluye usuarios especiales. exento_incidencias=true lista solo especiales."""
    return service.PersonalService.get_empleados(
        db,
        skip=skip,
        limit=limit,
        estado=estado,
        rol_id=rol_id,
        jefe_id=jefe_id,
        departamento_id=departamento_id,
        search=search,
        exento_incidencias=exento_incidencias,
        incluir_exentos=incluir_exentos,
    )


@router.post("/mi-area/ausencias-del-dia", response_model=List[schemas.MiAreaAusenciasDelDiaItem])
def mi_area_ausencias_del_dia(
    body: schemas.MiAreaAusenciasDelDiaRequest,
    ctx: dict = Depends(get_current_empleado_with_rol),
    db: Session = Depends(get_db),
):
    """
    Para la pestaña Personal de Mi área: indica si hoy (calendario México) cada empleado
    está en incapacidad activa o en periodo de vacaciones aprobadas (misma regla que asistencia).
    Solo se devuelven IDs que el usuario puede ver (sus departamentos o superusuario).
    """
    from app.core.timezone_utils import hoy_mexico
    from app.modules.incapacidades import models as inc_models
    from app.modules.asistencia.service import AsistenciaService

    raw_ids = list(dict.fromkeys(body.empleado_ids))[:900]
    if not raw_ids:
        return []

    is_superuser = bool(ctx.get("is_superuser"))
    depto_ids = ctx.get("departamento_ids_que_administro") or []

    if is_superuser:
        ids_ok = [
            row[0]
            for row in db.query(models.Empleado.id).filter(models.Empleado.id.in_(raw_ids)).all()
        ]
    else:
        if not depto_ids:
            return []
        ids_ok = [
            row[0]
            for row in db.query(models.Empleado.id).filter(
                models.Empleado.id.in_(raw_ids),
                models.Empleado.departamento_id.in_(depto_ids),
            ).all()
        ]

    if not ids_ok:
        return []

    hoy = hoy_mexico()
    incap_rows = (
        db.query(inc_models.Incapacidad.empleado_id)
        .filter(
            inc_models.Incapacidad.empleado_id.in_(ids_ok),
            inc_models.Incapacidad.estado == inc_models.EstadoIncapacidad.ACTIVA,
            inc_models.Incapacidad.fecha_inicio <= hoy,
            inc_models.Incapacidad.fecha_fin >= hoy,
        )
        .distinct()
        .all()
    )
    incap_set = {int(r[0]) for r in incap_rows}

    vac_todos = AsistenciaService.empleados_cubiertos_por_solicitud_vacaciones_aprobada(db, hoy)
    ids_ok_set = set(ids_ok)
    vac_set = vac_todos & ids_ok_set

    return [
        schemas.MiAreaAusenciasDelDiaItem(
            empleado_id=eid,
            en_incapacidad=eid in incap_set,
            en_vacaciones=eid in vac_set,
        )
        for eid in sorted(ids_ok_set)
    ]


@router.get("/me", response_model=schemas.EmpleadoResponse)
def get_me(
    current: dict = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Datos del empleado actual (portal del empleado). Requiere autenticación."""
    empleado_id = int(current["user_id"])
    db_empleado = service.PersonalService.get_empleado(db, empleado_id)
    if not db_empleado:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Empleado no encontrado"
        )
    return db_empleado


@router.get("/empleados/{empleado_id}", response_model=schemas.EmpleadoResponse)
def get_empleado(empleado_id: int, db: Session = Depends(get_db), _current: dict = Depends(get_current_user)):
    """Obtener empleado por ID"""
    db_empleado = service.PersonalService.get_empleado(db, empleado_id)
    if not db_empleado:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Empleado no encontrado"
        )
    return db_empleado


@router.put("/empleados/{empleado_id}", response_model=schemas.EmpleadoResponse)
def update_empleado(
    empleado_id: int,
    empleado: schemas.EmpleadoUpdate,
    current_extra: dict = Depends(get_current_empleado_with_rol),
    db: Session = Depends(get_db)
):
    """Actualizar empleado"""
    # Solo Administrador puede asignar puestos reservados
    if empleado.puesto_id is not None:
        service.PersonalService.ensure_puestos_reservados(db)
        puesto = service.PersonalService.get_puesto(db, empleado.puesto_id)
        if puesto and service.PersonalService._nombre_reservado(puesto.nombre):
            if not current_extra.get("is_superuser"):
                raise HTTPException(
                    status_code=403,
                    detail="Solo el Administrador puede asignar los puestos Director, Gerente General, RH, Gerente y Supervisor."
                )
    try:
        db_empleado = service.PersonalService.update_empleado(db, empleado_id, empleado)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))
    if not db_empleado:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Empleado no encontrado"
        )
    return db_empleado


@router.post("/empleados/{empleado_id}/restablecer-password")
def restablecer_password_empleado(
    empleado_id: int,
    request: Request,
    current_extra: dict = Depends(get_current_empleado_with_rol),
    db: Session = Depends(get_db),
):
    """
    RH o Admin genera una contraseña temporal aleatoria.
    El colaborador debe cambiarla al entrar (must_change_password).
    La clave se muestra una sola vez en la respuesta.
    """
    if not (current_extra.get("is_superuser") or current_extra.get("is_rh")):
        raise HTTPException(
            status_code=403,
            detail="Solo Administrador o RH pueden restablecer contraseñas temporales.",
        )
    temporal = service.PersonalService.restablecer_password_temporal(db, empleado_id)
    if temporal is None:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")
    actor_id = int(current_extra["user_id"])
    afectado = (
        db.query(models.Empleado)
        .filter(models.Empleado.id == empleado_id)
        .first()
    )
    num_afectado = (afectado.numero_empleado if afectado else None) or str(empleado_id)
    nombre_afectado = (
        " ".join(
            p for p in [
                afectado.nombre if afectado else None,
                afectado.apellido_paterno if afectado else None,
                afectado.apellido_materno if afectado else None,
            ] if p and str(p).strip()
        ) or None
    )
    empresa_afectado = (afectado.empresa.nombre if afectado and afectado.empresa else None)
    ActividadService.registrar(
        db,
        nivel="info",
        categoria="auth",
        mensaje=f"Contraseña temporal restablecida para No. empleado {num_afectado}",
        empleado_id=actor_id,
        ip_cliente=_client_ip(request) or None,
        metodo_http="POST",
        ruta=(f"{settings.API_V1_PREFIX}/personal/empleados/{empleado_id}/restablecer-password")[:500],
        codigo_http=200,
        contexto={
            "empleado_afectado_id": empleado_id,
            "empleado_afectado_numero": num_afectado,
            "empleado_afectado_nombre": nombre_afectado,
            "empleado_afectado_empresa": empresa_afectado,
            "accion": "restablecer_password_temporal",
            "actor_rol": "admin" if current_extra.get("is_superuser") else "rh",
        },
    )
    return {
        "empleado_id": empleado_id,
        "password_temporal": temporal,
        "mensaje": "Contraseña temporal generada. Entrégala al colaborador; deberá cambiarla al iniciar sesión.",
    }


@router.patch("/empleados/{empleado_id}/permisos-especiales", response_model=schemas.EmpleadoResponse)
def set_permisos_especiales(
    empleado_id: int,
    body: schemas.PermisosEspecialesUpdate,
    current_extra: dict = Depends(get_current_empleado_with_rol),
    db: Session = Depends(get_db),
):
    """
    Actualiza los permisos especiales de un empleado (exento_incidencias, puede_checar_remoto).
    Solo el Administrador (superuser) puede usar este endpoint.
    """
    if not current_extra.get("is_superuser"):
        raise HTTPException(status_code=403, detail="Solo el Administrador puede gestionar permisos especiales.")
    emp = service.PersonalService.get_empleado(db, empleado_id)
    if not emp:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")
    if body.exento_incidencias is not None:
        emp.exento_incidencias = body.exento_incidencias
    if body.puede_checar_remoto is not None:
        emp.puede_checar_remoto = body.puede_checar_remoto
    db.commit()
    db.refresh(emp)
    return emp


@router.post("/usuarios-especiales", response_model=schemas.EmpleadoResponse, status_code=status.HTTP_201_CREATED)
def create_usuario_especial(
    data: schemas.UsuarioEspecialCreate,
    current_extra: dict = Depends(get_current_empleado_with_rol),
    db: Session = Depends(get_db),
):
    """
    Alta simplificada de usuario especial:
    - No se captura número de empleado (se genera interno).
    - Siempre queda exento de incidencias.
    - Puede asignarse a cualquier empresa/departamento/puesto.
    Solo Administrador.
    """
    if not current_extra.get("is_superuser"):
        raise HTTPException(status_code=403, detail="Solo el Administrador puede crear usuarios especiales.")

    if not (data.nombre and data.nombre.strip()):
        raise HTTPException(status_code=400, detail="Nombre es obligatorio")

    empresa = service.PersonalService.get_empresa(db, data.empresa_id)
    if not empresa:
        raise HTTPException(status_code=400, detail="La empresa especificada no existe")

    depto = service.PersonalService.get_departamento(db, data.departamento_id)
    if not depto:
        raise HTTPException(status_code=400, detail="El departamento especificado no existe")
    if depto.empresa_id != data.empresa_id:
        raise HTTPException(status_code=400, detail="El departamento no pertenece a la empresa seleccionada")

    puesto = service.PersonalService.get_puesto(db, data.puesto_id)
    if not puesto:
        raise HTTPException(status_code=400, detail="El puesto especificado no existe")
    if puesto.empresa_id is not None and puesto.empresa_id != data.empresa_id:
        raise HTTPException(status_code=400, detail="El puesto no pertenece a la empresa seleccionada")
    if puesto.departamento_id is not None and puesto.departamento_id != data.departamento_id:
        raise HTTPException(status_code=400, detail="El puesto no pertenece al departamento seleccionado")

    if (puesto.nombre or "").strip().lower() in ("director", "subdirector", "gerente general"):
        ids = set(data.empresas_supervision_ids or [data.empresa_id])
        ids.add(data.empresa_id)
        for eid in ids:
            if not service.PersonalService.get_empresa(db, eid):
                raise HTTPException(status_code=400, detail=f"La empresa {eid} no existe")

    try:
        return service.PersonalService.create_usuario_especial(db, data)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.delete("/empleados/{empleado_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_empleado(empleado_id: int, db: Session = Depends(get_db), _admin: dict = Depends(require_superuser_or_rh)):
    """Eliminar empleado (dar de baja)"""
    success = service.PersonalService.delete_empleado(db, empleado_id)
    if not success:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Empleado no encontrado"
        )


@router.get("/empleados/{jefe_id}/subordinados", response_model=List[schemas.EmpleadoResponse])
def get_subordinados(jefe_id: int, db: Session = Depends(get_db), _current: dict = Depends(get_current_user)):
    """Obtener subordinados de un jefe"""
    jefe = service.PersonalService.get_empleado(db, jefe_id)
    if not jefe:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Jefe no encontrado"
        )
    return service.PersonalService.get_subordinados(db, jefe_id)


# ========== IMPORTACIÓN MASIVA DESDE XLSX ==========

@router.get("/importar/plantilla")
def descargar_plantilla_xlsx(
    download_token: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    _su: dict = Depends(require_superuser_download),
):
    """Descarga una plantilla XLSX con las columnas necesarias y catálogos en hojas auxiliares."""
    from io import BytesIO
    from fastapi.responses import Response
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    empresas = db.query(models.Empresa).filter(models.Empresa.activo == True).order_by(models.Empresa.nombre).all()
    deptos = db.query(models.Departamento).filter(models.Departamento.activo == True).order_by(models.Departamento.nombre).all()

    wb = Workbook()

    thin = Side(style="thin", color="B0B0B0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    hdr_fill = PatternFill(start_color="2B6CB0", end_color="2B6CB0", fill_type="solid")
    opt_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    align_c = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── Hoja principal: Empleados ──
    ws = wb.active
    ws.title = "Empleados"

    for ci, (key, label, width) in enumerate(EMPLEADOS_IMPORT_XLSX_COLUMNAS, 1):
        cell = ws.cell(row=1, column=ci, value=label)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = align_c
        cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = width
        # Columnas opcionales con fondo gris
        if "*" not in label:
            for r in range(2, 102):
                ws.cell(row=r, column=ci).fill = opt_fill

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    # ── Hoja catálogo: Empresas ──
    ws_emp = wb.create_sheet("Cat. Empresas")
    ws_emp.cell(row=1, column=1, value="Empresa").font = hdr_font
    ws_emp.cell(row=1, column=1).fill = hdr_fill
    ws_emp.column_dimensions["A"].width = 30
    for i, e in enumerate(empresas, 2):
        ws_emp.cell(row=i, column=1, value=e.nombre)

    # ── Hoja catálogo: Departamentos ──
    ws_dep = wb.create_sheet("Cat. Departamentos")
    ws_dep.cell(row=1, column=1, value="Departamento").font = hdr_font
    ws_dep.cell(row=1, column=1).fill = hdr_fill
    ws_dep.cell(row=1, column=2, value="Empresa").font = hdr_font
    ws_dep.cell(row=1, column=2).fill = hdr_fill
    ws_dep.column_dimensions["A"].width = 30
    ws_dep.column_dimensions["B"].width = 30
    emp_map = {e.id: e.nombre for e in empresas}
    for i, d in enumerate(deptos, 2):
        ws_dep.cell(row=i, column=1, value=d.nombre)
        ws_dep.cell(row=i, column=2, value=emp_map.get(d.empresa_id, ""))

    # Validaciones desplegable para Depto
    if deptos:
        dv_dep = DataValidation(type="list", formula1=f"'Cat. Departamentos'!$A$2:$A${len(deptos)+1}", allow_blank=False)
        ws.add_data_validation(dv_dep)
        dv_dep.add(f"E2:E1000")

    output = BytesIO()
    wb.save(output)
    contenido = output.getvalue()
    output.close()
    wb.close()
    return Response(
        content=contenido,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="plantilla_empleados.xlsx"',
            "Content-Length": str(len(contenido)),
            "Cache-Control": "no-store",
        },
    )


@router.post("/importar/xlsx")
async def importar_empleados_xlsx(
    file: UploadFile = File(...),
    empresa_id: int = Form(...),
    actualizar_existentes: bool = Form(False),
    db: Session = Depends(get_db),
    _su: dict = Depends(require_superuser),
):
    """
    Importa empleados desde un archivo XLSX.
    La empresa se selecciona en el modal y se aplica a todas las filas.
    Busca departamento por nombre dentro de esa empresa.
    Si el empleado (numero_empleado + empresa) ya existe:
    - por defecto, lo omite.
    - si actualizar_existentes=true, completa campos vacíos con valores del XLSX.
    Devuelve resumen de creados, actualizados, omitidos y errores.
    """
    from io import BytesIO
    from openpyxl import load_workbook
    from datetime import datetime, date
    from app.core.security import get_password_hash

    if not file.filename or not file.filename.endswith(('.xlsx', '.XLSX')):
        raise HTTPException(status_code=400, detail="El archivo debe ser .xlsx")

    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Archivo demasiado grande (máx 10 MB)")

    try:
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="No se pudo leer el archivo XLSX")

    ws = wb.active
    if ws is None:
        raise HTTPException(status_code=400, detail="El archivo no tiene hojas")

    # Leer encabezados (fila 1)
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value or "").strip().lower())

    COL_MAP = {
        "no. empleado": "numero_empleado", "numero_empleado": "numero_empleado", "no empleado": "numero_empleado",
        "nombre": "nombre",
        "ap. paterno": "apellido_paterno", "apellido_paterno": "apellido_paterno", "ap paterno": "apellido_paterno",
        "ap. materno": "apellido_materno", "apellido_materno": "apellido_materno", "ap materno": "apellido_materno",
        "departamento": "departamento",
        "fecha ingreso": "fecha_ingreso", "fecha_ingreso": "fecha_ingreso",
        "fecha nacimiento": "fecha_nacimiento", "fecha_nacimiento": "fecha_nacimiento",
        "email": "email", "correo": "email",
        "telefono": "telefono", "teléfono": "telefono",
        "curp": "curp", "rfc": "rfc", "nss": "nss",
        "direccion": "direccion", "dirección": "direccion",
        "colonia": "colonia", "c.p.": "cp", "cp": "cp",
        "ciudad": "ciudad",
        "contacto emergencia": "contacto_emergencia", "contacto_emergencia": "contacto_emergencia",
        "tel. emergencia": "telefono_emergencia", "telefono_emergencia": "telefono_emergencia",
        "vacaciones vigentes": "vacaciones_vigentes",
        "vacaciones vigentes (días lft por gozar)": "vacaciones_vigentes",
        "saldo lft neto (días reales; por gozar − adeudo; puede ser negativo)": "vacaciones_vigentes",
        "saldo lft neto (dias reales; por gozar - adeudo; puede ser negativo)": "vacaciones_vigentes",
        "vacaciones_vigentes": "vacaciones_vigentes",
        "contraseña": "password", "password": "password", "contraseña": "password",
    }

    col_idx = {}
    for i, h in enumerate(headers):
        clean = h.replace("*", "").replace("(dd/mm/aaaa)", "").strip()
        mapped = COL_MAP.get(clean)
        if mapped:
            col_idx[mapped] = i

    required = ["numero_empleado", "nombre", "apellido_paterno", "apellido_materno", "fecha_ingreso"]
    missing = [r for r in required if r not in col_idx]
    if missing:
        raise HTTPException(status_code=400, detail=f"Columnas faltantes: {', '.join(missing)}")

    # Cachear catálogos
    empresa = db.query(models.Empresa).filter(models.Empresa.id == empresa_id, models.Empresa.activo == True).first()
    if not empresa:
        raise HTTPException(status_code=400, detail="Empresa seleccionada inválida o inactiva")

    deptos_all = db.query(models.Departamento).filter(models.Departamento.activo == True).all()
    deptos_map = {}
    for d in deptos_all:
        deptos_map[(d.nombre.strip().lower(), d.empresa_id)] = d.id

    # Rol "empleado" por defecto
    rol_empleado = db.query(models.Rol).filter(models.Rol.nombre.ilike("%empleado%")).first()
    rol_id_default = rol_empleado.id if rol_empleado else None

    def get_cell(row_data, field):
        idx = col_idx.get(field)
        if idx is None or idx >= len(row_data):
            return None
        v = row_data[idx]
        if v is None:
            return None
        return str(v).strip() if not isinstance(v, (datetime, date)) else v

    def parse_date(val):
        if val is None:
            return None
        if isinstance(val, datetime):
            return val
        if isinstance(val, date):
            return datetime(val.year, val.month, val.day)
        s = str(val).strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue
        return None

    def parse_decimal(val):
        if val is None:
            return None
        if isinstance(val, (int, float, Decimal)):
            try:
                return Decimal(str(val))
            except InvalidOperation:
                return None
        s = str(val).strip().replace(",", ".")
        if s == "":
            return None
        try:
            return Decimal(s)
        except InvalidOperation:
            return None

    def _aplicar_saldo_lft_neto_desde_import(
        empleado_id: int, saldo_neto_objetivo: Decimal, *, do_commit: bool = True
    ):
        from app.modules.vacaciones import service as vacaciones_service

        vacaciones_service.VacacionesService.aplicar_saldo_lft_neto_import(
            db, empleado_id, saldo_neto_objetivo, do_commit=do_commit
        )

    creados = []
    actualizados = []
    omitidos = []
    errores = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_data = list(row)

        num_emp = get_cell(row_data, "numero_empleado")
        nombre = get_cell(row_data, "nombre")
        if not num_emp or not nombre:
            continue  # fila vacía

        ap_pat = get_cell(row_data, "apellido_paterno") or ""
        ap_mat = get_cell(row_data, "apellido_materno") or ""
        depto_name = get_cell(row_data, "departamento") or ""

        depto_id = None
        if depto_name:
            depto_id = deptos_map.get((depto_name.lower(), empresa_id))
            if not depto_id:
                errores.append({"fila": row_num, "error": f"Departamento '{depto_name}' no encontrado en empresa seleccionada"})
                continue

        fecha_ingreso = parse_date(get_cell(row_data, "fecha_ingreso"))
        if not fecha_ingreso:
            errores.append({"fila": row_num, "error": "Fecha de ingreso inválida o vacía"})
            continue

        # Ya existe?
        existing = db.query(models.Empleado).filter(
            models.Empleado.numero_empleado == str(num_emp),
            models.Empleado.empresa_id == empresa_id,
        ).first()
        if existing:
            if not actualizar_existentes:
                omitidos.append({"fila": row_num, "numero_empleado": str(num_emp), "nombre": nombre, "razon": "Ya existe"})
                continue

            cambios = []

            def _vac(v):
                if v is None:
                    return True
                if isinstance(v, str):
                    return v.strip() == ""
                return False

            def _set_if_empty(field, new_value):
                if _vac(new_value):
                    return
                old_value = getattr(existing, field, None)
                if _vac(old_value):
                    setattr(existing, field, new_value)
                    cambios.append(field)

            fecha_nac = parse_date(get_cell(row_data, "fecha_nacimiento"))
            email_val = get_cell(row_data, "email") or None
            if email_val and _vac(existing.email):
                exists_email = db.query(models.Empleado).filter(
                    models.Empleado.email == email_val,
                    models.Empleado.id != existing.id,
                ).first()
                if not exists_email:
                    _set_if_empty("email", email_val)

            _set_if_empty("nombre", nombre)
            _set_if_empty("apellido_paterno", ap_pat)
            _set_if_empty("apellido_materno", ap_mat)
            _set_if_empty("departamento_id", depto_id)
            _set_if_empty("fecha_ingreso", fecha_ingreso)
            _set_if_empty("fecha_nacimiento", fecha_nac)
            _set_if_empty("telefono", get_cell(row_data, "telefono"))
            _set_if_empty("curp", get_cell(row_data, "curp"))
            _set_if_empty("rfc", get_cell(row_data, "rfc"))
            _set_if_empty("nss", get_cell(row_data, "nss"))
            _set_if_empty("direccion", get_cell(row_data, "direccion"))
            _set_if_empty("colonia", get_cell(row_data, "colonia"))
            _set_if_empty("cp", get_cell(row_data, "cp"))
            _set_if_empty("ciudad", get_cell(row_data, "ciudad"))
            _set_if_empty("contacto_emergencia", get_cell(row_data, "contacto_emergencia"))
            _set_if_empty("telefono_emergencia", get_cell(row_data, "telefono_emergencia"))
            _set_if_empty("rol_id", rol_id_default)

            vacaciones_vigentes_val = parse_decimal(get_cell(row_data, "vacaciones_vigentes"))
            if vacaciones_vigentes_val is not None:
                try:
                    _aplicar_saldo_lft_neto_desde_import(
                        existing.id, vacaciones_vigentes_val, do_commit=False
                    )
                    cambios.append("vacaciones_vigentes")
                except ValueError as e:
                    errores.append({"fila": row_num, "error": str(e)[:200]})
                    continue

            password_raw = get_cell(row_data, "password")
            if password_raw and _vac(existing.password_hash):
                existing.password_hash = get_password_hash(password_raw)
                existing.must_change_password = True
                cambios.append("password_hash")

            if cambios:
                try:
                    db.commit()
                    actualizados.append({
                        "fila": row_num,
                        "id": existing.id,
                        "numero_empleado": str(num_emp),
                        "nombre": f"{existing.nombre or ''} {existing.apellido_paterno or ''}".strip(),
                        "campos": cambios,
                    })
                except Exception as e:
                    db.rollback()
                    errores.append({"fila": row_num, "error": str(e)[:200]})
            else:
                omitidos.append({
                    "fila": row_num,
                    "numero_empleado": str(num_emp),
                    "nombre": nombre,
                    "razon": "Ya existe y no había campos vacíos por completar",
                })
            continue

        fecha_nac = parse_date(get_cell(row_data, "fecha_nacimiento"))
        email_val = get_cell(row_data, "email") or None
        vacaciones_vigentes_val = parse_decimal(get_cell(row_data, "vacaciones_vigentes"))
        if email_val:
            exists_email = db.query(models.Empleado).filter(models.Empleado.email == email_val).first()
            if exists_email:
                email_val = None  # evitar duplicado, se deja sin email

        password_raw = get_cell(row_data, "password")
        rfc_val = get_cell(row_data, "rfc") or ""

        try:
            emp_data = schemas.EmpleadoCreate(
                numero_empleado=str(num_emp),
                nombre=nombre,
                apellido_paterno=ap_pat,
                apellido_materno=ap_mat,
                empresa_id=empresa_id,
                departamento_id=depto_id,
                puesto_id=None,
                fecha_ingreso=fecha_ingreso,
                fecha_nacimiento=fecha_nac,
                email=email_val,
                telefono=get_cell(row_data, "telefono"),
                curp=get_cell(row_data, "curp"),
                rfc=rfc_val,
                nss=get_cell(row_data, "nss"),
                direccion=get_cell(row_data, "direccion"),
                colonia=get_cell(row_data, "colonia"),
                cp=get_cell(row_data, "cp"),
                ciudad=get_cell(row_data, "ciudad"),
                contacto_emergencia=get_cell(row_data, "contacto_emergencia"),
                telefono_emergencia=get_cell(row_data, "telefono_emergencia"),
                rol_id=rol_id_default,
                password=password_raw,
                estado=models.EstadoEmpleado.ACTIVO,
            )
            db_emp = service.PersonalService.create_empleado(db, emp_data)
            if vacaciones_vigentes_val is not None:
                _aplicar_saldo_lft_neto_desde_import(db_emp.id, vacaciones_vigentes_val)
            creados.append({"fila": row_num, "id": db_emp.id, "numero_empleado": str(num_emp), "nombre": f"{nombre} {ap_pat}"})
        except Exception as e:
            db.rollback()
            errores.append({"fila": row_num, "error": str(e)[:200]})

    wb.close()

    return {
        "total_filas": len(creados) + len(actualizados) + len(omitidos) + len(errores),
        "creados": len(creados),
        "actualizados": len(actualizados),
        "omitidos": len(omitidos),
        "errores_count": len(errores),
        "detalle_creados": creados,
        "detalle_actualizados": actualizados,
        "detalle_omitidos": omitidos,
        "detalle_errores": errores,
    }


@router.get("/exportar/empleados")
def exportar_empleados_xlsx(
    empresa_id: Optional[int] = Query(None),
    estado: Optional[str] = Query(None, description="activo, inactivo o baja"),
    download_token: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    _auth: dict = Depends(require_superuser_download),
):
    """Exporta empleados a XLSX: mismas columnas que la plantilla de importación (GET /importar/plantilla).

    Una hoja por empresa (o «Sin datos»). Incluye saldo LFT neto (igual que Mi Vacaciones) y hojas
    Catálogo + validación de departamento como la plantilla, para reimportar sin reformatear.
    """
    from io import BytesIO
    from datetime import datetime, date

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from sqlalchemy.orm import joinedload

    def fmt_dt(val):
        if val is None:
            return ""
        if isinstance(val, datetime):
            return val.strftime("%d/%m/%Y")
        if isinstance(val, date):
            return val.strftime("%d/%m/%Y")
        return str(val)

    q = (
        db.query(models.Empleado)
        .options(
            joinedload(models.Empleado.empresa),
            joinedload(models.Empleado.departamento_rel),
        )
        .filter(models.Empleado.empresa_id.isnot(None))
    )
    if empresa_id is not None:
        q = q.filter(models.Empleado.empresa_id == empresa_id)
    from sqlalchemy import or_, func

    if estado:
        try:
            est = models.EstadoEmpleado(estado.lower())
        except ValueError:
            raise HTTPException(status_code=400, detail="Estado inválido (use activo, inactivo o baja)")
        if est == models.EstadoEmpleado.ACTIVO:
            q = q.filter(
                or_(
                    models.Empleado.estado == models.EstadoEmpleado.ACTIVO,
                    models.Empleado.estado.is_(None),
                )
            )
        else:
            q = q.filter(models.Empleado.estado == est)

    q = q.filter(func.coalesce(models.Empleado.exento_incidencias, False) == False)

    rows = q.order_by(
        models.Empleado.empresa_id,
        models.Empleado.apellido_paterno,
        models.Empleado.apellido_materno,
        models.Empleado.nombre,
    ).limit(25000).all()

    from app.modules.vacaciones.service import VacacionesService

    saldo_map = VacacionesService.saldo_lft_neto_por_empleados(db, [e.id for e in rows])

    empresas_cat = (
        db.query(models.Empresa).filter(models.Empresa.activo == True).order_by(models.Empresa.nombre).all()
    )
    deptos_cat = (
        db.query(models.Departamento)
        .filter(models.Departamento.activo == True)
        .order_by(models.Departamento.nombre)
        .all()
    )
    emp_map_cat = {e.id: e.nombre for e in empresas_cat}

    wb = Workbook()
    wb.remove(wb.active)

    thin = Side(style="thin", color="B0B0B0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    hdr_fill = PatternFill(start_color="2B6CB0", end_color="2B6CB0", fill_type="solid")
    opt_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    align_c = Alignment(horizontal="center", vertical="center", wrap_text=True)

    dept_col_idx = next(
        i for i, (k, _, _) in enumerate(EMPLEADOS_IMPORT_XLSX_COLUMNAS, 1) if k == "departamento"
    )
    dept_col_letter = get_column_letter(dept_col_idx)

    def fila_valores_import(emp: models.Empleado) -> list:
        depto_n = emp.departamento_rel.nombre if emp.departamento_rel else ""
        raw_saldo = saldo_map.get(emp.id)
        if raw_saldo is None:
            vac_val = ""
        elif abs(float(raw_saldo)) < 0.005:
            vac_val = 0
        else:
            vac_val = round(float(raw_saldo), 2)
        return [
            emp.numero_empleado or "",
            emp.nombre or "",
            emp.apellido_paterno or "",
            emp.apellido_materno or "",
            depto_n,
            fmt_dt(emp.fecha_ingreso),
            fmt_dt(emp.fecha_nacimiento),
            emp.email or "",
            emp.telefono or "",
            emp.curp or "",
            emp.rfc or "",
            emp.nss or "",
            emp.direccion or "",
            emp.colonia or "",
            emp.cp or "",
            emp.ciudad or "",
            emp.contacto_emergencia or "",
            emp.telefono_emergencia or "",
            vac_val,
            "",
        ]

    def escribir_hoja_empleados(ws, empleados_list: List[models.Empleado]):
        for ci, (_, label, width) in enumerate(EMPLEADOS_IMPORT_XLSX_COLUMNAS, 1):
            cell = ws.cell(row=1, column=ci, value=label)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = align_c
            cell.border = border
            ws.column_dimensions[get_column_letter(ci)].width = width
            if "*" not in label:
                for r in range(2, 102):
                    ws.cell(row=r, column=ci).fill = opt_fill
        ws.row_dimensions[1].height = 32
        ws.freeze_panes = "A2"
        for ri, emp in enumerate(empleados_list, 2):
            for ci, val in enumerate(fila_valores_import(emp), 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.border = border
        if deptos_cat:
            dv_dep = DataValidation(
                type="list",
                formula1=f"'Cat. Departamentos'!$A$2:$A${len(deptos_cat) + 1}",
                allow_blank=True,
            )
            ws.add_data_validation(dv_dep)
            dv_dep.add(f"{dept_col_letter}2:{dept_col_letter}1000")

    def safe_sheet_name(name: str, fallback: str = "Empresa") -> str:
        base = (name or fallback).strip()
        invalid = set('[]:*?/\\')
        cleaned = "".join(c for c in base if c not in invalid).strip()
        if not cleaned:
            cleaned = fallback
        return cleaned[:31]

    by_empresa: Dict[str, List[models.Empleado]] = {}
    for emp in rows:
        emp_name = emp.empresa.nombre if emp.empresa else "Sin empresa"
        by_empresa.setdefault(emp_name, []).append(emp)

    if not by_empresa:
        ws = wb.create_sheet(title="Sin datos")
        escribir_hoja_empleados(ws, [])
    else:
        used_names: set[str] = set()
        for empresa_nombre in sorted(by_empresa.keys()):
            base_name = safe_sheet_name(empresa_nombre, fallback="Empresa")
            sheet_name = base_name
            seq = 2
            while sheet_name.lower() in used_names:
                suffix = f" ({seq})"
                max_base_len = 31 - len(suffix)
                sheet_name = f"{base_name[:max_base_len]}{suffix}"
                seq += 1
            used_names.add(sheet_name.lower())
            ws = wb.create_sheet(title=sheet_name)
            escribir_hoja_empleados(ws, by_empresa[empresa_nombre])

    ws_emp = wb.create_sheet("Cat. Empresas")
    ws_emp.cell(row=1, column=1, value="Empresa").font = hdr_font
    ws_emp.cell(row=1, column=1).fill = hdr_fill
    ws_emp.column_dimensions["A"].width = 30
    for i, e in enumerate(empresas_cat, 2):
        ws_emp.cell(row=i, column=1, value=e.nombre)

    ws_dep = wb.create_sheet("Cat. Departamentos")
    ws_dep.cell(row=1, column=1, value="Departamento").font = hdr_font
    ws_dep.cell(row=1, column=1).fill = hdr_fill
    ws_dep.cell(row=1, column=2, value="Empresa").font = hdr_font
    ws_dep.cell(row=1, column=2).fill = hdr_fill
    ws_dep.column_dimensions["A"].width = 30
    ws_dep.column_dimensions["B"].width = 30
    for i, d in enumerate(deptos_cat, 2):
        ws_dep.cell(row=i, column=1, value=d.nombre)
        ws_dep.cell(row=i, column=2, value=emp_map_cat.get(d.empresa_id, ""))

    output = BytesIO()
    wb.save(output)
    contenido = output.getvalue()
    output.close()
    wb.close()
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    fname = f"empleados_export_{ts}.xlsx"
    from fastapi.responses import Response
    return Response(
        content=contenido,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{fname}"',
            "Content-Length": str(len(contenido)),
            "Cache-Control": "no-store",
        },
    )
