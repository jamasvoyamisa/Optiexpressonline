"""Exportación XLSX de detalle de nómina por periodo, una hoja por área."""
from __future__ import annotations

import io
import re
from decimal import Decimal
from typing import List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.modules.personal.models import Empleado, Empresa

from .models import DetalleNominaEmpleado, PeriodoEstado, PeriodoNomina
from .nomina_areas import (
    agrupar_detalles_por_area,
    cargar_detalles_periodo,
    departamento_de_empleado,
    listar_areas_periodo,
    slug_area,
)
from .numero_periodo import etiqueta_quincena_numero, numero_periodo_nomina

# ── Estilos ───────────────────────────────────────────────────────────────────

FILL_HEADER = PatternFill("solid", fgColor="1E40AF")
FILL_TITLE = PatternFill("solid", fgColor="EFF6FF")
FILL_SUBTOTAL = PatternFill("solid", fgColor="F3F4F6")
FILL_TOTAL = PatternFill("solid", fgColor="DBEAFE")
FONT_HEADER = Font(bold=True, color="FFFFFF", size=10)
FONT_TITLE = Font(bold=True, size=13, color="1E3A8A")
FONT_SUBTITLE = Font(size=10, color="374151")
FONT_TOTAL = Font(bold=True, size=10, color="1E3A8A")
MONEY_FMT = "#,##0.00"
THIN = Side(style="thin", color="D1D5DB")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADERS = [
    "Quincena",
    "Área",
    "No. empleado",
    "Empleado",
    "Días lab.",
    "Días pag.",
    "Fuente días",
    "Percepciones",
    "Gravado",
    "Exento",
    "Deducciones",
    "Subsidio",
    "Neto",
    "UUID CFDI",
]

RESUMEN_HEADERS = ["Área", "Recibos", "Percepciones", "Deducciones", "Neto"]


def _dec(v) -> Decimal:
    if v is None:
        return Decimal("0")
    return Decimal(str(v))


def _float(v: Decimal) -> float:
    return float(v.quantize(Decimal("0.01")))


def _totales_detalles(detalles: List[DetalleNominaEmpleado]) -> Tuple[int, Decimal, Decimal, Decimal]:
    perc = ded = neto = Decimal("0")
    for d in detalles:
        perc += _dec(d.total_percepciones)
        ded += _dec(d.total_deducciones)
        neto += _dec(d.total_neto)
    return len(detalles), perc, ded, neto


def _empleado_nombre(emp: Empleado | None) -> str:
    if not emp:
        return ""
    parts = [emp.nombre or "", emp.apellido_paterno or "", emp.apellido_materno or ""]
    return " ".join(p.strip() for p in parts if p and p.strip())


def _fila_detalle(quincena: str, area: str, d: DetalleNominaEmpleado) -> list:
    emp = d.empleado
    return [
        quincena,
        area,
        emp.numero_empleado if emp else "",
        _empleado_nombre(emp),
        _float(_dec(d.dias_laborados)) if d.dias_laborados is not None else None,
        _float(_dec(d.dias_pagados)) if d.dias_pagados is not None else None,
        d.dias_fuente or "",
        _float(_dec(d.total_percepciones)) if d.total_percepciones is not None else None,
        _float(_dec(d.total_gravado)) if d.total_gravado is not None else None,
        _float(_dec(d.total_exento)) if d.total_exento is not None else None,
        _float(_dec(d.total_deducciones)) if d.total_deducciones is not None else None,
        _float(_dec(d.subsidio_causado)) if d.subsidio_causado is not None else None,
        _float(_dec(d.total_neto)) if d.total_neto is not None else None,
        d.cfdi_uuid or "",
    ]


def _aplicar_borde_celda(cell, border: Border = BORDER_ALL) -> None:
    cell.border = border


def _celda_money(ws, row: int, col: int, value: float | None, bold: bool = False) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    if isinstance(value, (int, float)):
        cell.number_format = MONEY_FMT
    if bold:
        cell.font = FONT_TOTAL
    _aplicar_borde_celda(cell)


def _ajustar_columnas(ws, min_col: int, max_col: int, min_row: int = 1) -> None:
    for col in range(min_col, max_col + 1):
        letter = get_column_letter(col)
        max_len = 10
        for row in ws.iter_rows(min_row=min_row, min_col=col, max_col=col):
            val = row[0].value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[letter].width = min(max_len + 3, 42)


def _fmt_fecha_slash(val) -> str:
    s = str(val)[:10]
    if len(s) == 10 and s[4] == "-":
        return s.replace("-", "/")
    return s


def _escribir_hoja_resumen(
    ws,
    empresa_nombre: str,
    quincena: str,
    periodo: PeriodoNomina,
    filas_area: List[Tuple[str, int, Decimal, Decimal, Decimal]],
) -> None:
    fi = _fmt_fecha_slash(periodo.fecha_inicio)
    ff = _fmt_fecha_slash(periodo.fecha_fin)
    estado = periodo.estado.value if isinstance(periodo.estado, PeriodoEstado) else str(periodo.estado)

    ws.merge_cells("A1:E1")
    c1 = ws["A1"]
    c1.value = f"Nómina — {empresa_nombre}"
    c1.font = FONT_TITLE
    c1.fill = FILL_TITLE
    c1.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:E2")
    c2 = ws["A2"]
    c2.value = f"{quincena}  ·  Periodo {fi} al {ff}  ·  Estado: {estado}"
    c2.font = FONT_SUBTITLE
    c2.alignment = Alignment(horizontal="left")

    header_row = 4
    for col, title in enumerate(RESUMEN_HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col, value=title)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center")
        _aplicar_borde_celda(cell)

    tot_emp = tot_perc = tot_ded = tot_neto = Decimal("0")
    row = header_row + 1
    for area_nombre, n_emp, perc, ded, neto in filas_area:
        ws.cell(row=row, column=1, value=area_nombre).alignment = Alignment(horizontal="left")
        ws.cell(row=row, column=2, value=n_emp).alignment = Alignment(horizontal="center")
        _celda_money(ws, row, 3, _float(perc))
        _celda_money(ws, row, 4, _float(ded))
        _celda_money(ws, row, 5, _float(neto))
        for col in range(1, 6):
            _aplicar_borde_celda(ws.cell(row=row, column=col))
        tot_emp += n_emp
        tot_perc += perc
        tot_ded += ded
        tot_neto += neto
        row += 1

    for col in range(1, 6):
        cell = ws.cell(row=row, column=col)
        cell.fill = FILL_TOTAL
        cell.font = FONT_TOTAL
        _aplicar_borde_celda(cell)
    ws.cell(row=row, column=1, value="TOTAL GENERAL").alignment = Alignment(horizontal="left")
    ws.cell(row=row, column=2, value=int(tot_emp)).alignment = Alignment(horizontal="center")
    _celda_money(ws, row, 3, _float(tot_perc), bold=True)
    _celda_money(ws, row, 4, _float(tot_ded), bold=True)
    _celda_money(ws, row, 5, _float(tot_neto), bold=True)

    _ajustar_columnas(ws, 1, 5, min_row=header_row)
    ws.freeze_panes = f"A{header_row + 1}"


def _escribir_hoja_detalle(
    ws,
    quincena: str,
    area: str,
    detalles: List[DetalleNominaEmpleado],
) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    titulo = ws.cell(row=1, column=1, value=f"{area} — {quincena}")
    titulo.font = FONT_TITLE
    titulo.fill = FILL_TITLE
    titulo.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    header_row = 3
    for col, title in enumerate(HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col, value=title)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        _aplicar_borde_celda(cell)

    data_start = header_row + 1
    for row_idx, d in enumerate(detalles, start=data_start):
        _, dep_nombre = departamento_de_empleado(d.empleado)
        fila_area = area or dep_nombre
        for col_idx, value in enumerate(_fila_detalle(quincena, fila_area, d), start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx >= 8 and col_idx <= 13 and isinstance(value, (int, float)):
                cell.number_format = MONEY_FMT
            elif col_idx in (5, 6) and isinstance(value, (int, float)):
                cell.number_format = "0.##"
            _aplicar_borde_celda(cell)

    n, perc, ded, neto = _totales_detalles(detalles)
    sub_row = data_start + len(detalles)
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=sub_row, column=col)
        cell.fill = FILL_SUBTOTAL
        cell.font = FONT_TOTAL
        _aplicar_borde_celda(cell)
    ws.cell(row=sub_row, column=1, value="Subtotal área")
    ws.merge_cells(start_row=sub_row, start_column=1, end_row=sub_row, end_column=4)
    ws.cell(row=sub_row, column=5, value=f"{n} recibos").alignment = Alignment(horizontal="center")
    _celda_money(ws, sub_row, 8, _float(perc), bold=True)
    _celda_money(ws, sub_row, 11, _float(ded), bold=True)
    _celda_money(ws, sub_row, 13, _float(neto), bold=True)

    _ajustar_columnas(ws, 1, len(HEADERS), min_row=header_row)
    ws.freeze_panes = f"A{data_start}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(HEADERS))}{sub_row - 1}"


def _nombre_hoja_excel(nombre: str, usados: set[str]) -> str:
    base = re.sub(r"[\[\]:*?/\\]", "_", nombre).strip()[:28] or "Area"
    titulo = base[:31]
    if titulo not in usados:
        usados.add(titulo)
        return titulo
    n = 2
    while True:
        suf = f"_{n}"
        candidato = f"{base[: 31 - len(suf)]}{suf}"
        if candidato not in usados:
            usados.add(candidato)
            return candidato
        n += 1


def _etiqueta_quincena(periodo: PeriodoNomina) -> str:
    num = numero_periodo_nomina(periodo.periodicidad, periodo.fecha_fin)
    if (periodo.periodicidad or "04") == "04" and num is not None:
        return etiqueta_quincena_numero(num)
    return str(periodo.id)


def _nombre_archivo_export(periodo: PeriodoNomina, empresa_nombre: str) -> str:
    empresa_slug = slug_area(empresa_nombre, max_len=55)
    num = numero_periodo_nomina(periodo.periodicidad, periodo.fecha_fin)
    if (periodo.periodicidad or "04") == "04" and num is not None:
        return f"nomina_{empresa_slug}_quincena_{num:02d}.xlsx"
    if num is not None:
        return f"nomina_{empresa_slug}_P{num:02d}.xlsx"
    return f"nomina_{empresa_slug}.xlsx"


def generar_xlsx_periodo(
    db: Session,
    periodo_id: int,
    departamento_id: Optional[int] = None,
) -> tuple[str, bytes]:
    """Genera un libro: hoja Resumen + una hoja por área (departamento)."""
    del departamento_id

    periodo, detalles = cargar_detalles_periodo(db, periodo_id)
    if not detalles:
        raise ValueError("No hay recibos calculados para exportar.")

    empresa = db.query(Empresa).filter(Empresa.id == periodo.empresa_id).first()
    empresa_nombre = empresa.nombre if empresa else "Empresa"

    areas_meta = listar_areas_periodo(db, periodo_id)
    grupos = agrupar_detalles_por_area(detalles)
    quincena = _etiqueta_quincena(periodo)

    orden_areas: List[tuple[Optional[int], str, List[DetalleNominaEmpleado]]] = []
    meta_por_id = {a["departamento_id"]: a["departamento_nombre"] for a in areas_meta}
    for dep_id in sorted(
        grupos.keys(),
        key=lambda k: (meta_por_id.get(k) or "Sin área").lower(),
    ):
        nombre = meta_por_id.get(dep_id) or "Sin área"
        orden_areas.append((dep_id, nombre, grupos[dep_id]))

    filas_resumen: List[Tuple[str, int, Decimal, Decimal, Decimal]] = []
    for _, area_nombre, detalles_area in orden_areas:
        n, perc, ded, neto = _totales_detalles(detalles_area)
        filas_resumen.append((area_nombre, n, perc, ded, neto))

    wb = Workbook()
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    _escribir_hoja_resumen(ws_resumen, empresa_nombre, quincena, periodo, filas_resumen)

    hojas_usadas: set[str] = {"Resumen"}
    for _, area_nombre, detalles_area in orden_areas:
        ws = wb.create_sheet(title=_nombre_hoja_excel(area_nombre, hojas_usadas))
        _escribir_hoja_detalle(ws, quincena, area_nombre, detalles_area)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = _nombre_archivo_export(periodo, empresa_nombre)
    return filename, buf.getvalue()
