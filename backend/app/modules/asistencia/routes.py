from fastapi import APIRouter, Depends, HTTPException, status, Query, Header, UploadFile, File, Form, Request
from sqlalchemy.orm import Session, joinedload
from typing import List, Optional
from datetime import datetime, timezone, timedelta, date
from app.core.database import get_db
from app.core.config import settings
from app.core.security import get_current_user
from app.core.deps import get_current_empleado_with_rol, require_superuser_download, require_superuser, require_superuser_or_rh_download
from app.modules.audit.service import ActividadService
from app.modules.personal import models as personal_models
from app.modules.personal.service import PersonalService
from . import schemas, service, models
from .biometric.sync_service import SyncService


def _parse_fecha_mexico_a_utc(fecha_str: str, es_fin: bool = False) -> Optional[datetime]:
    """
    Convierte un string de fecha/hora ingresado como hora México a UTC.
    Si es_fin=True, extiende el rango hasta el final del día México (23:59:59 → +6h UTC).
    El dispositivo y la BD guardan en UTC con offset México (UTC-6 CST / UTC-5 CDT).
    Se usa UTC-6 como offset fijo (CST), que cubre la mayor parte del año en México.
    """
    try:
        dt = datetime.fromisoformat(fecha_str)
        # Si ya tiene timezone, dejarlo como está
        if dt.tzinfo is not None:
            return dt.replace(tzinfo=None)  # quitar tz, ya es UTC si viene del frontend
        # Naive → asumir que es hora México (CST = UTC-6)
        # Convertir a UTC sumando 6 horas
        MEXICO_UTC_OFFSET = timedelta(hours=6)
        return dt + MEXICO_UTC_OFFSET
    except Exception:
        return None

router = APIRouter(prefix=f"{settings.API_V1_PREFIX}/asistencia", tags=["Asistencia"])


# ========== SINCRONIZACIÓN DESDE AGENTE LOCAL ==========

@router.post("/device-sync", response_model=schemas.AsistenciaResponse, status_code=status.HTTP_201_CREATED)
def sync_attendance(
    sync_data: schemas.AsistenciaSync,
    x_api_key: str = Header(..., alias="X-API-Key", description="API Key del agente"),
    db: Session = Depends(get_db)
):
    """
    Endpoint para recibir checadas del agente local
    
    Este es el endpoint principal que usa el agente para enviar datos del dispositivo biométrico.
    
    Headers requeridos:
    - X-API-Key: API Key del dispositivo registrado
    
    Body:
    - user_id: Número de empleado en el dispositivo
    - timestamp: Fecha y hora de la checada (ISO format)
    - device_id: ID del dispositivo (opcional, se usa la API key)
    - tipo: "entrada" o "salida" (opcional, default: "entrada")
    """
    try:
        asistencia = SyncService.sync_attendance(db, sync_data, x_api_key)
        return asistencia
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al sincronizar asistencia: {str(e)}"
        )


# ========== DISPOSITIVOS ==========
# Nota: ADMS (conexión directa dispositivo→servidor) ya no se usa. Solo el agente local sincroniza.

@router.post("/devices", response_model=schemas.DispositivoResponse, status_code=status.HTTP_201_CREATED)
def create_dispositivo(dispositivo: schemas.DispositivoCreate, db: Session = Depends(get_db)):
    """Registrar nuevo dispositivo/agente"""
    return service.AsistenciaService.create_dispositivo(db, dispositivo)


@router.get("/devices", response_model=List[schemas.DispositivoResponse])
def get_dispositivos(
    activo: Optional[bool] = None,
    db: Session = Depends(get_db)
):
    """Listar dispositivos registrados"""
    return service.AsistenciaService.get_dispositivos(db, activo=activo)


@router.get("/devices/{device_id}", response_model=schemas.DispositivoResponse)
def get_dispositivo(device_id: int, db: Session = Depends(get_db)):
    """Obtener dispositivo por ID"""
    dispositivo = service.AsistenciaService.get_dispositivo(db, device_id)
    if not dispositivo:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Dispositivo no encontrado"
        )
    return dispositivo


@router.patch("/devices/{device_id}", response_model=schemas.DispositivoResponse)
def update_dispositivo(device_id: int, data: schemas.DispositivoUpdate, db: Session = Depends(get_db)):
    """Actualizar dispositivo (nombre, ip_local, ubicación, etc.)"""
    dispositivo = service.AsistenciaService.update_dispositivo(db, device_id, data)
    if not dispositivo:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Dispositivo no encontrado")
    return dispositivo


@router.delete("/devices/{device_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_dispositivo(device_id: int, db: Session = Depends(get_db)):
    """Eliminar dispositivo. No permite si tiene checadas registradas."""
    try:
        service.AsistenciaService.delete_dispositivo(db, device_id)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


@router.post("/devices/{device_id}/test-connection", response_model=schemas.TestConnectionResponse)
def test_connection(device_id: int, db: Session = Depends(get_db)):
    """Probar configuración: crea registro de prueba"""
    result = service.AsistenciaService.test_connection(db, device_id)
    if not result["success"]:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=result["message"]
        )
    return result


@router.post("/devices/{device_id}/test-device-connection")
def test_real_device_connection(device_id: int, db: Session = Depends(get_db)):
    """Probar conexión REAL con el dispositivo (pyzk, puerto 4370). El backend debe estar en la misma red."""
    result = service.AsistenciaService.test_device_connection(db, device_id)
    if not result["success"]:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=result["message"]
        )
    return result


@router.post("/devices/{device_id}/enqueue-user", response_model=schemas.UsuarioPendienteResponse, status_code=status.HTTP_201_CREATED)
def enqueue_user(
    device_id: int,
    data: schemas.EnqueueUserRequest,
    db: Session = Depends(get_db)
):
    """Agregar usuario a la cola para alta remota en un dispositivo"""
    try:
        return service.AsistenciaService.enqueue_user(db, device_id, data)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


@router.post("/enqueue-user-multi")
def enqueue_user_multi(
    data: schemas.EnqueueUserRequest,
    dispositivo_ids: List[int] = Query(..., description="IDs de dispositivos destino"),
    db: Session = Depends(get_db)
):
    """Agregar usuario a la cola en multiples dispositivos a la vez"""
    results = []
    for did in dispositivo_ids:
        try:
            service.AsistenciaService.enqueue_user(db, did, data)
            results.append({"dispositivo_id": did, "ok": True})
        except ValueError as e:
            results.append({"dispositivo_id": did, "ok": False, "error": str(e)})
    return {"results": results}


@router.post("/devices/{device_id}/enqueue-replicate", response_model=schemas.PendingReplicateResponse, status_code=status.HTTP_201_CREATED)
def enqueue_replicate(
    device_id: int,
    data: schemas.EnqueueReplicateRequest,
    db: Session = Depends(get_db),
):
    """Encola la replicación de huella de un empleado hacia un dispositivo destino.
    Requiere que el empleado ya tenga template almacenado en el backend."""
    numero = data.numero_empleado.strip()

    dispositivo = service.AsistenciaService.get_dispositivo(db, device_id)
    if not dispositivo:
        raise HTTPException(status_code=404, detail="Dispositivo no encontrado")

    # Verificar que el EMPLEADO destino tenga huella, no cualquier empleado con el mismo numero.
    # Resolver al empleado destino vía usuarios_pendientes_dispositivo (pin_checador único).
    upd = db.query(models.UsuarioPendienteDispositivo).filter(
        models.UsuarioPendienteDispositivo.dispositivo_id == device_id,
        models.UsuarioPendienteDispositivo.numero_empleado == numero,
    ).order_by(models.UsuarioPendienteDispositivo.id.desc()).first()
    emp = None
    if upd and upd.pin_checador:
        emp = db.query(personal_models.Empleado).filter(
            personal_models.Empleado.pin_checador == upd.pin_checador
        ).first()
    if not emp:
        candidatos = db.query(personal_models.Empleado).filter(
            personal_models.Empleado.numero_empleado == numero
        ).all()
        if len(candidatos) == 1:
            emp = candidatos[0]

    if emp is None:
        raise HTTPException(
            status_code=400,
            detail=f"No se pudo resolver al empleado {numero} para este dispositivo. "
                   "Asegúrate de haberlo enviado primero al checador."
        )

    tiene_template = db.query(models.FingerprintTemplate).filter(
        models.FingerprintTemplate.empleado_id == int(emp.id)
    ).first()
    if not tiene_template:
        raise HTTPException(
            status_code=400,
            detail=f"El empleado {numero} ({emp.nombre}) no tiene huellas almacenadas en el sistema. "
                   "Primero realiza un enroll en algún dispositivo."
        )

    existente = db.query(models.PendingReplicate).filter(
        models.PendingReplicate.dispositivo_id == device_id,
        models.PendingReplicate.numero_empleado == numero,
    ).first()
    if existente:
        if existente.procesado:
            existente.procesado = False
            existente.procesado_at = None
            db.commit()
            db.refresh(existente)
        return existente

    pr = models.PendingReplicate(
        dispositivo_id=device_id,
        numero_empleado=numero,
    )
    db.add(pr)
    db.commit()
    db.refresh(pr)
    return pr


@router.get("/devices/{device_id}/pending-replicate", response_model=List[schemas.PendingReplicateResponse])
def get_pending_replicate_for_device(device_id: int, db: Session = Depends(get_db)):
    """Ver cola de replicaciones pendientes para un dispositivo"""
    return db.query(models.PendingReplicate).filter(
        models.PendingReplicate.dispositivo_id == device_id,
        models.PendingReplicate.procesado == False,
    ).all()


@router.get("/fingerprint-templates/{numero_empleado}", response_model=List[schemas.FingerprintTemplateResponse])
def get_templates_for_employee(
    numero_empleado: str,
    empleado_id: Optional[int] = Query(None),
    pin_checador: Optional[str] = Query(None, description="Si se provee, resuelve al empleado único por pin_checador (preferido por agentes)"),
    db: Session = Depends(get_db),
):
    """Ver si un empleado tiene templates de huella almacenados, con nombre del dispositivo origen.

    Cuando hay numero_empleado duplicado entre empresas, los agentes deben enviar
    pin_checador (único globalmente) para evitar mezclar plantillas.
    """
    from app.modules.personal import models as pm

    q = db.query(models.FingerprintTemplate)
    if empleado_id is not None:
        q = q.filter(models.FingerprintTemplate.empleado_id == int(empleado_id))
    elif pin_checador:
        emp = db.query(pm.Empleado).filter(pm.Empleado.pin_checador == pin_checador.strip()).first()
        if not emp:
            return []
        q = q.filter(models.FingerprintTemplate.empleado_id == int(emp.id))
    else:
        # Sin discriminador: si numero_empleado está duplicado entre empresas, devolver vacío
        # para que un agente viejo no asuma que ya tiene huella (y no se salte un upload válido).
        candidatos = db.query(pm.Empleado).filter(
            pm.Empleado.numero_empleado == numero_empleado.strip()
        ).count()
        if candidatos > 1:
            return []
        q = q.filter(models.FingerprintTemplate.numero_empleado == numero_empleado.strip())
    templates = q.all()
    # Cargar nombres de dispositivos
    device_ids = {t.source_device_id for t in templates if t.source_device_id}
    dispositivos = {d.id: d.nombre for d in db.query(models.Dispositivo).filter(
        models.Dispositivo.id.in_(device_ids)
    ).all()} if device_ids else {}
    result = []
    for t in templates:
        item = schemas.FingerprintTemplateResponse.model_validate(t)
        item.source_device_nombre = dispositivos.get(t.source_device_id)
        result.append(item)
    return result


@router.post("/devices/{device_id}/pending-users/{pending_id}/retry")
def retry_pending_user(device_id: int, pending_id: int, db: Session = Depends(get_db)):
    """Reintentar envío: marca como no enviado para que el agente lo reenvíe en el próximo ciclo"""
    pendiente = db.query(models.UsuarioPendienteDispositivo).filter(
        models.UsuarioPendienteDispositivo.id == pending_id,
        models.UsuarioPendienteDispositivo.dispositivo_id == device_id
    ).first()
    if not pendiente:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Usuario no encontrado")
    pendiente.enviado = False
    pendiente.enviado_at = None
    db.commit()
    return {"status": "ok", "message": "Se reenviará cuando el agente sincronice"}


@router.get("/devices/{device_id}/pending-users", response_model=List[schemas.UsuarioPendienteResponse])
def get_pending_users(
    device_id: int,
    include_sent: bool = Query(False, description="Incluir usuarios ya enviados"),
    db: Session = Depends(get_db)
):
    """Listar usuarios en cola (pendientes y opcionalmente enviados)"""
    return service.AsistenciaService.get_pending_users(db, device_id, include_sent=include_sent)


@router.post("/devices/{device_id}/start-enroll", response_model=schemas.PendingEnrollResponse, status_code=status.HTTP_201_CREATED)
def start_enroll(
    device_id: int,
    data: schemas.StartEnrollRequest,
    db: Session = Depends(get_db)
):
    """Iniciar registro de huella para un usuario ya enviado al dispositivo. Requiere agente en la misma red."""
    try:
        return service.AsistenciaService.start_enroll(
            db,
            device_id,
            data.numero_empleado,
            empleado_id=data.empleado_id,
            empresa_id=data.empresa_id,
        )
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


@router.get("/devices/{device_id}/enroll-status/{enroll_id}", response_model=schemas.PendingEnrollResponse)
def get_enroll_status(
    device_id: int,
    enroll_id: int,
    db: Session = Depends(get_db)
):
    """Consultar el estado de un enroll pendiente (para polling desde el frontend)."""
    pe = db.query(models.PendingEnroll).filter(
        models.PendingEnroll.id == enroll_id,
        models.PendingEnroll.dispositivo_id == device_id,
    ).first()
    if not pe:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Enroll no encontrado")
    return pe


# ========== ENDPOINTS POR EMPLEADO (vista 360°) ==========

@router.get(
    "/empleados/{empleado_id}/dispositivos",
    response_model=List[schemas.EmpleadoDispositivoEstado],
)
def get_empleado_dispositivos(empleado_id: int, db: Session = Depends(get_db)):
    """
    Devuelve el estado del empleado en cada dispositivo activo:
    - si está dado de alta (enviado), su id de cola y su pin del checador
    - si tiene plantilla en BD para replicación
    - cuántas checadas y última checada (indica que sí enroló físicamente)
    - id de pending_enroll/pending_delete activos para el dispositivo
    """
    empleado = db.query(personal_models.Empleado).filter(
        personal_models.Empleado.id == empleado_id
    ).first()
    if not empleado:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")

    dispositivos = db.query(models.Dispositivo).filter(
        models.Dispositivo.activo == True
    ).order_by(models.Dispositivo.nombre).all()

    numero = (empleado.numero_empleado or "").strip()
    pin = (empleado.pin_checador or "").strip() if empleado.pin_checador else None

    # Pre-cargar info por dispositivo en una sola consulta cada uno
    dev_ids = [d.id for d in dispositivos]
    if not dev_ids:
        return []

    pendientes_q = db.query(models.UsuarioPendienteDispositivo).filter(
        models.UsuarioPendienteDispositivo.dispositivo_id.in_(dev_ids),
        models.UsuarioPendienteDispositivo.numero_empleado == numero,
    )
    if pin:
        pendientes_q = pendientes_q.filter(
            (models.UsuarioPendienteDispositivo.pin_checador == pin)
            | (models.UsuarioPendienteDispositivo.pin_checador.is_(None))
        )
    pendientes_map = {p.dispositivo_id: p for p in pendientes_q.all()}

    enrolls = db.query(models.PendingEnroll).filter(
        models.PendingEnroll.dispositivo_id.in_(dev_ids),
        models.PendingEnroll.numero_empleado == numero,
        models.PendingEnroll.status == "pending",
    ).all()
    enroll_map = {e.dispositivo_id: e for e in enrolls}

    deletes = db.query(models.PendingDelete).filter(
        models.PendingDelete.dispositivo_id.in_(dev_ids),
        models.PendingDelete.numero_empleado == numero,
        models.PendingDelete.procesado == False,
    ).all()
    delete_map = {d.dispositivo_id: d for d in deletes}

    templates = db.query(models.FingerprintTemplate).filter(
        models.FingerprintTemplate.empleado_id == empleado_id,
    ).all()
    templates_by_dev: dict[int, list[int]] = {}
    finger_indices_servidor: list[int] = []
    origen_dev_id: int | None = None
    for t in templates:
        finger_indices_servidor.append(t.finger_index)
        if t.source_device_id is not None:
            templates_by_dev.setdefault(t.source_device_id, []).append(t.finger_index)
            if origen_dev_id is None:
                origen_dev_id = t.source_device_id
    finger_indices_servidor = sorted(set(finger_indices_servidor))
    origen_nombre: str | None = None
    if origen_dev_id is not None:
        dev_origen = db.query(models.Dispositivo).filter(models.Dispositivo.id == origen_dev_id).first()
        origen_nombre = dev_origen.nombre if dev_origen else None

    replicates = db.query(models.PendingReplicate).filter(
        models.PendingReplicate.dispositivo_id.in_(dev_ids),
        models.PendingReplicate.numero_empleado == numero,
    ).all()
    replicate_pending: set[int] = set()
    replicate_done: set[int] = set()
    for r in replicates:
        if r.procesado:
            replicate_done.add(r.dispositivo_id)
        else:
            replicate_pending.add(r.dispositivo_id)

    from sqlalchemy import func as sa_func
    checadas_rows = db.query(
        models.Asistencia.dispositivo_id,
        sa_func.count(models.Asistencia.id),
        sa_func.max(models.Asistencia.timestamp),
    ).filter(
        models.Asistencia.empleado_id == empleado_id,
        models.Asistencia.dispositivo_id.in_(dev_ids),
    ).group_by(models.Asistencia.dispositivo_id).all()
    checadas_map = {row[0]: (row[1], row[2]) for row in checadas_rows}

    result: list[schemas.EmpleadoDispositivoEstado] = []
    for d in dispositivos:
        p = pendientes_map.get(d.id)
        e = enroll_map.get(d.id)
        de = delete_map.get(d.id)
        finger_idx = sorted(set(templates_by_dev.get(d.id, [])))
        checadas_total, ultima = checadas_map.get(d.id, (0, None))
        checadas_n = int(checadas_total or 0)
        enviado = bool(p.enviado) if p else False
        result.append(schemas.EmpleadoDispositivoEstado(
            dispositivo_id=d.id,
            dispositivo_nombre=d.nombre,
            dispositivo_ubicacion=d.ubicacion,
            enviado=enviado,
            enviado_at=p.enviado_at if p else None,
            pending_user_id=p.id if p else None,
            pending_enroll_id=e.id if e else None,
            pending_delete_id=de.id if de else None,
            tiene_huella_en_bd=len(finger_idx) > 0,
            finger_indices=finger_idx,
            huella_en_servidor=len(finger_indices_servidor) > 0,
            finger_indices_servidor=finger_indices_servidor,
            huella_origen_dispositivo_id=origen_dev_id,
            huella_origen_dispositivo_nombre=origen_nombre,
            replicacion_pendiente=d.id in replicate_pending,
            replicacion_completada=d.id in replicate_done,
            presente_en_checador=enviado or checadas_n > 0,
            checadas_total=checadas_n,
            ultima_checada=ultima,
        ))
    return result


@router.post("/devices/{device_id}/queue-delete", status_code=status.HTTP_201_CREATED)
def queue_delete_user(
    device_id: int,
    data: schemas.QueueDeleteRequest,
    db: Session = Depends(get_db),
):
    """
    Encola la eliminación de un empleado del dispositivo y limpia su rastro local:
    - Inserta fila en pending_delete (la consume el agente)
    - Marca usuarios_pendientes_dispositivo.enviado=False (deja de considerarse 'dado de alta')
    - Cancela cualquier pending_enroll abierto para ese empleado en ese dispositivo
    - Borra la plantilla local en fingerprint_templates (para que no se replique)
    """
    dispositivo = db.query(models.Dispositivo).filter(models.Dispositivo.id == device_id).first()
    if not dispositivo:
        raise HTTPException(status_code=404, detail="Dispositivo no encontrado")

    empleado = None
    if data.empleado_id is not None:
        empleado = db.query(personal_models.Empleado).filter(
            personal_models.Empleado.id == int(data.empleado_id)
        ).first()
    elif data.numero_empleado:
        empleado = db.query(personal_models.Empleado).filter(
            personal_models.Empleado.numero_empleado == data.numero_empleado.strip()
        ).first()
    if not empleado:
        raise HTTPException(status_code=400, detail="Empleado no encontrado")

    numero = (empleado.numero_empleado or "").strip()

    # Si ya hay un pending_delete sin procesar, reutilizarlo
    pending_del = db.query(models.PendingDelete).filter(
        models.PendingDelete.dispositivo_id == device_id,
        models.PendingDelete.numero_empleado == numero,
        models.PendingDelete.procesado == False,
    ).first()
    if not pending_del:
        pending_del = models.PendingDelete(
            dispositivo_id=device_id,
            numero_empleado=numero,
            procesado=False,
        )
        db.add(pending_del)

    db.query(models.UsuarioPendienteDispositivo).filter(
        models.UsuarioPendienteDispositivo.dispositivo_id == device_id,
        models.UsuarioPendienteDispositivo.numero_empleado == numero,
    ).update({
        models.UsuarioPendienteDispositivo.enviado: False,
        models.UsuarioPendienteDispositivo.enviado_at: None,
    }, synchronize_session=False)

    db.query(models.PendingEnroll).filter(
        models.PendingEnroll.dispositivo_id == device_id,
        models.PendingEnroll.numero_empleado == numero,
        models.PendingEnroll.status == "pending",
    ).update({
        models.PendingEnroll.status: "failed",
        models.PendingEnroll.completed_at: datetime.now(timezone.utc),
    }, synchronize_session=False)

    db.query(models.FingerprintTemplate).filter(
        models.FingerprintTemplate.empleado_id == empleado.id,
        models.FingerprintTemplate.source_device_id == device_id,
    ).delete(synchronize_session=False)

    db.commit()
    db.refresh(pending_del)
    return {
        "ok": True,
        "pending_delete_id": pending_del.id,
        "dispositivo_id": device_id,
        "numero_empleado": numero,
        "empleado_id": empleado.id,
    }


# ========== ENDPOINTS PARA AGENTE (X-API-Key) ==========

def _get_device_from_api_key(
    request: Request,
    x_api_key: str = Header(..., alias="X-API-Key"),
    db: Session = Depends(get_db),
):
    from .biometric.agent_auth import verify_api_key
    dispositivo = verify_api_key(db, x_api_key)
    if not dispositivo:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="API Key inválida o dispositivo inactivo")
    dispositivo.ultima_sync_agente = datetime.now(timezone.utc)
    ip = (request.headers.get("X-Real-IP") or request.headers.get("X-Forwarded-For") or "").split(",")[0].strip()
    if not ip and request.client:
        ip = request.client.host or ""
    if ip:
        dispositivo.ultima_ip_conexion = ip[:50]
    db.commit()
    return dispositivo


@router.get("/agent/pin-to-numero")
def agent_get_pin_to_numero(
    dispositivo: models.Dispositivo = Depends(_get_device_from_api_key),
    db: Session = Depends(get_db)
):
    """Mapeo pin_checador -> numero_empleado para que el agente suba templates con numero_empleado correcto.
    El dispositivo guarda user_id=pin (1,2,3); al subir al backend necesitamos numero_empleado (124)."""
    from app.modules.personal import models as pm
    empleados = db.query(pm.Empleado).filter(
        pm.Empleado.pin_checador.isnot(None),
        pm.Empleado.pin_checador != ""
    ).all()
    mapping = {str(e.pin_checador).strip(): str(e.numero_empleado).strip() for e in empleados if e.pin_checador}
    # Incluir también usuarios pendientes de este dispositivo (por si aún no están en empleados con pin)
    pendientes = db.query(models.UsuarioPendienteDispositivo).filter(
        models.UsuarioPendienteDispositivo.dispositivo_id == dispositivo.id,
        models.UsuarioPendienteDispositivo.pin_checador.isnot(None),
    ).all()
    for p in pendientes:
        if p.pin_checador and p.numero_empleado:
            mapping[str(p.pin_checador).strip()] = str(p.numero_empleado).strip()
    return mapping


@router.get("/agent/diagnostic")
def agent_diagnostic(
    dispositivo: models.Dispositivo = Depends(_get_device_from_api_key),
    db: Session = Depends(get_db)
):
    """Diagnostico: verifica API Key y muestra usuarios pendientes (para depurar)"""
    pendientes = service.AsistenciaService.get_pending_users(db, dispositivo.id, include_sent=False)
    return {
        "ok": True,
        "dispositivo": {"id": dispositivo.id, "nombre": dispositivo.nombre},
        "pendientes": len(pendientes),
        "usuarios": [{"id": p.id, "numero_empleado": p.numero_empleado, "nombre": p.nombre} for p in pendientes]
    }


@router.get("/agent/pending-users", response_model=List[schemas.UsuarioPendienteResponse])
def agent_get_pending_users(
    dispositivo: models.Dispositivo = Depends(_get_device_from_api_key),
    db: Session = Depends(get_db)
):
    """Obtener usuarios pendientes de enviar (para agente con pyzk set_user)"""
    return service.AsistenciaService.get_pending_users(db, dispositivo.id, include_sent=False)


@router.post("/agent/pending-users/mark-sent")
def agent_mark_users_sent(
    data: schemas.MarkSentRequest,
    dispositivo: models.Dispositivo = Depends(_get_device_from_api_key),
    db: Session = Depends(get_db)
):
    """Marcar usuarios como enviados al dispositivo (llamado por el agente tras set_user)"""
    count = service.AsistenciaService.mark_users_sent(db, data.ids, dispositivo.id)
    return {"marked": count}


@router.get("/agent/pending-enroll", response_model=List[schemas.PendingEnrollResponse])
def agent_get_pending_enroll(
    dispositivo: models.Dispositivo = Depends(_get_device_from_api_key),
    db: Session = Depends(get_db)
):
    """Obtener enrolls pendientes (para agente con pyzk enroll_user)"""
    return service.AsistenciaService.get_pending_enrolls(db, dispositivo.id)


@router.post("/agent/pending-enroll/{enroll_id}/mark-done")
def agent_mark_enroll_done(
    enroll_id: int,
    success: bool = Query(True),
    dispositivo: models.Dispositivo = Depends(_get_device_from_api_key),
    db: Session = Depends(get_db)
):
    """Marcar enroll como completado (llamado por el agente tras enroll_user)"""
    ok = service.AsistenciaService.mark_enroll_done(db, enroll_id, dispositivo.id, success=success)
    if not ok:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Enroll no encontrado o ya procesado")
    return {"ok": True}


@router.post("/agent/upload-template")
def agent_upload_template(
    data: schemas.UploadTemplateRequest,
    dispositivo: models.Dispositivo = Depends(_get_device_from_api_key),
    db: Session = Depends(get_db)
):
    """El agente sube un template de huella despues del enroll exitoso"""
    numero = (data.numero_empleado or "").strip()
    if not numero:
        raise HTTPException(status_code=400, detail="numero_empleado es obligatorio")

    empleado = None
    if data.empleado_id is not None:
        empleado = db.query(personal_models.Empleado).filter(
            personal_models.Empleado.id == int(data.empleado_id)
        ).first()
        if not empleado:
            raise HTTPException(status_code=400, detail=f"No existe empleado_id={data.empleado_id}")
        if (empleado.numero_empleado or "").strip() != numero:
            raise HTTPException(status_code=400, detail="empleado_id no coincide con numero_empleado")
    elif (data.pin_checador or "").strip():
        pin = (data.pin_checador or "").strip()
        empleado = db.query(personal_models.Empleado).filter(
            personal_models.Empleado.pin_checador == pin
        ).first()
        if not empleado:
            raise HTTPException(status_code=400, detail=f"No existe empleado con pin_checador={pin}")
        if (empleado.numero_empleado or "").strip() != numero:
            raise HTTPException(status_code=400, detail="pin_checador no coincide con numero_empleado")
    else:
        # Compatibilidad con agentes antiguos (solo numero_empleado):
        # - Si el número es único, resolver directo.
        # - Si está duplicado entre empresas, usar el último enroll de este dispositivo.
        candidatos = db.query(personal_models.Empleado).filter(
            personal_models.Empleado.numero_empleado == numero
        ).all()
        if len(candidatos) == 1:
            empleado = candidatos[0]
        elif len(candidatos) > 1:
            pe = db.query(models.PendingEnroll).filter(
                models.PendingEnroll.dispositivo_id == dispositivo.id,
                models.PendingEnroll.numero_empleado == numero,
            ).order_by(models.PendingEnroll.created_at.desc()).first()
            if pe and pe.pin_checador:
                empleado = db.query(personal_models.Empleado).filter(
                    personal_models.Empleado.pin_checador == pe.pin_checador
                ).first()
            if not empleado:
                raise HTTPException(
                    status_code=409,
                    detail=(
                        f"numero_empleado {numero} existe en más de una empresa. "
                        "Actualiza el agente para enviar pin_checador o empleado_id."
                    ),
                )

    if not empleado:
        raise HTTPException(
            status_code=404,
            detail=f"No se pudo resolver el empleado para numero_empleado={numero}",
        )

    empleado_id = int(empleado.id)
    numero_real = (empleado.numero_empleado or numero).strip()
    existing = db.query(models.FingerprintTemplate).filter(
        models.FingerprintTemplate.empleado_id == empleado_id,
        models.FingerprintTemplate.finger_index == data.finger_index,
    ).first()
    if existing:
        existing.template_data = data.template_data
        existing.source_device_id = dispositivo.id
        existing.numero_empleado = numero_real
    else:
        tpl = models.FingerprintTemplate(
            empleado_id=empleado_id,
            numero_empleado=numero_real,
            finger_index=data.finger_index,
            template_data=data.template_data,
            source_device_id=dispositivo.id,
        )
        db.add(tpl)
    db.commit()
    return {"ok": True, "numero_empleado": numero_real, "finger_index": data.finger_index}


@router.get("/agent/pending-replicate")
def agent_get_pending_replicate(
    dispositivo: models.Dispositivo = Depends(_get_device_from_api_key),
    db: Session = Depends(get_db),
):
    """Replicaciones de huella pendientes para este dispositivo.
    Retorna la lista con template_data incluido para que el agente los suba directamente.

    Resuelve el empleado destino por pin_checador (único globalmente) para evitar mezclar
    plantillas entre empresas cuando numero_empleado está duplicado.
    """
    pending = db.query(models.PendingReplicate).filter(
        models.PendingReplicate.dispositivo_id == dispositivo.id,
        models.PendingReplicate.procesado == False,
    ).all()

    result = []
    for pr in pending:
        # Resolver el empleado destino: primero por usuarios_pendientes_dispositivo (más confiable),
        # luego por numero_empleado solo si es único.
        upd = db.query(models.UsuarioPendienteDispositivo).filter(
            models.UsuarioPendienteDispositivo.dispositivo_id == dispositivo.id,
            models.UsuarioPendienteDispositivo.numero_empleado == pr.numero_empleado,
        ).order_by(models.UsuarioPendienteDispositivo.id.desc()).first()

        emp = None
        if upd and upd.pin_checador:
            emp = db.query(personal_models.Empleado).filter(
                personal_models.Empleado.pin_checador == upd.pin_checador
            ).first()
        if not emp:
            candidatos = db.query(personal_models.Empleado).filter(
                personal_models.Empleado.numero_empleado == pr.numero_empleado
            ).all()
            if len(candidatos) == 1:
                emp = candidatos[0]
        if not emp:
            # No podemos resolver con seguridad: omitimos para no replicar la huella equivocada.
            continue

        templates = db.query(models.FingerprintTemplate).filter(
            models.FingerprintTemplate.empleado_id == int(emp.id),
        ).all()
        if not templates:
            continue

        pin = str(emp.pin_checador).strip() if emp.pin_checador else pr.numero_empleado
        nombre = f"{emp.nombre or ''} {emp.apellido_paterno or ''}".strip() or pr.numero_empleado

        for tpl in templates:
            result.append({
                "id": pr.id,
                "numero_empleado": pr.numero_empleado,
                "user_id": pin,
                "nombre": nombre,
                "finger_index": tpl.finger_index,
                "template_data": tpl.template_data,
                "create_user_first": True,
            })
    return result


@router.post("/agent/pending-replicate/{replicate_id}/mark-done")
def agent_mark_replicate_done(
    replicate_id: int,
    success: bool = Query(True),
    dispositivo: models.Dispositivo = Depends(_get_device_from_api_key),
    db: Session = Depends(get_db),
):
    """El agente marca una replicación como procesada (éxito o fallo)."""
    pr = db.query(models.PendingReplicate).filter(
        models.PendingReplicate.id == replicate_id,
        models.PendingReplicate.dispositivo_id == dispositivo.id,
    ).first()
    if not pr:
        raise HTTPException(status_code=404, detail="No encontrado")
    if success:
        pr.procesado = True
        pr.procesado_at = datetime.now(timezone.utc)
        emp = None
        upd = db.query(models.UsuarioPendienteDispositivo).filter(
            models.UsuarioPendienteDispositivo.dispositivo_id == dispositivo.id,
            models.UsuarioPendienteDispositivo.numero_empleado == pr.numero_empleado,
        ).order_by(models.UsuarioPendienteDispositivo.id.desc()).first()
        if upd and upd.pin_checador:
            emp = db.query(personal_models.Empleado).filter(
                personal_models.Empleado.pin_checador == upd.pin_checador
            ).first()
        if not emp:
            candidatos = db.query(personal_models.Empleado).filter(
                personal_models.Empleado.numero_empleado == pr.numero_empleado
            ).all()
            if len(candidatos) == 1:
                emp = candidatos[0]
        if emp:
            service.AsistenciaService.marcar_empleado_enviado_en_dispositivo(
                db, dispositivo.id, emp
            )
    db.commit()
    return {"ok": True, "success": success}


@router.get("/agent/pending-templates")
def agent_get_pending_templates(
    dispositivo: models.Dispositivo = Depends(_get_device_from_api_key),
    db: Session = Depends(get_db)
):
    """Usuarios ya enviados a este dispositivo con template de otro dispositivo (sin cola de replicación).

    Resuelve el empleado por pin_checador (único globalmente) para evitar mezclar plantillas
    entre empresas cuando hay numero_empleado duplicado.
    """
    from app.modules.personal import models as pm

    sent_users = db.query(models.UsuarioPendienteDispositivo).filter(
        models.UsuarioPendienteDispositivo.dispositivo_id == dispositivo.id,
        models.UsuarioPendienteDispositivo.enviado == True,
    ).all()
    if not sent_users:
        return []

    pins = [(u.pin_checador or "").strip() for u in sent_users if (u.pin_checador or "").strip()]
    if not pins:
        return []

    empleados = db.query(pm.Empleado).filter(pm.Empleado.pin_checador.in_(pins)).all()
    pin_to_emp = {(e.pin_checador or "").strip(): e for e in empleados if e.pin_checador}
    if not pin_to_emp:
        return []

    empleado_ids = [int(e.id) for e in pin_to_emp.values()]
    templates = db.query(models.FingerprintTemplate).filter(
        models.FingerprintTemplate.empleado_id.in_(empleado_ids),
        models.FingerprintTemplate.source_device_id != dispositivo.id,
    ).all()

    emp_id_to_pin = {int(e.id): (e.pin_checador or "").strip() for e in pin_to_emp.values()}

    return [
        {
            "numero_empleado": t.numero_empleado,
            "user_id": emp_id_to_pin.get(int(t.empleado_id), t.numero_empleado) if t.empleado_id else t.numero_empleado,
            "finger_index": t.finger_index,
            "template_data": t.template_data,
            "create_user_first": False,
            "pending_replicate": False,
        }
        for t in templates
        if t.empleado_id and int(t.empleado_id) in emp_id_to_pin
    ]


@router.get("/agent/pending-deletes")
def agent_get_pending_deletes(
    dispositivo: models.Dispositivo = Depends(_get_device_from_api_key),
    db: Session = Depends(get_db)
):
    """Obtiene usuarios pendientes de eliminar. Incluye pin_checador para que el agente elimine por ID del dispositivo.

    Resuelve el pin desde usuarios_pendientes_dispositivo (única correspondencia por dispositivo)
    para evitar borrar al usuario equivocado cuando numero_empleado está duplicado entre empresas.
    """
    pending = db.query(models.PendingDelete).filter(
        models.PendingDelete.dispositivo_id == dispositivo.id,
        models.PendingDelete.procesado == False,
    ).all()
    result = []
    for p in pending:
        pin = None
        upd = db.query(models.UsuarioPendienteDispositivo).filter(
            models.UsuarioPendienteDispositivo.dispositivo_id == dispositivo.id,
            models.UsuarioPendienteDispositivo.numero_empleado == p.numero_empleado,
        ).order_by(models.UsuarioPendienteDispositivo.id.desc()).first()
        if upd and upd.pin_checador:
            pin = str(upd.pin_checador).strip()
        else:
            candidatos = db.query(personal_models.Empleado).filter(
                personal_models.Empleado.numero_empleado == p.numero_empleado
            ).all()
            if len(candidatos) == 1 and candidatos[0].pin_checador:
                pin = str(candidatos[0].pin_checador).strip()
            elif len(candidatos) > 1:
                # No podemos resolver con seguridad: omitimos para no borrar el usuario equivocado.
                continue
        result.append({
            "id": p.id,
            "numero_empleado": p.numero_empleado,
            "pin_checador": pin or p.numero_empleado,
        })
    return result


@router.post("/agent/pending-deletes/{delete_id}/mark-done")
def agent_mark_delete_done(
    delete_id: int,
    dispositivo: models.Dispositivo = Depends(_get_device_from_api_key),
    db: Session = Depends(get_db)
):
    """Marcar eliminacion como procesada"""
    pd = db.query(models.PendingDelete).filter(
        models.PendingDelete.id == delete_id,
        models.PendingDelete.dispositivo_id == dispositivo.id,
        models.PendingDelete.procesado == False,
    ).first()
    if not pd:
        raise HTTPException(status_code=404, detail="No encontrado o ya procesado")
    pd.procesado = True
    pd.procesado_at = datetime.now(timezone.utc)
    db.commit()
    return {"ok": True}


@router.post("/agent/sync-device-users", response_model=schemas.SyncDeviceUsersResponse)
def agent_sync_device_users(
    data: schemas.SyncDeviceUsersRequest,
    dispositivo: models.Dispositivo = Depends(_get_device_from_api_key),
    db: Session = Depends(get_db),
):
    """
    El agente reporta la lista completa de usuarios del reloj (PIN + nombre).
    El servidor compara con su mapa PIN→empleado y devuelve:
      - cuántos están reconocidos
      - cuáles PINs no tienen empleado mapeado (sin_mapeo)
    Útil para detectar gente enrolada directo en el reloj sin pasar por el sistema.
    """
    from app.modules.personal import models as pm

    # Mapa completo PIN → empleado
    empleados_por_pin = {
        str(e.pin_checador).strip(): e
        for e in db.query(pm.Empleado).filter(
            pm.Empleado.pin_checador.isnot(None),
            pm.Empleado.pin_checador != "",
        ).all()
        if e.pin_checador
    }
    # Incluir pendientes de este dispositivo como mapeo adicional
    pendientes_pin = {
        str(p.pin_checador).strip()
        for p in db.query(models.UsuarioPendienteDispositivo).filter(
            models.UsuarioPendienteDispositivo.dispositivo_id == dispositivo.id,
            models.UsuarioPendienteDispositivo.pin_checador.isnot(None),
        ).all()
        if p.pin_checador
    }

    desconocidos = []
    reconocidos = 0
    for u in data.usuarios:
        pin = str(u.pin).strip()
        if pin in empleados_por_pin or pin in pendientes_pin:
            reconocidos += 1
        else:
            desconocidos.append(u)

    total = len(data.usuarios)
    sin_mapeo = len(desconocidos)
    ctx_base = {
        "dispositivo_id": dispositivo.id,
        "dispositivo_nombre": dispositivo.nombre,
        "total_en_reloj": total,
        "reconocidos": reconocidos,
        "sin_mapeo": sin_mapeo,
    }
    if total == 0:
        ActividadService.registrar(
            db,
            nivel="warning",
            categoria="checador",
            mensaje=f"Reloj «{dispositivo.nombre}»: el agente no leyó usuarios del dispositivo (lista vacía)",
            contexto=ctx_base,
            metodo_http="POST",
            ruta="/asistencia/agent/sync-device-users",
            codigo_http=200,
        )
    elif sin_mapeo:
        ctx_base["desconocidos"] = [
            {"pin": u.pin, "nombre": u.nombre} for u in desconocidos
        ]
        ActividadService.registrar(
            db,
            nivel="warning",
            categoria="checador",
            mensaje=(
                f"Reloj «{dispositivo.nombre}»: {sin_mapeo} PIN(s) sin empleado en el sistema "
                f"({reconocidos}/{total} reconocidos)"
            ),
            contexto=ctx_base,
            metodo_http="POST",
            ruta="/asistencia/agent/sync-device-users",
            codigo_http=200,
        )
    else:
        ActividadService.registrar(
            db,
            nivel="info",
            categoria="checador",
            mensaje=(
                f"Reloj «{dispositivo.nombre}»: sincronización de usuarios OK "
                f"({reconocidos}/{total} reconocidos)"
            ),
            contexto=ctx_base,
            metodo_http="POST",
            ruta="/asistencia/agent/sync-device-users",
            codigo_http=200,
        )

    return schemas.SyncDeviceUsersResponse(
        total_en_reloj=total,
        reconocidos=reconocidos,
        sin_mapeo=sin_mapeo,
        desconocidos=desconocidos,
    )


# ========== DASHBOARD ==========

@router.get("/dashboard")
def get_dashboard_stats(
    departamento_ids: Optional[str] = Query(None, description="IDs de departamentos separados por coma para filtrar por área (solo para vista general)"),
    tipo_grafica: Optional[str] = Query("global", description="Tipo de gráfica de asistencia: global, empresa, area"),
    db: Session = Depends(get_db),
    current_extra: dict = Depends(get_current_empleado_with_rol),
):
    """
    Datos del dashboard. Admin/Director/Gerente General/RH ven todo; pueden filtrar por área con departamento_ids.
    Gerentes y supervisores ven solo datos de su área.
    """
    puede_ver = current_extra.get("puede_ver_dashboard") or current_extra.get("puede_ver_mi_area")
    if not puede_ver:
        raise HTTPException(status_code=403, detail="No tienes permiso para ver el dashboard")
    from sqlalchemy import or_
    from app.core.timezone_utils import mexico_date_to_utc_range, hoy_mexico
    from app.modules.personal import models as pm

    depto_ids = current_extra.get("departamento_ids_que_administro") or []
    tiene_vista_general = current_extra.get("puede_ver_dashboard")
    filtro_manual_ids = []
    if departamento_ids and tiene_vista_general:
        try:
            filtro_manual_ids = [int(x.strip()) for x in departamento_ids.split(",") if x.strip()]
        except ValueError:
            filtro_manual_ids = []
    if filtro_manual_ids:
        depto_ids = filtro_manual_ids
        solo_mi_area = True
    else:
        solo_mi_area = current_extra.get("puede_ver_mi_area") and not tiene_vista_general and depto_ids

    hoy = hoy_mexico()
    inicio_mes = hoy.replace(day=1)
    fin_mes = inicio_mes + timedelta(days=32)
    fin_mes = fin_mes.replace(day=1) - timedelta(days=1)

    # Misma base que listado de personal: empresa asignada, no exentos (coalesce), sin Admin/Superuser
    q_empleados = PersonalService.empleados_operativos_dashboard_query(db, solo_mi_area, depto_ids)

    total_empleados = q_empleados.count()
    empleados_activos = q_empleados.filter(
        or_(pm.Empleado.estado == pm.EstadoEmpleado.ACTIVO, pm.Empleado.estado.is_(None))
    ).count()
    empleados_inactivos = q_empleados.filter(pm.Empleado.estado == pm.EstadoEmpleado.INACTIVO).count()
    empleados_baja = q_empleados.filter(pm.Empleado.estado == pm.EstadoEmpleado.BAJA).count()

    # Empresas y departamentos (para mi área: solo los de su ámbito)
    if solo_mi_area:
        total_departamentos = db.query(pm.Departamento).filter(
            pm.Departamento.activo == True,
            pm.Departamento.id.in_(depto_ids),
        ).count()
        deptos_area = db.query(pm.Departamento).filter(pm.Departamento.id.in_(depto_ids)).all()
        empresa_ids = list({d.empresa_id for d in deptos_area if d.empresa_id})
        total_empresas = db.query(pm.Empresa).filter(
            pm.Empresa.activo == True,
            pm.Empresa.id.in_(empresa_ids),
        ).count() if empresa_ids else 0
    else:
        total_empresas = db.query(pm.Empresa).filter(pm.Empresa.activo == True).count()
        total_departamentos = db.query(pm.Departamento).filter(pm.Departamento.activo == True).count()

    # Empleados del área para filtrar checadas e incidencias (misma base que KPIs)
    empleado_ids_area = None
    if solo_mi_area:
        empleado_ids_area = [r[0] for r in q_empleados.with_entities(pm.Empleado.id).all()]

    dt_inicio, _ = mexico_date_to_utc_range(inicio_mes)
    _, dt_fin = mexico_date_to_utc_range(fin_mes + timedelta(days=1))

    q_checadas = db.query(models.Asistencia).filter(
        models.Asistencia.timestamp >= dt_inicio,
        models.Asistencia.timestamp < dt_fin,
    )
    if empleado_ids_area is not None:
        q_checadas = q_checadas.filter(models.Asistencia.empleado_id.in_(empleado_ids_area))
    checadas_mes = q_checadas.count()

    q_incidencias = db.query(models.Incidencia).filter(
        models.Incidencia.fecha >= dt_inicio,
        models.Incidencia.fecha < dt_fin,
    )
    if empleado_ids_area is not None:
        q_incidencias = q_incidencias.filter(models.Incidencia.empleado_id.in_(empleado_ids_area))
    incidencias_mes = q_incidencias.count()

    # Checadas por mes (últimos 12 meses)
    checadas_por_mes: list[dict] = []
    d = inicio_mes
    for _ in range(12):
        mes_fin = d + timedelta(days=32)
        mes_fin = mes_fin.replace(day=1) - timedelta(days=1)
        di, _ = mexico_date_to_utc_range(d)
        _, df = mexico_date_to_utc_range(mes_fin + timedelta(days=1))
        q = db.query(models.Asistencia).filter(
            models.Asistencia.timestamp >= di,
            models.Asistencia.timestamp < df,
        )
        if empleado_ids_area is not None:
            q = q.filter(models.Asistencia.empleado_id.in_(empleado_ids_area))
        count = q.count()
        checadas_por_mes.append({
            "mes": d.strftime("%Y-%m"),
            "label": d.strftime("%b %Y"),
            "checadas": count,
        })
        d = (d.replace(day=1) - timedelta(days=1)).replace(day=1)

    checadas_por_mes.reverse()

    # Gráfica de asistencia vs personal (hoy)
    hoy_inicio, _ = mexico_date_to_utc_range(hoy)
    _, hoy_fin = mexico_date_to_utc_range(hoy + timedelta(days=1))
    empleados_con_checada_hoy = set(
        r[0] for r in db.query(models.Asistencia.empleado_id).filter(
            models.Asistencia.timestamp >= hoy_inicio,
            models.Asistencia.timestamp < hoy_fin,
        ).distinct().all()
    )
    if empleado_ids_area is not None:
        empleados_con_checada_hoy &= set(empleado_ids_area)

    asistencia_grafica: dict = {}
    tipo_g = (tipo_grafica or "global").lower()

    activos_operativos_ids = {
        r[0]
        for r in q_empleados.filter(
            or_(pm.Empleado.estado == pm.EstadoEmpleado.ACTIVO, pm.Empleado.estado.is_(None))
        ).with_entities(pm.Empleado.id).all()
    }

    if tipo_g == "global":
        personal_activos = len(activos_operativos_ids)
        con_asistencia = len(empleados_con_checada_hoy & activos_operativos_ids)
        asistencia_grafica = {
            "tipo": "global",
            "items": [{"label": "Personal", "personal": personal_activos, "con_asistencia": con_asistencia}],
        }
    elif tipo_g == "empresa":
        empresas = db.query(pm.Empresa).filter(pm.Empresa.activo == True).order_by(pm.Empresa.nombre).all()
        if empleado_ids_area is not None:
            deptos_area = db.query(pm.Departamento).filter(pm.Departamento.id.in_(depto_ids)).all()
            empresa_ids_scope = {d.empresa_id for d in deptos_area if d.empresa_id}
            empresas = [e for e in empresas if e.id in empresa_ids_scope]
        items = []
        for emp in empresas:
            emp_ids = [r[0] for r in db.query(pm.Empleado.id).filter(
                pm.Empleado.empresa_id == emp.id,
                pm.Empleado.id.in_(activos_operativos_ids),
            ).all()]
            personal = len(emp_ids)
            con_asistencia = len(empleados_con_checada_hoy & set(emp_ids))
            items.append({"label": emp.nombre, "personal": personal, "con_asistencia": con_asistencia})
        asistencia_grafica = {"tipo": "empresa", "items": items}
    elif tipo_g == "area":
        deptos = db.query(pm.Departamento).filter(pm.Departamento.activo == True).order_by(pm.Departamento.nombre).all()
        if empleado_ids_area is not None:
            deptos = [d for d in deptos if d.id in depto_ids]
        items = []
        for dept in deptos:
            emp_ids = [r[0] for r in db.query(pm.Empleado.id).filter(
                pm.Empleado.departamento_id == dept.id,
                pm.Empleado.id.in_(activos_operativos_ids),
            ).all()]
            personal = len(emp_ids)
            con_asistencia = len(empleados_con_checada_hoy & set(emp_ids))
            items.append({"label": dept.nombre, "personal": personal, "con_asistencia": con_asistencia})
        asistencia_grafica = {"tipo": "area", "items": items}
    else:
        asistencia_grafica = {"tipo": "global", "items": []}

    # Ausentes hoy: activos operativos en alcance sin checada (misma base que KPI «Ausentes hoy»)
    ausentes_ids = sorted(activos_operativos_ids - empleados_con_checada_hoy)
    ausentes_hoy: list[dict] = []
    if ausentes_ids:
        from app.modules.incapacidades import models as inc_models

        incap_set = {
            int(r[0])
            for r in db.query(inc_models.Incapacidad.empleado_id)
            .filter(
                inc_models.Incapacidad.empleado_id.in_(ausentes_ids),
                inc_models.Incapacidad.estado == inc_models.EstadoIncapacidad.ACTIVA,
                inc_models.Incapacidad.fecha_inicio <= hoy,
                inc_models.Incapacidad.fecha_fin >= hoy,
            )
            .distinct()
            .all()
        }
        vac_set = service.AsistenciaService.empleados_cubiertos_por_solicitud_vacaciones_aprobada(
            db, hoy
        ) & set(ausentes_ids)

        empleados_ausentes = (
            db.query(pm.Empleado)
            .options(
                joinedload(pm.Empleado.empresa),
                joinedload(pm.Empleado.departamento_rel),
            )
            .filter(pm.Empleado.id.in_(ausentes_ids))
            .order_by(
                pm.Empleado.apellido_paterno.asc(),
                pm.Empleado.apellido_materno.asc(),
                pm.Empleado.nombre.asc(),
            )
            .limit(500)
            .all()
        )
        for emp in empleados_ausentes:
            nombre = " ".join(
                x
                for x in [emp.nombre, emp.apellido_paterno, emp.apellido_materno]
                if (x or "").strip()
            ).strip()
            eid = int(emp.id)
            ausentes_hoy.append(
                {
                    "empleado_id": eid,
                    "nombre_completo": nombre or (emp.numero_empleado or f"#{eid}"),
                    "numero_empleado": emp.numero_empleado,
                    "empresa_nombre": (emp.empresa.nombre if emp.empresa else None),
                    "departamento_nombre": (
                        emp.departamento_rel.nombre if emp.departamento_rel else None
                    ),
                    "en_vacaciones": eid in vac_set,
                    "en_incapacidad": eid in incap_set,
                }
            )

    return {
        "empleados": {
            "total": total_empleados,
            "activos": empleados_activos,
            "inactivos": empleados_inactivos,
            "baja": empleados_baja,
        },
        "empresas": total_empresas,
        "departamentos": total_departamentos,
        "checadas_mes_actual": checadas_mes,
        "incidencias_mes_actual": incidencias_mes,
        "checadas_por_mes": checadas_por_mes,
        "solo_mi_area": solo_mi_area,
        "asistencia_grafica": asistencia_grafica,
        "ausentes_hoy": ausentes_hoy,
    }


# ========== ASISTENCIAS ==========

@router.get("/checadas", response_model=List[schemas.AsistenciaResponse])
def get_checadas(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    empleado_id: Optional[int] = None,
    dispositivo_id: Optional[int] = None,
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Listar checadas con filtros. Las fechas se interpretan como hora México (CST) y se convierten a UTC."""
    fecha_inicio_dt = _parse_fecha_mexico_a_utc(fecha_inicio) if fecha_inicio else None
    fecha_fin_dt = _parse_fecha_mexico_a_utc(fecha_fin) if fecha_fin else None

    return service.AsistenciaService.get_asistencias(
        db,
        skip=skip,
        limit=limit,
        empleado_id=empleado_id,
        dispositivo_id=dispositivo_id,
        fecha_inicio=fecha_inicio_dt,
        fecha_fin=fecha_fin_dt
    )


@router.get("/mis-checadas", response_model=List[schemas.AsistenciaResponse])
def get_mis_checadas(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    current: dict = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Checadas del empleado actual (portal del empleado). Las fechas se interpretan como hora México."""
    fecha_inicio_dt = _parse_fecha_mexico_a_utc(fecha_inicio) if fecha_inicio else None
    fecha_fin_dt = _parse_fecha_mexico_a_utc(fecha_fin) if fecha_fin else None
    empleado_id = int(current["user_id"])
    return service.AsistenciaService.get_asistencias(
        db,
        skip=skip,
        limit=limit,
        empleado_id=empleado_id,
        fecha_inicio=fecha_inicio_dt,
        fecha_fin=fecha_fin_dt
    )


@router.get("/mis-contexto-dias", response_model=List[schemas.DiaContextoLaboralResponse])
def get_mis_contexto_dias(
    fecha_inicio: str = Query(..., description="YYYY-MM-DD (calendario México)"),
    fecha_fin: str = Query(..., description="YYYY-MM-DD"),
    current: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Contexto de cada día en el rango: incapacidad, vacación general aplicada, solicitud de vacaciones,
    festivo o jornada normal (sábado, entre semana, etc.). Alineado con la generación de incidencias.
    """
    try:
        fi = datetime.strptime(fecha_inicio.strip(), "%Y-%m-%d").date()
        ff = datetime.strptime(fecha_fin.strip(), "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Use fechas YYYY-MM-DD")
    if ff < fi:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="fecha_fin debe ser >= fecha_inicio")
    if (ff - fi).days > 95:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Rango máximo 96 días")
    empleado_id = int(current["user_id"])
    raw = service.AsistenciaService.listar_contexto_dias_empleado_rango(db, empleado_id, fi, ff)
    return [schemas.DiaContextoLaboralResponse(**row) for row in raw]


@router.post("/cleanup-employees")
def cleanup_employees(
    keep_numeros: List[str] = Query(..., description="Numeros de empleado a conservar"),
    db: Session = Depends(get_db)
):
    """
    Elimina empleados que NO estan en la lista keep_numeros.
    Tambien elimina sus checadas y registros pendientes.
    """
    from app.modules.personal import models as pm
    all_emps = db.query(pm.Empleado).all()
    deleted_names = []
    for emp in all_emps:
        if emp.numero_empleado not in keep_numeros:
            db.query(models.Asistencia).filter(models.Asistencia.empleado_id == emp.id).delete()
            # Filtrar por pin_checador del empleado para no afectar a otro empleado con el mismo
            # numero_empleado en otra empresa.
            upd_filter = [models.UsuarioPendienteDispositivo.numero_empleado == emp.numero_empleado]
            pe_filter = [models.PendingEnroll.numero_empleado == emp.numero_empleado]
            if emp.pin_checador:
                upd_filter.append(models.UsuarioPendienteDispositivo.pin_checador == emp.pin_checador)
                pe_filter.append(models.PendingEnroll.pin_checador == emp.pin_checador)
            db.query(models.UsuarioPendienteDispositivo).filter(*upd_filter).delete(synchronize_session=False)
            db.query(models.PendingEnroll).filter(*pe_filter).delete(synchronize_session=False)
            deleted_names.append(f"{emp.numero_empleado} ({emp.nombre})")
            db.delete(emp)
    db.commit()
    return {"deleted": deleted_names, "kept": keep_numeros}


# ========== HORARIOS ==========

@router.post("/horarios", response_model=schemas.HorarioResponse, status_code=status.HTTP_201_CREATED)
def create_horario(horario: schemas.HorarioCreate, db: Session = Depends(get_db)):
    """Crear nuevo horario"""
    return service.AsistenciaService.create_horario(db, horario)


@router.get("/horarios", response_model=List[schemas.HorarioResponse])
def get_horarios(
    activo: Optional[bool] = None,
    db: Session = Depends(get_db)
):
    """Listar horarios"""
    return service.AsistenciaService.get_horarios(db, activo=activo)


# ========== INCIDENCIAS ==========

@router.post("/incidencias", response_model=schemas.IncidenciaResponse, status_code=status.HTTP_201_CREATED)
def create_incidencia(incidencia: schemas.IncidenciaCreate, db: Session = Depends(get_db)):
    """Crear nueva incidencia"""
    return service.AsistenciaService.create_incidencia(db, incidencia)


@router.get("/incidencias", response_model=List[schemas.IncidenciaResponse])
def get_incidencias(
    empleado_id: Optional[int] = None,
    tipo: Optional[str] = None,
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Listar incidencias con filtros"""
    fecha_inicio_dt = None
    fecha_fin_dt = None
    
    if fecha_inicio:
        try:
            fecha_inicio_dt = datetime.fromisoformat(fecha_inicio)
        except:
            pass
    
    if fecha_fin:
        try:
            fecha_fin_dt = datetime.fromisoformat(fecha_fin)
        except:
            pass
    
    return service.AsistenciaService.get_incidencias(
        db,
        empleado_id=empleado_id,
        tipo=tipo,
        fecha_inicio=fecha_inicio_dt,
        fecha_fin=fecha_fin_dt
    )


@router.get("/checadas/mi-area", response_model=List[schemas.AsistenciaResponse])
def get_checadas_mi_area(
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    limit: int = Query(2000, ge=1, le=5000),
    departamento_id: Optional[int] = Query(
        None,
        description="Solo superusuario: filtrar por un departamento. Sin parámetro = todos.",
    ),
    empleado_id: Optional[int] = Query(
        None,
        description="Filtrar por un empleado concreto. Debe pertenecer al área permitida (o existir si es superusuario sin filtro de departamento).",
    ),
    current_extra: dict = Depends(get_current_empleado_with_rol),
    db: Session = Depends(get_db)
):
    """Checadas del personal del área del gerente/supervisor autenticado. Requiere autenticación."""
    current_emp_id = current_extra["user_id"]
    is_superuser = current_extra.get("is_superuser") is True

    if is_superuser and departamento_id is not None:
        empleados = db.query(personal_models.Empleado).filter(
            personal_models.Empleado.departamento_id == departamento_id
        ).all()
        empleado_ids = [e.id for e in empleados]
        if not empleado_ids:
            return []
    elif is_superuser:
        empleado_ids = None
    else:
        from app.modules.personal import service as personal_service
        depto_ids = personal_service.PersonalService.get_departamento_ids_que_administro(db, current_emp_id)
        if not depto_ids:
            return []
        empleados = db.query(personal_models.Empleado).filter(
            personal_models.Empleado.departamento_id.in_(depto_ids)
        ).all()
        empleado_ids = [e.id for e in empleados]
        if not empleado_ids:
            return []

    if empleado_id is not None:
        if empleado_ids is not None and empleado_id not in empleado_ids:
            return []
        if empleado_ids is None:
            existe = (
                db.query(personal_models.Empleado.id)
                .filter(personal_models.Empleado.id == empleado_id)
                .first()
            )
            if not existe:
                return []

    fecha_inicio_dt = _parse_fecha_mexico_a_utc(fecha_inicio) if fecha_inicio else None
    fecha_fin_dt = _parse_fecha_mexico_a_utc(fecha_fin) if fecha_fin else None

    if empleado_id is not None:
        return service.AsistenciaService.get_asistencias(
            db,
            skip=0,
            limit=limit,
            empleado_id=empleado_id,
            fecha_inicio=fecha_inicio_dt,
            fecha_fin=fecha_fin_dt,
        )

    if empleado_ids is None:
        return service.AsistenciaService.get_asistencias(
            db, skip=0, limit=limit,
            fecha_inicio=fecha_inicio_dt, fecha_fin=fecha_fin_dt
        )

    # Por cada empleado del área pedir hasta 500 checadas (suficiente para una quincena);
    # luego unir, ordenar y devolver hasta `limit` para no truncar solo a un subconjunto de empleados
    limit_por_empleado = 500
    resultados = []
    for eid in empleado_ids:
        resultados += service.AsistenciaService.get_asistencias(
            db, skip=0, limit=limit_por_empleado,
            empleado_id=eid,
            fecha_inicio=fecha_inicio_dt, fecha_fin=fecha_fin_dt
        )
    resultados.sort(key=lambda a: a.timestamp, reverse=True)
    return resultados[:limit]


@router.get("/incidencias/mi-area", response_model=List[schemas.IncidenciaResponse])
def get_incidencias_mi_area(
    tipo: Optional[str] = None,
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    departamento_id: Optional[int] = Query(
        None,
        description="Solo superusuario: filtrar por un departamento. Sin parámetro = todas.",
    ),
    empleado_id: Optional[int] = Query(
        None,
        description="Filtrar por un empleado concreto (mismas reglas de alcance que checadas/mi-area).",
    ),
    current_extra: dict = Depends(get_current_empleado_with_rol),
    db: Session = Depends(get_db)
):
    """
    Lista incidencias: si es jefe de área, las de su equipo; si es superuser, todas.
    Requiere autenticación.
    """
    current_emp_id = current_extra["user_id"]
    is_superuser = current_extra.get("is_superuser") is True

    if is_superuser and departamento_id is not None:
        empleados = db.query(personal_models.Empleado).filter(
            personal_models.Empleado.departamento_id == departamento_id
        ).all()
        empleado_ids = [e.id for e in empleados]
        if not empleado_ids:
            return []
    elif is_superuser:
        empleado_ids = None
    else:
        # Área que administro: departamentos donde soy jefe (gerente) o donde soy supervisor
        from app.modules.personal import service as personal_service
        depto_ids = personal_service.PersonalService.get_departamento_ids_que_administro(db, current_emp_id)
        if not depto_ids:
            return []
        empleados = db.query(personal_models.Empleado).filter(
            personal_models.Empleado.departamento_id.in_(depto_ids)
        ).all()
        empleado_ids = [e.id for e in empleados]
        if not empleado_ids:
            return []

    if empleado_id is not None:
        if empleado_ids is not None and empleado_id not in empleado_ids:
            return []
        if empleado_ids is None:
            existe = (
                db.query(personal_models.Empleado.id)
                .filter(personal_models.Empleado.id == empleado_id)
                .first()
            )
            if not existe:
                return []

    fecha_inicio_dt = None
    fecha_fin_dt = None
    if fecha_inicio:
        try:
            fecha_inicio_dt = datetime.fromisoformat(fecha_inicio)
        except Exception:
            pass
    if fecha_fin:
        try:
            fecha_fin_dt = datetime.fromisoformat(fecha_fin)
        except Exception:
            pass
    if empleado_id is not None:
        incidencias = service.AsistenciaService.get_incidencias(
            db,
            empleado_id=empleado_id,
            empleado_ids=None,
            tipo=tipo,
            fecha_inicio=fecha_inicio_dt,
            fecha_fin=fecha_fin_dt,
        )
    else:
        incidencias = service.AsistenciaService.get_incidencias(
            db,
            empleado_ids=empleado_ids,
            tipo=tipo,
            fecha_inicio=fecha_inicio_dt,
            fecha_fin=fecha_fin_dt,
        )
    all_emp_ids = list({inc.empleado_id for inc in incidencias})
    empleados_map = {
        e.id: f"{e.nombre} {e.apellido_paterno or ''} {e.apellido_materno or ''}".strip()
        for e in db.query(personal_models.Empleado).filter(
            personal_models.Empleado.id.in_(all_emp_ids)
        ).all()
    }
    for inc in incidencias:
        inc.empleado_nombre = empleados_map.get(inc.empleado_id) or ""
    return incidencias


@router.patch("/incidencias/{incidencia_id}", response_model=schemas.IncidenciaResponse)
def update_incidencia(
    incidencia_id: int,
    body: schemas.IncidenciaUpdate,
    current_extra: dict = Depends(get_current_empleado_with_rol),
    db: Session = Depends(get_db)
):
    """
    Actualizar incidencia (justificada, comentarios).
    Gerente: puede justificar incidencias de empleados y supervisores de su área.
    Supervisor: puede justificar solo incidencias de empleados (no de supervisores).
    """
    inc = service.AsistenciaService.get_incidencia(db, incidencia_id)
    if not inc:
        raise HTTPException(status_code=404, detail="Incidencia no encontrada")
    current_id = current_extra["user_id"]
    if not current_extra.get("is_superuser"):
        from app.modules.personal import service as personal_service
        empleado = db.query(personal_models.Empleado).options(
            joinedload(personal_models.Empleado.puesto_rel)
        ).filter(personal_models.Empleado.id == inc.empleado_id).first()
        aprobadores = personal_service.PersonalService.get_ids_aprobadores_area(db, empleado.departamento_id if empleado else None)
        depto_ids_admin = current_extra.get("departamento_ids_que_administro") or []
        gg_puede_justificar_su_area = (
            current_extra.get("is_gerente_general") is True
            and empleado
            and empleado.departamento_id
            and empleado.departamento_id in depto_ids_admin
        )
        if current_id not in aprobadores and not gg_puede_justificar_su_area:
            raise HTTPException(
                status_code=403,
                detail="Solo el gerente o supervisor del área del empleado puede justificar esta incidencia"
            )
        # Supervisor solo puede justificar empleados; gerente puede justificar empleados y supervisores
        empleado_puesto = (empleado.puesto_rel.nombre or "").strip().lower() if (empleado and empleado.puesto_rel) else ""
        empleado_es_supervisor = "supervisor" in empleado_puesto
        empleado_es_gerente = "gerente" in empleado_puesto
        current_emp = db.query(personal_models.Empleado).options(
            joinedload(personal_models.Empleado.puesto_rel)
        ).filter(personal_models.Empleado.id == current_id).first()
        current_puesto = (current_emp.puesto_rel.nombre or "").strip().lower() if (current_emp and current_emp.puesto_rel) else ""
        current_es_solo_supervisor = "supervisor" in current_puesto and "gerente" not in current_puesto
        # Jefe de departamento se considera gerente para este fin
        current_es_jefe = db.query(personal_models.Departamento).filter(personal_models.Departamento.jefe_id == current_id).count() > 0
        current_es_gerente = "gerente" in current_puesto or current_es_jefe
        if current_es_solo_supervisor and (empleado_es_supervisor or empleado_es_gerente):
            raise HTTPException(
                status_code=403,
                detail="El supervisor solo puede justificar incidencias de empleados. Las de supervisores las justifica el gerente del área."
            )
    update_data = body.dict(exclude_unset=True)
    if body.justificada:
        update_data["justificado_por_id"] = current_id
    updated = service.AsistenciaService.update_incidencia(db, incidencia_id, update_data)
    return updated


# ========== HORARIOS ==========

@router.get("/horarios", response_model=List[schemas.HorarioResponse])
def get_horarios(
    activo: Optional[bool] = None,
    current_extra: dict = Depends(get_current_empleado_with_rol),
    db: Session = Depends(get_db)
):
    """Lista todos los horarios. Acceso: RH y superadmin."""
    return service.AsistenciaService.get_horarios(db, activo=activo)


@router.post("/horarios", response_model=schemas.HorarioResponse, status_code=status.HTTP_201_CREATED)
def create_horario(
    horario: schemas.HorarioCreate,
    current_extra: dict = Depends(get_current_empleado_with_rol),
    db: Session = Depends(get_db)
):
    """Crear horario. Solo RH y superadmin."""
    if not current_extra.get("is_superuser") and not current_extra.get("is_rh"):
        raise HTTPException(status_code=403, detail="Solo RH o superadmin pueden gestionar horarios")
    return service.AsistenciaService.create_horario(db, horario)


@router.put("/horarios/{horario_id}", response_model=schemas.HorarioResponse)
def update_horario(
    horario_id: int,
    horario: schemas.HorarioUpdate,
    current_extra: dict = Depends(get_current_empleado_with_rol),
    db: Session = Depends(get_db)
):
    """Actualizar horario. Solo RH y superadmin."""
    if not current_extra.get("is_superuser") and not current_extra.get("is_rh"):
        raise HTTPException(status_code=403, detail="Solo RH o superadmin pueden gestionar horarios")
    h = service.AsistenciaService.update_horario(db, horario_id, horario)
    if not h:
        raise HTTPException(status_code=404, detail="Horario no encontrado")
    return h


@router.delete("/horarios/{horario_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_horario(
    horario_id: int,
    current_extra: dict = Depends(get_current_empleado_with_rol),
    db: Session = Depends(get_db)
):
    """Desactivar horario. Solo RH y superadmin."""
    if not current_extra.get("is_superuser") and not current_extra.get("is_rh"):
        raise HTTPException(status_code=403, detail="Solo RH o superadmin pueden gestionar horarios")
    if not service.AsistenciaService.delete_horario(db, horario_id):
        raise HTTPException(status_code=404, detail="Horario no encontrado")


# ========== ASIGNACIÓN DE HORARIO A EMPLEADO ==========

@router.get("/empleados/{empleado_id}/horarios", response_model=list[schemas.EmpleadoHorarioResponse])
def get_horarios_empleado(
    empleado_id: int,
    current_extra: dict = Depends(get_current_empleado_with_rol),
    db: Session = Depends(get_db)
):
    """Lista todos los horarios activos del empleado (puede haber varios para días distintos)."""
    return service.AsistenciaService.get_horarios_activos_empleado(db, empleado_id)


@router.get("/empleados/{empleado_id}/horario", response_model=schemas.EmpleadoHorarioResponse)
def get_horario_empleado(
    empleado_id: int,
    current_extra: dict = Depends(get_current_empleado_with_rol),
    db: Session = Depends(get_db)
):
    """Obtiene el primer horario activo del empleado (compatibilidad)."""
    eh = service.AsistenciaService.get_horario_activo_empleado(db, empleado_id)
    if not eh:
        raise HTTPException(status_code=404, detail="El empleado no tiene horario asignado")
    return eh


@router.post("/empleados/{empleado_id}/horario", response_model=schemas.EmpleadoHorarioResponse, status_code=status.HTTP_201_CREATED)
def assign_horario_empleado(
    empleado_id: int,
    body: schemas.AsignarHorarioRequest,
    current_extra: dict = Depends(get_current_empleado_with_rol),
    db: Session = Depends(get_db)
):
    """
    Asigna un horario al empleado. Solo RH y superadmin.
    Si el nuevo horario comparte días con uno existente, reemplaza ese bloque de días.
    Horarios con días distintos coexisten (ej: L-V + Sábado).
    """
    if not current_extra.get("is_superuser") and not current_extra.get("is_rh"):
        raise HTTPException(status_code=403, detail="Solo RH o superadmin pueden asignar horarios")
    horario = service.AsistenciaService.get_horario(db, body.horario_id)
    if not horario:
        raise HTTPException(status_code=404, detail="Horario no encontrado")
    empleado = db.query(personal_models.Empleado).filter(personal_models.Empleado.id == empleado_id).first()
    if not empleado:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")
    try:
        return service.AsistenciaService.assign_horario_empleado(db, empleado_id, body.horario_id)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.delete("/empleados/{empleado_id}/horario/{asignacion_id}", status_code=status.HTTP_204_NO_CONTENT)
def remove_horario_asignacion(
    empleado_id: int,
    asignacion_id: int,
    current_extra: dict = Depends(get_current_empleado_with_rol),
    db: Session = Depends(get_db)
):
    """Quita una asignación de horario específica del empleado. Solo RH y superadmin."""
    if not current_extra.get("is_superuser") and not current_extra.get("is_rh"):
        raise HTTPException(status_code=403, detail="Solo RH o superadmin pueden gestionar horarios")
    if not service.AsistenciaService.remove_horario_empleado(db, empleado_id, asignacion_id):
        raise HTTPException(status_code=404, detail="Asignación no encontrada")


@router.delete("/empleados/{empleado_id}/horario", status_code=status.HTTP_204_NO_CONTENT)
def remove_todos_horarios_empleado(
    empleado_id: int,
    current_extra: dict = Depends(get_current_empleado_with_rol),
    db: Session = Depends(get_db)
):
    """Quita TODOS los horarios activos del empleado. Solo RH y superadmin."""
    if not current_extra.get("is_superuser") and not current_extra.get("is_rh"):
        raise HTTPException(status_code=403, detail="Solo RH o superadmin pueden gestionar horarios")
    if not service.AsistenciaService.remove_horario_empleado(db, empleado_id):
        raise HTTPException(status_code=404, detail="El empleado no tiene horario asignado")


# ========== PROCESO DIARIO ==========

@router.post("/horarios/procesar-dia")
def procesar_dia(
    fecha: Optional[str] = Query(None, description="Fecha YYYY-MM-DD (default: ayer)"),
    current_extra: dict = Depends(get_current_empleado_with_rol),
    db: Session = Depends(get_db)
):
    """
    Detecta faltas y checadas incompletas para todos los empleados con horario asignado.
    Si no se indica fecha, procesa el día de ayer.
    Solo RH y superadmin.
    """
    if not current_extra.get("is_superuser") and not current_extra.get("is_rh"):
        raise HTTPException(status_code=403, detail="Solo RH o superadmin pueden ejecutar este proceso")
    try:
        resultado = service.AsistenciaService.procesar_dia(db, fecha)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    return resultado


@router.post(
    "/incidencias/reconciliar-faltas-contexto",
    response_model=schemas.ReconciliarFaltasContextoResponse,
)
def reconciliar_faltas_contexto(
    fecha_inicio: str = Query(..., description="YYYY-MM-DD"),
    fecha_fin: str = Query(..., description="YYYY-MM-DD"),
    current_extra: dict = Depends(get_current_empleado_with_rol),
    db: Session = Depends(get_db),
):
    """
    Marca como **justificadas** las faltas automáticas en el rango cuando, con las reglas actuales
    (vacación general aplicada, solicitud aprobada, incapacidad, festivo), ese día no requería asistencia.
    Solo RH o superadmin.
    """
    if not current_extra.get("is_superuser") and not current_extra.get("is_rh"):
        raise HTTPException(status_code=403, detail="Solo RH o superadmin pueden ejecutar esta acción")
    try:
        fi = datetime.strptime(fecha_inicio.strip(), "%Y-%m-%d").date()
        ff = datetime.strptime(fecha_fin.strip(), "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Use fechas YYYY-MM-DD")
    if ff < fi:
        raise HTTPException(status_code=400, detail="fecha_fin debe ser >= fecha_inicio")
    if (ff - fi).days > 120:
        raise HTTPException(status_code=400, detail="Rango máximo 121 días")
    r = service.AsistenciaService.reconciliar_faltas_automaticas_con_contexto(db, fi, ff)
    return schemas.ReconciliarFaltasContextoResponse(**r)


# ========== DÍAS FESTIVOS ==========

@router.get("/festivos", response_model=list[schemas.DiaFestivoResponse])
def get_dias_festivos(
    año: Optional[int] = Query(None, description="Filtrar por año"),
    solo_activos: bool = Query(True),
    current_extra: dict = Depends(get_current_empleado_with_rol),
    db: Session = Depends(get_db)
):
    """Lista días festivos. Con ?solo_activos=false devuelve todos."""
    return service.AsistenciaService.get_dias_festivos(db, año=año, solo_activos=solo_activos)


@router.post("/festivos", response_model=schemas.DiaFestivoResponse, status_code=status.HTTP_201_CREATED)
def create_dia_festivo(
    data: schemas.DiaFestivoCreate,
    current_extra: dict = Depends(get_current_empleado_with_rol),
    db: Session = Depends(get_db)
):
    """Crear día festivo manualmente. Solo superadmin y RH."""
    if not current_extra.get("is_superuser") and not current_extra.get("is_rh"):
        raise HTTPException(status_code=403, detail="Solo RH o superadmin")
    try:
        return service.AsistenciaService.create_dia_festivo(db, data)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.patch("/festivos/{festivo_id}", response_model=schemas.DiaFestivoResponse)
def update_dia_festivo(
    festivo_id: int,
    data: schemas.DiaFestivoUpdate,
    current_extra: dict = Depends(get_current_empleado_with_rol),
    db: Session = Depends(get_db)
):
    """Actualizar nombre, tipo o activo de un día festivo. Solo superadmin y RH."""
    if not current_extra.get("is_superuser") and not current_extra.get("is_rh"):
        raise HTTPException(status_code=403, detail="Solo RH o superadmin")
    festivo = service.AsistenciaService.update_dia_festivo(db, festivo_id, data)
    if not festivo:
        raise HTTPException(status_code=404, detail="Festivo no encontrado")
    return festivo


@router.delete("/festivos/{festivo_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_dia_festivo(
    festivo_id: int,
    current_extra: dict = Depends(get_current_empleado_with_rol),
    db: Session = Depends(get_db)
):
    """Eliminar día festivo. Solo superadmin y RH."""
    if not current_extra.get("is_superuser") and not current_extra.get("is_rh"):
        raise HTTPException(status_code=403, detail="Solo RH o superadmin")
    if not service.AsistenciaService.delete_dia_festivo(db, festivo_id):
        raise HTTPException(status_code=404, detail="Festivo no encontrado")


@router.post("/festivos/generar/{anio}", status_code=status.HTTP_200_OK)
def generar_festivos_año(
    anio: int,
    current_extra: dict = Depends(get_current_empleado_with_rol),
    db: Session = Depends(get_db)
):
    """
    Auto-genera los días festivos LFT para el año indicado (Art. 74 + Semana Santa).
    Omite los que ya existen. Solo superadmin y RH.
    """
    if not current_extra.get("is_superuser") and not current_extra.get("is_rh"):
        raise HTTPException(status_code=403, detail="Solo RH o superadmin")
    if anio < 2020 or anio > 2099:
        raise HTTPException(status_code=400, detail="Año inválido (2020-2099)")
    return service.AsistenciaService.generar_festivos_año(db, anio)


def _require_superuser_checadas_especiales(current: dict):
    if not current.get("is_superuser"):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Solo el administrador puede gestionar checadas especiales",
        )


@router.get("/checadas-especiales", response_model=List[schemas.ChecadaEspecialResponse])
def listar_checadas_especiales(
    db: Session = Depends(get_db),
    current: dict = Depends(get_current_empleado_with_rol),
):
    _require_superuser_checadas_especiales(current)
    items = service.AsistenciaService.listar_checadas_especiales(db)
    return [service.AsistenciaService.map_checada_especial_response(x) for x in items]


@router.post("/checadas-especiales", response_model=schemas.ChecadaEspecialResponse, status_code=status.HTTP_201_CREATED)
def crear_checada_especial(
    body: schemas.ChecadaEspecialCreate,
    db: Session = Depends(get_db),
    current: dict = Depends(get_current_empleado_with_rol),
):
    _require_superuser_checadas_especiales(current)
    try:
        ce = service.AsistenciaService.crear_checada_especial(db, body)
        return service.AsistenciaService.map_checada_especial_response(ce)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


@router.put("/checadas-especiales/{checada_id}", response_model=schemas.ChecadaEspecialResponse)
def actualizar_checada_especial(
    checada_id: int,
    body: schemas.ChecadaEspecialUpdate,
    db: Session = Depends(get_db),
    current: dict = Depends(get_current_empleado_with_rol),
):
    _require_superuser_checadas_especiales(current)
    try:
        row = service.AsistenciaService.actualizar_checada_especial(db, checada_id, body)
        if not row:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Registro no encontrado")
        return service.AsistenciaService.map_checada_especial_response(row)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


@router.delete("/checadas-especiales/{checada_id}", status_code=status.HTTP_204_NO_CONTENT)
def eliminar_checada_especial(
    checada_id: int,
    db: Session = Depends(get_db),
    current: dict = Depends(get_current_empleado_with_rol),
):
    _require_superuser_checadas_especiales(current)
    if not service.AsistenciaService.eliminar_checada_especial(db, checada_id):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Registro no encontrado")


# ========== REPORTES ==========

@router.get("/reporte-resumen")
def reporte_resumen_asistencia(
    fecha_inicio: str = Query(..., description="YYYY-MM-DD"),
    fecha_fin: str = Query(..., description="YYYY-MM-DD"),
    empresa_id: Optional[int] = None,
    departamento_id: Optional[int] = None,
    empleado_id: Optional[int] = None,
    db: Session = Depends(get_db),
    _current: dict = Depends(get_current_user),
):
    """
    Reporte de asistencia resumido por empleado.
    Devuelve para cada empleado: días del periodo (naturales, incl. descansos),
    asistencias, faltas, retardos, salidas anticipadas e incapacidades.
    """
    from datetime import date

    try:
        fi = date.fromisoformat(fecha_inicio)
        ff = date.fromisoformat(fecha_fin)
    except ValueError:
        raise HTTPException(status_code=400, detail="Formato de fecha inválido (use YYYY-MM-DD)")

    if (ff - fi).days > 366:
        raise HTTPException(status_code=400, detail="El rango máximo es 1 año")

    return service.AsistenciaService.build_reporte_resumen(
        db, fi, ff,
        empleado_id=empleado_id,
        departamento_id=departamento_id,
        empresa_id=empresa_id,
    )


@router.get("/mi-resumen-asistencia", response_model=schemas.ResumenAsistenciaEmpleadoResponse)
def mi_resumen_asistencia(
    fecha_inicio: str = Query(..., description="YYYY-MM-DD"),
    fecha_fin: str = Query(..., description="YYYY-MM-DD"),
    db: Session = Depends(get_db),
    current: dict = Depends(get_current_user),
):
    """Resumen de asistencia y puntualidad del empleado autenticado (Mis asistencias)."""
    from datetime import date
    from app.modules.personal import models as pm

    try:
        fi = date.fromisoformat(fecha_inicio)
        ff = date.fromisoformat(fecha_fin)
    except ValueError:
        raise HTTPException(status_code=400, detail="Formato de fecha inválido (use YYYY-MM-DD)")

    if ff < fi:
        raise HTTPException(status_code=400, detail="fecha_fin debe ser >= fecha_inicio")
    if (ff - fi).days > 95:
        raise HTTPException(status_code=400, detail="El rango máximo es 96 días")

    empleado_id = int(current["user_id"])
    emp = db.query(pm.Empleado).filter(pm.Empleado.id == empleado_id).first()
    if not emp or emp.exento_incidencias:
        raise HTTPException(status_code=403, detail="No disponible para este usuario")

    rows = service.AsistenciaService.build_reporte_resumen(db, fi, ff, empleado_id=empleado_id)
    if not rows:
        total = service.AsistenciaService.dias_periodo_rango(fi, ff)
        ff_eval, en_curso = service.AsistenciaService.rango_evaluado_reporte(fi, ff)
        evaluados = service.AsistenciaService.dias_periodo_rango(fi, ff_eval)
        return {
            "empleado_id": empleado_id,
            "total_dias_periodo": total,
            "dias_periodo_evaluados": evaluados,
            "periodo_en_curso": en_curso,
            "dias_asistio": 0,
            "dias_completos": 0,
            "faltas": 0,
            "faltas_justificadas": 0,
            "incompletas": 0,
            "retardos": 0,
            "salidas_anticipadas": 0,
            "dias_incapacidad": 0,
            "dias_vacaciones": 0,
            "puntualidad_pct": 0,
        }
    row = rows[0]
    return {
        "empleado_id": row["empleado_id"],
        "total_dias_periodo": row["total_dias_periodo"],
        "dias_periodo_evaluados": row["dias_periodo_evaluados"],
        "periodo_en_curso": row["periodo_en_curso"],
        "dias_asistio": row["dias_asistio"],
        "dias_completos": row["dias_completos"],
        "faltas": row["faltas"],
        "faltas_justificadas": row["faltas_justificadas"],
        "incompletas": row.get("incompletas", 0),
        "retardos": row["retardos"],
        "salidas_anticipadas": row["salidas_anticipadas"],
        "dias_incapacidad": row["dias_incapacidad"],
        "dias_vacaciones": row["dias_vacaciones"],
        "puntualidad_pct": row["puntualidad_pct"],
    }


@router.get("/reporte-detalle/{empleado_id}")
def reporte_detalle_empleado(
    empleado_id: int,
    fecha_inicio: str = Query(...),
    fecha_fin: str = Query(...),
    db: Session = Depends(get_db),
    _current: dict = Depends(get_current_user),
):
    """
    Detalle día a día de un empleado en el período: checadas, incidencias e incapacidades.
    """
    from datetime import date, timedelta, datetime as dt
    from app.modules.incapacidades import models as inc_models
    from app.core.timezone_utils import to_mexico, mexico_date_to_utc_range
    from app.modules.personal import models as pm

    # Bloquear acceso al detalle de cuentas de sistema o usuarios exentos
    emp_check = db.query(pm.Empleado).filter(pm.Empleado.id == empleado_id).first()
    if not emp_check or emp_check.empresa_id is None or emp_check.exento_incidencias:
        raise HTTPException(status_code=403, detail="No disponible para este usuario")

    try:
        fi = date.fromisoformat(fecha_inicio)
        ff = date.fromisoformat(fecha_fin)
    except ValueError:
        raise HTTPException(status_code=400, detail="Formato de fecha inválido")

    # Rango en UTC para el período (días en calendario México)
    dt_inicio_utc, _ = mexico_date_to_utc_range(fi)
    ff_end_utc, _ = mexico_date_to_utc_range(ff + timedelta(days=1))

    # Checadas en ese rango (por día en México)
    checadas = db.query(models.Asistencia).filter(
        models.Asistencia.empleado_id == empleado_id,
        models.Asistencia.timestamp >= dt_inicio_utc,
        models.Asistencia.timestamp < ff_end_utc,
    ).order_by(models.Asistencia.timestamp).all()

    # Incidencias (mismo rango UTC)
    incidencias = db.query(models.Incidencia).filter(
        models.Incidencia.empleado_id == empleado_id,
        models.Incidencia.fecha >= dt_inicio_utc,
        models.Incidencia.fecha < ff_end_utc,
    ).order_by(models.Incidencia.fecha).all()

    # Obtener nombres de quienes justificaron (consulta explícita, evita problemas de relación entre módulos)
    justificador_ids = {inc.justificado_por_id for inc in incidencias if inc.justificado_por_id}
    justificadores_map = {}
    if justificador_ids:
        for e in db.query(personal_models.Empleado).filter(personal_models.Empleado.id.in_(justificador_ids)).all():
            justificadores_map[e.id] = f"{e.nombre} {e.apellido_paterno or ''}".strip()

    # Incapacidades
    incapacidades = db.query(inc_models.Incapacidad).filter(
        inc_models.Incapacidad.empleado_id == empleado_id,
        inc_models.Incapacidad.estado != inc_models.EstadoIncapacidad.CANCELADA,
        inc_models.Incapacidad.fecha_inicio <= ff,
        inc_models.Incapacidad.fecha_fin >= fi,
    ).all()
    incap_ranges = [(i.fecha_inicio, i.fecha_fin, i.tipo) for i in incapacidades]

    # Vacaciones aprobadas
    from app.modules.vacaciones import models as vac_models
    from datetime import datetime as dt_vac
    vacaciones = db.query(vac_models.SolicitudVacaciones).filter(
        vac_models.SolicitudVacaciones.empleado_id == empleado_id,
        vac_models.SolicitudVacaciones.estado.in_(
            (vac_models.EstadoSolicitud.APROBADA, vac_models.EstadoSolicitud.APROBADA_JEFE)
        ),
        vac_models.SolicitudVacaciones.fecha_inicio <= dt_vac(ff.year, ff.month, ff.day, 23, 59, 59),
        vac_models.SolicitudVacaciones.fecha_fin >= dt_vac(fi.year, fi.month, fi.day),
    ).all()
    vac_ranges = []
    for v in vacaciones:
        v_fi = v.fecha_inicio.date() if hasattr(v.fecha_inicio, "date") else v.fecha_inicio
        v_ff = v.fecha_fin.date() if hasattr(v.fecha_fin, "date") else v.fecha_fin
        vac_ranges.append((v_fi, v_ff))

    # Construir vista diaria
    festivos_bd = db.query(models.DiaFestivo).filter(
        models.DiaFestivo.activo == True,
        models.DiaFestivo.fecha >= fi,
        models.DiaFestivo.fecha <= ff,
    ).all()
    festivos_map = {f.fecha: f.nombre for f in festivos_bd}

    # Agrupar checadas por día (hora mostrada en México)
    checadas_por_dia: dict[str, list] = {}
    for c in checadas:
        ts_mex = to_mexico(c.timestamp) or c.timestamp
        dia = ts_mex.strftime("%Y-%m-%d")
        checadas_por_dia.setdefault(dia, []).append({
            "hora": ts_mex.strftime("%H:%M"),
            "tipo": c.tipo.value if hasattr(c.tipo, "value") else str(c.tipo),
        })

    # Agrupar incidencias por día (fecha en México)
    incidencias_por_dia: dict[str, list] = {}
    for inc in incidencias:
        ts_mex = to_mexico(inc.fecha) or inc.fecha
        dia = ts_mex.strftime("%Y-%m-%d")
        justificado_por_nombre = justificadores_map.get(inc.justificado_por_id) if inc.justificado_por_id else None
        incidencias_por_dia.setdefault(dia, []).append({
            "tipo": inc.tipo.value if hasattr(inc.tipo, "value") else str(inc.tipo),
            "descripcion": inc.descripcion,
            "justificada": inc.justificada,
            "comentarios": inc.comentarios,  # observaciones del que justificó
            "justificado_por_nombre": justificado_por_nombre,
            "origen": inc.origen,
        })

    dias = []
    d = fi
    while d <= ff:
        dia_str = d.isoformat()
        es_domingo = d.weekday() == 6
        festivo = festivos_map.get(d)
        en_incapacidad = any(i[0] <= d <= i[1] for i in incap_ranges)
        en_vacaciones = any(v[0] <= d <= v[1] for v in vac_ranges)
        dias.append({
            "fecha": dia_str,
            "dia_semana": ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"][d.weekday()],
            "es_domingo": es_domingo,
            "es_festivo": festivo is not None,
            "festivo_nombre": festivo,
            "en_incapacidad": en_incapacidad,
            "en_vacaciones": en_vacaciones,
            "checadas": checadas_por_dia.get(dia_str, []),
            "incidencias": incidencias_por_dia.get(dia_str, []),
        })
        d += timedelta(days=1)

    return {"empleado_id": empleado_id, "fecha_inicio": fecha_inicio, "fecha_fin": fecha_fin, "dias": dias}


@router.get("/reporte-export-detalle")
def reporte_export_detalle(
    fecha_inicio: str = Query(..., description="YYYY-MM-DD"),
    fecha_fin: str = Query(..., description="YYYY-MM-DD"),
    empresa_id: Optional[int] = None,
    departamento_id: Optional[int] = None,
    db: Session = Depends(get_db),
    _current: dict = Depends(get_current_user),
):
    """
    Detalle día a día de TODOS los empleados del período, para exportar XLSX completo.
    Devuelve una lista de empleados con info + array de días con checadas/incidencias.
    """
    from datetime import date, timedelta
    from app.modules.personal import models as pm
    from app.modules.incapacidades import models as inc_models
    from app.core.timezone_utils import to_mexico, mexico_date_to_utc_range

    try:
        fi = date.fromisoformat(fecha_inicio)
        ff = date.fromisoformat(fecha_fin)
    except ValueError:
        raise HTTPException(status_code=400, detail="Formato de fecha inválido")

    if (ff - fi).days > 45:
        raise HTTPException(status_code=400, detail="Máximo 45 días para el export detallado")

    q = db.query(pm.Empleado).filter(
        pm.Empleado.estado == pm.EstadoEmpleado.ACTIVO,
        pm.Empleado.empresa_id.isnot(None),
        pm.Empleado.exento_incidencias == False,
    )
    if departamento_id:
        q = q.filter(pm.Empleado.departamento_id == departamento_id)
    elif empresa_id:
        q = q.filter(pm.Empleado.empresa_id == empresa_id)

    empleados = q.order_by(pm.Empleado.apellido_paterno, pm.Empleado.nombre).all()
    if not empleados:
        return []

    emp_ids = [e.id for e in empleados]

    dep_ids = {e.departamento_id for e in empleados if e.departamento_id}
    emp_empresa_ids = {e.empresa_id for e in empleados if e.empresa_id}
    deptos_map = {
        d.id: d.nombre
        for d in db.query(pm.Departamento).filter(pm.Departamento.id.in_(dep_ids)).all()
    } if dep_ids else {}
    empresas_map = {
        em.id: em.nombre
        for em in db.query(pm.Empresa).filter(pm.Empresa.id.in_(emp_empresa_ids)).all()
    } if emp_empresa_ids else {}

    dt_inicio_utc, _ = mexico_date_to_utc_range(fi)
    ff_end_utc, _ = mexico_date_to_utc_range(ff + timedelta(days=1))

    checadas_all = db.query(models.Asistencia).filter(
        models.Asistencia.empleado_id.in_(emp_ids),
        models.Asistencia.timestamp >= dt_inicio_utc,
        models.Asistencia.timestamp < ff_end_utc,
    ).order_by(models.Asistencia.timestamp).all()

    incidencias_all = db.query(models.Incidencia).filter(
        models.Incidencia.empleado_id.in_(emp_ids),
        models.Incidencia.fecha >= dt_inicio_utc,
        models.Incidencia.fecha < ff_end_utc,
    ).order_by(models.Incidencia.fecha).all()

    incapacidades_all = db.query(inc_models.Incapacidad).filter(
        inc_models.Incapacidad.empleado_id.in_(emp_ids),
        inc_models.Incapacidad.estado != inc_models.EstadoIncapacidad.CANCELADA,
        inc_models.Incapacidad.fecha_inicio <= ff,
        inc_models.Incapacidad.fecha_fin >= fi,
    ).all()

    from app.modules.vacaciones import models as vac_models
    from datetime import datetime as dt_vac
    vacaciones_all = db.query(vac_models.SolicitudVacaciones).filter(
        vac_models.SolicitudVacaciones.empleado_id.in_(emp_ids),
        vac_models.SolicitudVacaciones.estado.in_(
            (vac_models.EstadoSolicitud.APROBADA, vac_models.EstadoSolicitud.APROBADA_JEFE)
        ),
        vac_models.SolicitudVacaciones.fecha_inicio <= dt_vac(ff.year, ff.month, ff.day, 23, 59, 59),
        vac_models.SolicitudVacaciones.fecha_fin >= dt_vac(fi.year, fi.month, fi.day),
    ).all()

    festivos_bd = db.query(models.DiaFestivo).filter(
        models.DiaFestivo.activo == True,
        models.DiaFestivo.fecha >= fi,
        models.DiaFestivo.fecha <= ff,
    ).all()
    festivos_map = {f.fecha: f.nombre for f in festivos_bd}

    checadas_idx: dict[int, dict[str, list]] = {eid: {} for eid in emp_ids}
    for c in checadas_all:
        ts_mex = to_mexico(c.timestamp) or c.timestamp
        dia = ts_mex.strftime("%Y-%m-%d")
        checadas_idx[c.empleado_id].setdefault(dia, []).append({
            "hora": ts_mex.strftime("%H:%M"),
            "tipo": c.tipo.value if hasattr(c.tipo, "value") else str(c.tipo),
        })

    incidencias_idx: dict[int, dict[str, list]] = {eid: {} for eid in emp_ids}
    for inc in incidencias_all:
        ts_mex = to_mexico(inc.fecha) or inc.fecha
        dia = ts_mex.strftime("%Y-%m-%d")
        incidencias_idx[inc.empleado_id].setdefault(dia, []).append({
            "tipo": inc.tipo.value if hasattr(inc.tipo, "value") else str(inc.tipo),
            "justificada": inc.justificada,
        })

    incap_idx: dict[int, list] = {}
    for ic in incapacidades_all:
        incap_idx.setdefault(ic.empleado_id, []).append((ic.fecha_inicio, ic.fecha_fin))

    vac_idx: dict[int, list] = {}
    for v in vacaciones_all:
        v_fi = v.fecha_inicio.date() if hasattr(v.fecha_inicio, "date") else v.fecha_inicio
        v_ff = v.fecha_fin.date() if hasattr(v.fecha_fin, "date") else v.fecha_fin
        vac_idx.setdefault(v.empleado_id, []).append((v_fi, v_ff))

    all_dias = []
    d = fi
    while d <= ff:
        all_dias.append(d)
        d += timedelta(days=1)

    resultado = []
    for emp in empleados:
        emp_checadas = checadas_idx.get(emp.id, {})
        emp_incidencias = incidencias_idx.get(emp.id, {})
        emp_incap = incap_idx.get(emp.id, [])
        emp_vac = vac_idx.get(emp.id, [])

        dias = []
        for d in all_dias:
            dia_str = d.isoformat()
            es_domingo = d.weekday() == 6
            festivo = festivos_map.get(d)
            en_incapacidad = any(i[0] <= d <= i[1] for i in emp_incap)
            en_vacaciones = any(v[0] <= d <= v[1] for v in emp_vac)
            dias.append({
                "fecha": dia_str,
                "dia_semana": ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"][d.weekday()],
                "es_domingo": es_domingo,
                "es_festivo": festivo is not None,
                "festivo_nombre": festivo,
                "en_incapacidad": en_incapacidad,
                "en_vacaciones": en_vacaciones,
                "checadas": emp_checadas.get(dia_str, []),
                "incidencias": emp_incidencias.get(dia_str, []),
            })

        resultado.append({
            "empleado_id": emp.id,
            "numero_empleado": emp.numero_empleado,
            "nombre": f"{emp.nombre} {emp.apellido_paterno or ''}".strip(),
            "empresa": empresas_map.get(emp.empresa_id, "") if emp.empresa_id else "",
            "departamento": deptos_map.get(emp.departamento_id, "") if emp.departamento_id else "",
            "dias": dias,
        })

    return resultado


@router.get("/reporte-export-xlsx")
def reporte_export_xlsx(
    fecha_inicio: str = Query(..., description="YYYY-MM-DD"),
    fecha_fin: str = Query(..., description="YYYY-MM-DD"),
    empresa_id: Optional[int] = None,
    departamento_id: Optional[int] = None,
    download_token: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    _current: dict = Depends(require_superuser_or_rh_download),
):
    """Genera XLSX: hoja 1 = resumen, luego una hoja por empresa con detalle de checadas."""
    from io import BytesIO
    from collections import defaultdict
    from datetime import date, timedelta
    from datetime import datetime as dt_cls

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from app.modules.personal import models as pm
    from app.modules.incapacidades import models as inc_models
    from app.core.timezone_utils import to_mexico, mexico_date_to_utc_range
    from sqlalchemy.orm import joinedload

    try:
        fi = date.fromisoformat(fecha_inicio)
        ff = date.fromisoformat(fecha_fin)
    except ValueError:
        raise HTTPException(status_code=400, detail="Formato de fecha inválido")

    if (ff - fi).days > 45:
        raise HTTPException(status_code=400, detail="Máximo 45 días")

    # ── Empleados ──
    q = db.query(pm.Empleado).filter(
        pm.Empleado.estado == pm.EstadoEmpleado.ACTIVO,
        pm.Empleado.empresa_id.isnot(None),
        pm.Empleado.exento_incidencias == False,
    )
    if departamento_id:
        q = q.filter(pm.Empleado.departamento_id == departamento_id)
    elif empresa_id:
        q = q.filter(pm.Empleado.empresa_id == empresa_id)
    empleados = q.options(joinedload(pm.Empleado.empresa)).order_by(pm.Empleado.apellido_paterno, pm.Empleado.nombre).all()
    if not empleados:
        raise HTTPException(status_code=404, detail="No hay empleados")

    emp_ids = [e.id for e in empleados]
    empleados_by_id = {e.id: e for e in empleados}
    dep_ids = {e.departamento_id for e in empleados if e.departamento_id}
    emp_empresa_ids = {e.empresa_id for e in empleados if e.empresa_id}
    deptos_map = {d.id: d.nombre for d in db.query(pm.Departamento).filter(pm.Departamento.id.in_(dep_ids)).all()} if dep_ids else {}
    _empresas_rows = db.query(pm.Empresa).filter(pm.Empresa.id.in_(emp_empresa_ids)).all() if emp_empresa_ids else []
    empresas_map = {em.id: em.nombre for em in _empresas_rows}
    # Etiqueta corta: usa siglas si están definidas, si no el nombre completo
    empresas_label = {em.id: (em.siglas.strip() if em.siglas and em.siglas.strip() else em.nombre) for em in _empresas_rows}

    # ── Datos ──
    dt_inicio_utc, _ = mexico_date_to_utc_range(fi)
    ff_end_utc, _ = mexico_date_to_utc_range(ff + timedelta(days=1))

    checadas_all = db.query(models.Asistencia).filter(
        models.Asistencia.empleado_id.in_(emp_ids),
        models.Asistencia.timestamp >= dt_inicio_utc,
        models.Asistencia.timestamp < ff_end_utc,
    ).order_by(models.Asistencia.timestamp).all()

    incidencias_all = db.query(models.Incidencia).filter(
        models.Incidencia.empleado_id.in_(emp_ids),
        models.Incidencia.fecha >= dt_inicio_utc,
        models.Incidencia.fecha < ff_end_utc,
    ).order_by(models.Incidencia.fecha).all()

    incapacidades_all = db.query(inc_models.Incapacidad).filter(
        inc_models.Incapacidad.empleado_id.in_(emp_ids),
        inc_models.Incapacidad.estado != inc_models.EstadoIncapacidad.CANCELADA,
        inc_models.Incapacidad.fecha_inicio <= ff,
        inc_models.Incapacidad.fecha_fin >= fi,
    ).all()

    from app.modules.vacaciones import models as vac_models
    vacaciones_all = db.query(vac_models.SolicitudVacaciones).filter(
        vac_models.SolicitudVacaciones.empleado_id.in_(emp_ids),
        vac_models.SolicitudVacaciones.estado.in_(
            (vac_models.EstadoSolicitud.APROBADA, vac_models.EstadoSolicitud.APROBADA_JEFE)
        ),
        vac_models.SolicitudVacaciones.fecha_inicio <= dt_cls(ff.year, ff.month, ff.day, 23, 59, 59),
        vac_models.SolicitudVacaciones.fecha_fin >= dt_cls(fi.year, fi.month, fi.day),
    ).all()

    festivos_bd = db.query(models.DiaFestivo).filter(
        models.DiaFestivo.activo == True,
        models.DiaFestivo.fecha >= fi,
        models.DiaFestivo.fecha <= ff,
    ).all()
    festivos_set = {f.fecha for f in festivos_bd}
    festivos_map_local = {f.fecha: f.nombre for f in festivos_bd}

    ff_eval, periodo_en_curso = service.AsistenciaService.rango_evaluado_reporte(fi, ff)
    dias_lab_fn = service.AsistenciaService.dias_laborables_rango
    dias_periodo_fn = service.AsistenciaService.dias_periodo_rango
    total_dias_periodo = dias_periodo_fn(fi, ff)
    total_dias_evaluados = dias_periodo_fn(fi, ff_eval)
    dias_lab_evaluados = dias_lab_fn(fi, ff_eval, festivos_set)

    # Indexar checadas
    checadas_idx: dict = {eid: {} for eid in emp_ids}
    dias_con_checada: dict = {eid: set() for eid in emp_ids}
    checadas_por_emp_dia: dict = {}
    for c in checadas_all:
        ts_mex = to_mexico(c.timestamp) or c.timestamp
        dia = ts_mex.strftime("%Y-%m-%d")
        checadas_idx[c.empleado_id].setdefault(dia, []).append({
            "hora": ts_mex.strftime("%H:%M"),
            "tipo": c.tipo.value if hasattr(c.tipo, "value") else str(c.tipo),
        })
        dias_con_checada[c.empleado_id].add(dia)
        key = (c.empleado_id, dia)
        checadas_por_emp_dia[key] = checadas_por_emp_dia.get(key, 0) + 1

    dias_completos = service.AsistenciaService.contar_dias_completos_reporte(
        db, empleados_by_id, checadas_por_emp_dia, ff_eval
    )

    # Indexar incidencias
    incidencias_idx: dict = {eid: {} for eid in emp_ids}
    faltas: dict = {eid: 0 for eid in emp_ids}
    faltas_just: dict = {eid: 0 for eid in emp_ids}
    incompletas: dict = {eid: 0 for eid in emp_ids}
    retardos: dict = {eid: 0 for eid in emp_ids}
    sal_antic: dict = {eid: 0 for eid in emp_ids}
    for inc in incidencias_all:
        ts_mex = to_mexico(inc.fecha) or inc.fecha
        dia = ts_mex.strftime("%Y-%m-%d")
        incidencias_idx[inc.empleado_id].setdefault(dia, []).append({
            "tipo": inc.tipo.value if hasattr(inc.tipo, "value") else str(inc.tipo),
            "justificada": inc.justificada,
            "comentarios": (inc.comentarios or "").strip() if hasattr(inc, "comentarios") else "",
        })
        eid = inc.empleado_id
        if inc.tipo == models.TipoIncidencia.FALTA:
            if inc.justificada:
                faltas_just[eid] = faltas_just.get(eid, 0) + 1
            else:
                faltas[eid] = faltas.get(eid, 0) + 1
        elif inc.tipo == models.TipoIncidencia.INCOMPLETA:
            incompletas[eid] = incompletas.get(eid, 0) + 1
        elif inc.tipo == models.TipoIncidencia.RETARDO:
            retardos[eid] = retardos.get(eid, 0) + 1
        elif inc.tipo == models.TipoIncidencia.SALIDA_ANTICIPADA:
            sal_antic[eid] = sal_antic.get(eid, 0) + 1

    # Indexar incapacidades
    incap_idx: dict = {}
    dias_incap: dict = {eid: 0 for eid in emp_ids}
    for ic in incapacidades_all:
        incap_idx.setdefault(ic.empleado_id, []).append((ic.fecha_inicio, ic.fecha_fin))
        inicio_real = max(ic.fecha_inicio, fi)
        fin_real = min(ic.fecha_fin, ff_eval)
        if fin_real >= inicio_real:
            dias_incap[ic.empleado_id] = dias_incap.get(ic.empleado_id, 0) + dias_lab_fn(inicio_real, fin_real, festivos_set)

    # Indexar vacaciones
    vac_idx: dict = {}
    dias_vac: dict = {eid: 0 for eid in emp_ids}
    for v in vacaciones_all:
        v_fi = v.fecha_inicio.date() if hasattr(v.fecha_inicio, "date") else v.fecha_inicio
        v_ff = v.fecha_fin.date() if hasattr(v.fecha_fin, "date") else v.fecha_fin
        vac_idx.setdefault(v.empleado_id, []).append((v_fi, v_ff))
        inicio_real = max(v_fi, fi)
        fin_real = min(v_ff, ff_eval)
        if fin_real >= inicio_real:
            dias_vac[v.empleado_id] = dias_vac.get(v.empleado_id, 0) + dias_lab_fn(inicio_real, fin_real, festivos_set)

    all_dias = []
    dd = fi
    while dd <= ff:
        all_dias.append(dd)
        dd += timedelta(days=1)

    TIPO_INC = {"falta": "Falta", "incompleta": "Incompleta", "retardo": "Retardo",
                "salida_anticipada": "Sal. Anticipada", "horas_extra": "Hrs Extra"}
    DIAS_SEM = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]
    MESES_ES = ["Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
    periodo_txt = f"{fi.day} de {MESES_ES[fi.month-1]} al {ff.day} de {MESES_ES[ff.month-1]} de {ff.year}"
    dias_periodo_txt = (
        f"{total_dias_evaluados} de {total_dias_periodo} días del periodo (quincena en curso)"
        if periodo_en_curso
        else f"{total_dias_periodo} días del periodo (incluye domingos de descanso)"
    )

    # Agrupar empleados en hojas según el filtro solicitado:
    #  - Filtro por departamento  → una sola hoja con ese departamento
    #  - Filtro por empresa       → una hoja por departamento de esa empresa
    #  - Global (sin filtro)      → una hoja por empresa (con bloques por macro-área: Operaciones / Administración / Otras)
    grupos_detalle: dict = defaultdict(list)
    if departamento_id:
        depto_name = deptos_map.get(departamento_id, "Departamento")
        for emp in empleados:
            grupos_detalle[depto_name].append(emp)
    elif empresa_id:
        for emp in empleados:
            key = deptos_map.get(emp.departamento_id, "Sin Depto") if emp.departamento_id else "Sin Depto"
            grupos_detalle[key].append(emp)
    else:
        for emp in empleados:
            key = empresas_map.get(emp.empresa_id, "Sin Empresa") if emp.empresa_id else "Sin Empresa"
            grupos_detalle[key].append(emp)

    def clasificar_macro_area(nombre_depto: str) -> str:
        """Operaciones vs Administración vs resto (según nombre del departamento en catálogo)."""
        n = (nombre_depto or "").strip().lower()
        if not n:
            return "OTRAS ÁREAS"
        adm_kw = (
            "admin",
            "administrac",
            "rrhh",
            "recursos humano",
            "rh ",
            " contab",
            "contabi",
            "finanz",
            "tesorer",
            "sistemas",
            "tic",
            "legal",
            "direc",
            "corporativ",
            "oficina central",
            "gerenc",
        )
        if any(k in n for k in adm_kw):
            return "ADMINISTRACIÓN"
        op_kw = (
            "operac",
            "venta",
            "tienda",
            "sucursal",
            "almacén",
            "almacen",
            "distrib",
            "producc",
            "plant",
            "taller",
            "logíst",
            "logist",
            "bodega",
            "mostrador",
        )
        if any(k in n for k in op_kw):
            return "OPERACIONES"
        return "OTRAS ÁREAS"

    MACRO_ORDEN = ("OPERACIONES", "ADMINISTRACIÓN", "OTRAS ÁREAS")

    def iter_detalle_por_area_global(emp_list: list) -> list:
        """Secuencia de ('sec', etiqueta, n) o ('emp', empleado) para export global por empresa."""
        buckets = {k: [] for k in MACRO_ORDEN}
        for emp in emp_list:
            dn = deptos_map.get(emp.departamento_id, "") if emp.departamento_id else ""
            buckets[clasificar_macro_area(dn)].append(emp)
        out = []
        for m in MACRO_ORDEN:
            if not buckets[m]:
                continue
            buckets[m].sort(key=lambda e: (e.apellido_paterno or "", e.nombre or ""))
            out.append(("sec", m, len(buckets[m])))
            for emp in buckets[m]:
                out.append(("emp", emp))
        return out

    # ── Estilos ──
    thin = Side(style="thin", color="B0B0B0")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    font_title = Font(name="Calibri", size=14, bold=True, color="1A365D")
    font_subtitle = Font(name="Calibri", size=11, bold=True, color="2D3748")
    font_periodo = Font(name="Calibri", size=10, color="4A5568")
    font_emp_name = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_emp_info = Font(name="Calibri", size=10, color="E2E8F0")
    font_header = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
    font_data = Font(name="Calibri", size=9)
    font_data_num = Font(name="Calibri", size=9)
    font_data_zero = Font(name="Calibri", size=9, color="A0AEC0")
    font_inc_falta = Font(name="Calibri", size=9, bold=True, color="C53030")
    font_inc_retardo = Font(name="Calibri", size=9, color="B7791F")
    font_inc_vacaciones = Font(name="Calibri", size=9, bold=True, color="166534")
    font_total_lbl = Font(name="Calibri", size=10, bold=True, color="1A365D")
    font_total_val = Font(name="Calibri", size=10, bold=True, color="1A365D")
    font_footer = Font(name="Calibri", size=8, italic=True, color="718096")
    font_resumen_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    font_resumen_data = Font(name="Calibri", size=10)
    font_resumen_name = Font(name="Calibri", size=10, bold=True, color="1A365D")

    fill_emp_bg = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
    fill_col_header = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
    fill_row_even = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid")
    fill_row_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_domingo = PatternFill(start_color="EDF2F7", end_color="EDF2F7", fill_type="solid")
    fill_festivo = PatternFill(start_color="FFFBEB", end_color="FFFBEB", fill_type="solid")
    fill_incap_fill = PatternFill(start_color="EBF8FF", end_color="EBF8FF", fill_type="solid")
    fill_vacaciones = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")
    fill_falta = PatternFill(start_color="FFF5F5", end_color="FFF5F5", fill_type="solid")
    fill_total_row = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    fill_resumen_hdr = PatternFill(start_color="2B6CB0", end_color="2B6CB0", fill_type="solid")
    fill_resumen_even = PatternFill(start_color="EBF4FF", end_color="EBF4FF", fill_type="solid")
    font_area_title = Font(name="Calibri", size=12, bold=True, color="1A202C")
    fill_area_operaciones = PatternFill(start_color="C6F6D5", end_color="C6F6D5", fill_type="solid")
    fill_area_admin = PatternFill(start_color="E9D8FD", end_color="E9D8FD", fill_type="solid")
    fill_area_otras = PatternFill(start_color="CBD5E0", end_color="CBD5E0", fill_type="solid")

    def fill_por_macro_etiqueta(m: str):
        if m == "OPERACIONES":
            return fill_area_operaciones
        if m == "ADMINISTRACIÓN":
            return fill_area_admin
        return fill_area_otras

    align_c = Alignment(horizontal="center", vertical="center")
    align_l = Alignment(horizontal="left", vertical="center")
    align_r = Alignment(horizontal="right", vertical="center")
    align_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def to_min(h: str) -> int:
        hh, mm = h.split(":")
        return int(hh) * 60 + int(mm)

    wb = Workbook()

    # ═══════════════════════════════════════════════════════════════════════════
    #  HOJA 1: RESUMEN
    # ═══════════════════════════════════════════════════════════════════════════
    ws_res = wb.active
    ws_res.title = "Resumen"
    RES_COLS = 13
    res_widths = [6, 10, 28, 20, 18, 10, 10, 10, 10, 10, 10, 11, 10]
    for i, w in enumerate(res_widths, 1):
        ws_res.column_dimensions[get_column_letter(i)].width = w

    row = 1
    ws_res.merge_cells(start_row=row, start_column=1, end_row=row, end_column=RES_COLS)
    ws_res.cell(row=row, column=1, value="INFORME DE ASISTENCIA — RESUMEN GENERAL").font = font_title
    ws_res.cell(row=row, column=1).alignment = align_c
    ws_res.row_dimensions[row].height = 30

    row = 2
    ws_res.merge_cells(start_row=row, start_column=1, end_row=row, end_column=RES_COLS)
    ws_res.cell(row=row, column=1, value=f"Período: {periodo_txt}   |   {dias_periodo_txt}").font = font_periodo
    ws_res.cell(row=row, column=1).alignment = align_c

    row = 4
    res_headers = ["#", "No. Emp", "Nombre", "Empresa", "Departamento", "Asistencias",
                   "Completos", "Faltas", "Faltas Just.", "Retardos", "Incapacidad", "Vacaciones", "% Punt."]
    for ci, h in enumerate(res_headers, 1):
        cell = ws_res.cell(row=row, column=ci, value=h)
        cell.font = font_resumen_header
        cell.fill = fill_resumen_hdr
        cell.alignment = align_c
        cell.border = border_all
    ws_res.row_dimensions[row].height = 22

    for idx, emp in enumerate(empleados, 1):
        row += 1
        f_val = faltas.get(emp.id, 0)
        fj_val = faltas_just.get(emp.id, 0)
        di_val = dias_incap.get(emp.id, 0)
        dv_val = dias_vac.get(emp.id, 0)
        dc_val = dias_completos.get(emp.id, 0)
        denominador = dias_lab_evaluados - di_val - dv_val - f_val - fj_val
        pct = round((dc_val / max(1, denominador)) * 100, 1) if denominador > 0 else 0

        vals = [
            idx,
            emp.numero_empleado,
            f"{emp.nombre} {emp.apellido_paterno or ''}".strip(),
            empresas_label.get(emp.empresa_id, "") if emp.empresa_id else "",
            deptos_map.get(emp.departamento_id, "") if emp.departamento_id else "",
            len(dias_con_checada.get(emp.id, set())),
            dc_val,
            f_val,
            fj_val,
            retardos.get(emp.id, 0),
            di_val,
            dv_val,
            f"{pct}%",
        ]
        fill = fill_resumen_even if idx % 2 == 0 else fill_row_odd
        for ci, v in enumerate(vals, 1):
            cell = ws_res.cell(row=row, column=ci, value=v)
            cell.border = border_all
            cell.fill = fill
            cell.alignment = align_c if ci not in (3, 4, 5) else align_l
            if ci == 3:
                cell.font = font_resumen_name
            elif isinstance(v, int) and v == 0:
                cell.font = font_data_zero
            else:
                cell.font = font_resumen_data

    row += 2
    ws_res.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws_res.cell(row=row, column=1, value=f"Generado: {dt_cls.now().strftime('%d/%m/%Y %H:%M')}").font = font_footer
    ws_res.merge_cells(start_row=row, start_column=7, end_row=row, end_column=RES_COLS)
    ws_res.cell(row=row, column=7, value=f"Total: {len(empleados)} empleados").font = font_footer
    ws_res.cell(row=row, column=7).alignment = align_r

    ws_res.freeze_panes = "A5"
    ws_res.auto_filter.ref = f"A4:{get_column_letter(RES_COLS)}{4 + len(empleados)}"

    # ═══════════════════════════════════════════════════════════════════════════
    #  HOJAS DE DETALLE (agrupadas según filtro)
    #  - Filtro departamento → 1 hoja con ese departamento
    #  - Filtro empresa      → 1 hoja por departamento
    #  - Global (todas las empresas) → 1 hoja por empresa, bloques por macro-área
    # ═══════════════════════════════════════════════════════════════════════════
    DET_COLS = 10
    det_widths = [14, 7, 10, 12, 12, 10, 22, 12, 28, 13]

    for grupo_nombre in sorted(grupos_detalle.keys()):
        emp_list = grupos_detalle[grupo_nombre]
        sheet_name = grupo_nombre[:31]
        ws = wb.create_sheet(title=sheet_name)

        for i, w in enumerate(det_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        row = 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=DET_COLS)
        ws.cell(row=row, column=1, value=grupo_nombre.upper()).font = font_title
        ws.cell(row=row, column=1).alignment = align_c
        ws.row_dimensions[row].height = 30

        row = 2
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=DET_COLS)
        ws.cell(row=row, column=1, value="Detalle de Checadas por Empleado").font = font_subtitle
        ws.cell(row=row, column=1).alignment = align_c

        row = 3
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=DET_COLS)
        subt_detalle = f"Período: {periodo_txt}   |   {len(emp_list)} empleados"
        if not departamento_id and not empresa_id:
            subt_detalle += "   |   Por área: Operaciones → Administración → Otras"
        ws.cell(row=row, column=1, value=subt_detalle).font = font_periodo
        ws.cell(row=row, column=1).alignment = align_c

        row = 4

        export_global_por_empresa = not departamento_id and not empresa_id
        if export_global_por_empresa:
            loop_items = iter_detalle_por_area_global(emp_list)
        else:
            loop_items = [
                ("emp", e)
                for e in sorted(emp_list, key=lambda x: (x.apellido_paterno or "", x.nombre or ""))
            ]

        for item in loop_items:
            if item[0] == "sec":
                area_label = item[1]
                n_personas = item[2]
                row += 1
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=DET_COLS)
                fa = fill_por_macro_etiqueta(area_label)
                if area_label == "ADMINISTRACIÓN":
                    texto_banner = (
                        f"▶  ÁREA: ADMINISTRACIÓN  —  {n_personas} persona(s)\n"
                        f"    Identificador: personal administrativo (bloque separado de Operaciones)"
                    )
                else:
                    texto_banner = f"▶  ÁREA: {area_label}  —  {n_personas} persona(s)"
                c_banner = ws.cell(row=row, column=1, value=texto_banner)
                c_banner.font = font_area_title
                c_banner.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                c_banner.fill = fa
                for col in range(1, DET_COLS + 1):
                    cell_b = ws.cell(row=row, column=col)
                    cell_b.fill = fa
                    cell_b.border = border_all
                ws.row_dimensions[row].height = 28
                continue

            emp = item[1]
            emp_name = f"{emp.nombre} {emp.apellido_paterno or ''}".strip()
            emp_depto = deptos_map.get(emp.departamento_id, "") if emp.departamento_id else ""
            emp_empresa = empresas_label.get(emp.empresa_id, "") if emp.empresa_id else ""
            emp_ch = checadas_idx.get(emp.id, {})
            emp_inc = incidencias_idx.get(emp.id, {})
            emp_icap = incap_idx.get(emp.id, [])
            emp_vacaciones = vac_idx.get(emp.id, [])

            row += 1
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
            ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=DET_COLS)
            ws.cell(row=row, column=1, value=emp_name).font = font_emp_name
            ws.cell(row=row, column=5, value=f"No. {emp.numero_empleado}").font = font_emp_info
            info_col7 = emp_depto if not departamento_id else emp_empresa
            ws.cell(row=row, column=7, value=info_col7).font = font_emp_info
            for col in range(1, DET_COLS + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = fill_emp_bg
                cell.border = border_all
                cell.alignment = align_c if col >= 5 else align_l
            ws.row_dimensions[row].height = 24

            row += 1
            for ci, h in enumerate(["Fecha", "Día", "Entrada", "Sal. Comer", "Reg. Comer", "Salida", "Incidencia", "Justificación", "Motivo", "Tiempo"], 1):
                cell = ws.cell(row=row, column=ci, value=h)
                cell.font = font_header
                cell.fill = fill_col_header
                cell.alignment = align_c
                cell.border = border_all
            ws.row_dimensions[row].height = 20

            total_min = 0
            for idx_d, d_date in enumerate(all_dias):
                dia_str = d_date.isoformat()
                es_dom = d_date.weekday() == 6
                festivo = festivos_map_local.get(d_date)
                en_incap = any(i[0] <= d_date <= i[1] for i in emp_icap)
                en_vac = any(v[0] <= d_date <= v[1] for v in emp_vacaciones)
                dc = emp_ch.get(dia_str, [])
                di = emp_inc.get(dia_str, [])

                cmap: dict = {}
                for cc in dc:
                    cmap[cc["tipo"]] = cc["hora"]

                inc_text = ""
                just_text = ""
                motivo_text = ""
                if es_dom:
                    inc_text = "Descanso"
                elif en_vac:
                    inc_text = "Vacaciones"
                elif en_incap:
                    inc_text = "Incapacidad"
                elif festivo:
                    inc_text = f"Festivo: {festivo}"
                else:
                    parts = []
                    motivos: list[str] = []
                    todas_just = bool(di) and all(ii["justificada"] for ii in di)
                    alguna_just = any(ii["justificada"] for ii in di)
                    for ii in di:
                        lbl = TIPO_INC.get(ii["tipo"], ii["tipo"])
                        if ii["justificada"]:
                            obs = (ii.get("comentarios") or "").strip()
                            if obs and obs not in motivos:
                                motivos.append(obs)
                            lbl += " (J)"
                        parts.append(lbl)
                    inc_text = ", ".join(parts)
                    if di:
                        just_text = "Sí" if todas_just else ("Parcial" if alguna_just else "No")
                    motivo_text = "\n".join(motivos)

                tiempo_str = ""
                if cmap.get("entrada") and cmap.get("salida"):
                    t = to_min(cmap["salida"]) - to_min(cmap["entrada"])
                    if cmap.get("salida_comer") and cmap.get("regreso_comer"):
                        t -= to_min(cmap["regreso_comer"]) - to_min(cmap["salida_comer"])
                    if t > 0:
                        total_min += t
                        tiempo_str = f"{t // 60}:{t % 60:02d}"

                row += 1
                vals = [d_date.strftime("%d/%m/%Y"), DIAS_SEM[d_date.weekday()],
                        cmap.get("entrada", ""), cmap.get("salida_comer", ""),
                        cmap.get("regreso_comer", ""), cmap.get("salida", ""),
                        inc_text, just_text, motivo_text, tiempo_str]

                if es_dom:
                    fill = fill_domingo
                elif en_vac:
                    fill = fill_vacaciones
                elif en_incap:
                    fill = fill_incap_fill
                elif festivo:
                    fill = fill_festivo
                elif any(ii["tipo"] == "falta" and not ii.get("justificada") for ii in di):
                    fill = fill_falta
                else:
                    fill = fill_row_even if idx_d % 2 == 0 else fill_row_odd

                for ci, v in enumerate(vals, 1):
                    cell = ws.cell(row=row, column=ci, value=v)
                    cell.font = font_data
                    cell.fill = fill
                    cell.border = border_all
                    cell.alignment = align_wrap if ci in (8, 9) else (align_c if ci >= 2 else align_l)

                inc_cell = ws.cell(row=row, column=7)
                if inc_text == "Vacaciones":
                    inc_cell.font = font_inc_vacaciones
                elif "Falta" in inc_text and "(J)" not in inc_text:
                    inc_cell.font = font_inc_falta
                elif "Retardo" in inc_text:
                    inc_cell.font = font_inc_retardo

                motivo_lines = (motivo_text.count("\n") + 1) if motivo_text else 1
                ws.row_dimensions[row].height = int(min(18 + (motivo_lines - 1) * 12, 80))

            row += 1
            th_val = total_min // 60
            tm_val = total_min % 60
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            for col in range(1, DET_COLS + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = fill_total_row
                cell.border = border_all
            ws.cell(row=row, column=7, value="TOTAL HORAS").font = font_total_lbl
            ws.cell(row=row, column=7).alignment = align_r
            ws.cell(row=row, column=10, value=f"{th_val}:{tm_val:02d}").font = font_total_val
            ws.cell(row=row, column=10).alignment = align_c
            ws.row_dimensions[row].height = 22

        row += 2
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.cell(row=row, column=1, value=f"Generado: {dt_cls.now().strftime('%d/%m/%Y %H:%M')}").font = font_footer
        ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=DET_COLS)
        ws.cell(row=row, column=6, value=f"Total: {len(emp_list)} empleados").font = font_footer
        ws.cell(row=row, column=6).alignment = align_r

        ws.freeze_panes = "A5"

    # ── Guardar ──
    output = BytesIO()
    wb.save(output)
    contenido = output.getvalue()
    output.close()
    wb.close()

    # Construir nombre descriptivo según el filtro aplicado
    import re as _re
    def _slug(s: str) -> str:
        return _re.sub(r'[^a-zA-Z0-9]+', '_', s).strip('_')

    periodo = f"{fecha_inicio}_{fecha_fin}"
    if departamento_id:
        dep_nombre = deptos_map.get(departamento_id, str(departamento_id))
        emp_nombre = empresas_map.get(empresa_id, str(empresa_id)) if empresa_id else (
            empresas_map.get(next(iter({e.empresa_id for e in empleados if e.departamento_id == departamento_id}), None), "")
        )
        filename = f"{_slug(emp_nombre)}_{_slug(dep_nombre)}_{periodo}.xlsx"
    elif empresa_id:
        emp_nombre = empresas_map.get(empresa_id, str(empresa_id))
        filename = f"{_slug(emp_nombre)}_{periodo}.xlsx"
    else:
        filename = f"Reporte_General_{periodo}.xlsx"
    from fastapi.responses import Response
    ActividadService.registrar(
        db,
        nivel="info",
        categoria="negocio",
        mensaje="Exportación de reporte de asistencia (XLSX)",
        contexto={
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin,
            "empresa_id": empresa_id,
            "departamento_id": departamento_id,
            "filename": filename,
            "total_empleados": len(empleados),
        },
        empleado_id=int(_current.get("user_id")) if _current and _current.get("user_id") else None,
    )
    return Response(
        content=contenido,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Length": str(len(contenido)),
            "Cache-Control": "no-store",
        },
    )


# ════════════════════════════════════════════════════════════════════════════
# Importación histórica: endpoints deshabilitados (410). Código en importar_historico.py
# se conserva por si se reactiva puntualmente desde el servidor.
# ════════════════════════════════════════════════════════════════════════════

@router.post("/importar-historico/vista-previa/xlsx", tags=["importacion-historica"])
async def importar_historico_vista_previa_xlsx(
    file: UploadFile = File(...),
    empresa_id: Optional[int] = Form(None),
    db: Session = Depends(get_db),
    _su: dict = Depends(require_superuser),
):
    """Importación histórica deshabilitada (antes: vista previa sin escribir en BD)."""
    del file, empresa_id, db  # firma estable para clientes / OpenAPI
    raise HTTPException(
        status_code=status.HTTP_410_GONE,
        detail="La importación histórica de checadas está deshabilitada.",
    )


@router.post("/importar-historico/xlsx", tags=["importacion-historica"])
async def importar_historico_xlsx(
    file: UploadFile = File(...),
    empresa_id: Optional[int] = Form(None),
    db: Session = Depends(get_db),
    _su: dict = Depends(require_superuser),
):
    """Importación histórica deshabilitada."""
    del file, empresa_id, db
    raise HTTPException(
        status_code=status.HTTP_410_GONE,
        detail="La importación histórica de checadas está deshabilitada.",
    )


@router.get("/importar-historico/plantilla", tags=["importacion-historica"])
def descargar_plantilla_importacion(
    _su: dict = Depends(require_superuser_download),
):
    """Plantilla de importación histórica deshabilitada."""
    raise HTTPException(
        status_code=status.HTTP_410_GONE,
        detail="La importación histórica de checadas está deshabilitada.",
    )
