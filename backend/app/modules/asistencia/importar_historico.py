"""
╔══════════════════════════════════════════════════════════════════════╗
║  IMPORTACIÓN HISTÓRICA DE CHECADAS  — USO ÚNICO                     ║
║  Migración inicial de registros LFT desde ZKTeco BioTime.           ║
║  Una vez completada la importación, puede eliminarse junto con:      ║
║    • Este archivo                                                     ║
║    • La ruta en routes.py marcada con # [IMPORTACION-HISTORICA]      ║
║    • El tab "Importar Histórico" en ConfiguracionPage.tsx            ║
╚══════════════════════════════════════════════════════════════════════╝

Formatos soportados (se detectan automáticamente por el encabezado):

① Plantilla simple (recomendado):
  Numero_Empleado | Fecha      | Hora_Entrada | Hora_Salida
  013             | 01/04/2025 | 08:46        | 19:37
  134             | 01/04/2025 | 10:01        |           ← sin salida

  También puede incluir la columna Nombre (opcional). Si Numero_Empleado está vacío,
  se busca por nombre completo (como en el sistema: nombre + apellidos). Si vienen
  ambos, prevalece el número y el nombre solo sirve de referencia.

  En lun–vie (4 checadas requeridas), dos horas se expanden a cuatro marcas
  (entrada, salida comer, regreso comer, salida) interpolando la comida en el intervalo,
  para alinear con el motor de incidencias. Sáb/dom y checada especial de 2 marcas: solo 2.

② BioTime "Historial de registros" (una fila por evento):
  Tiempo               | ID de usuario | Nombre | Sección | ...
  2026-04-16 10:01:20  | 134           | ...    | ...     | ...

  Formatos de archivo: .xlsx (openpyxl) y .xls binario legacy (xlrd).
"""

from __future__ import annotations

import difflib
import secrets
import unicodedata
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta, timezone
from typing import Any, Dict, Optional, Set, Tuple

import openpyxl
from sqlalchemy.orm import Session

from app.core.timezone_utils import ZONE_MEXICO, to_utc
from app.modules.asistencia import models as am
from app.modules.personal import models as pm


# ── Dispositivo virtual para registros importados ───────────────────────────
DISPOSITIVO_IMPORTACION_NOMBRE = "Importación Histórica"
DISPOSITIVO_IMPORTACION_UBICACION = "Importación masiva — registro único LFT"

# Reportes .xls/.xlsx con título y metadatos antes del encabezado real (p. ej. Almacén).
_FILAS_MUESTRA_DETECCION_FORMATO = 60

# ── Umbral para considerar dos eventos como duplicado del mismo toque ────────
UMBRAL_DUPLICADO_SEG = 60

# ── Coincidencia aproximada de nombre (typo en Excel vs BD, ej. Carrizales/Carrizalez)
FUZZY_NOMBRE_MIN_RATIO = 0.89
FUZZY_NOMBRE_MIN_LEN = 12
FUZZY_NOMBRE_GAP_SEGUNDO = 0.025

# ── Columnas esperadas en export BioTime (posición por defecto) ─────────────
COL_TIEMPO = 0          # "Tiempo"
COL_ID_USUARIO = 1      # "ID de usuario"  — número de empleado

_OLE_XLS_SIG = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


def _archivo_es_xls(filename: Optional[str], contenido: bytes) -> bool:
    """True si es libro Excel 97-2003 (.xls OLE), no .xlsx."""
    fn = (filename or "").lower()
    if fn.endswith(".xlsx"):
        return False
    if fn.endswith(".xls"):
        return True
    return len(contenido) >= 8 and contenido[:8] == _OLE_XLS_SIG


def _filas_desde_xls(contenido: bytes) -> list[tuple]:
    """Primera hoja; fechas/horas como datetime (xlrd)."""
    import xlrd
    from xlrd import XL_CELL_BLANK, XL_CELL_DATE, XL_CELL_EMPTY
    from xlrd.xldate import xldate_as_datetime

    book = xlrd.open_workbook(file_contents=contenido)
    sh = book.sheet_by_index(0)
    out: list[tuple] = []
    for rx in range(sh.nrows):
        row: list = []
        for cx in range(sh.ncols):
            ctype = sh.cell_type(rx, cx)
            val = sh.cell_value(rx, cx)
            if ctype == XL_CELL_DATE and val != "" and val is not None:
                try:
                    val = xldate_as_datetime(float(val), book.datemode)
                except Exception:
                    pass
            elif ctype in (XL_CELL_EMPTY, XL_CELL_BLANK):
                val = None
            row.append(val)
        out.append(tuple(row))
    return out


def _cargar_filas_hoja_activa(
    contenido: bytes, filename: Optional[str] = None
) -> list[tuple]:
    """Filas de la hoja activa: .xls (xlrd) o .xlsx (openpyxl)."""
    import io

    if _archivo_es_xls(filename, contenido):
        try:
            return _filas_desde_xls(contenido)
        except Exception as e:
            raise ValueError(f"No se pudo leer el archivo .xls: {e}") from e
    try:
        wb = openpyxl.load_workbook(filename=io.BytesIO(contenido), data_only=True)
        ws = wb.active
        return list(ws.iter_rows(values_only=True))
    except Exception as e:
        raise ValueError(f"No se pudo abrir el archivo Excel: {e}") from e


def _norm_texto_nombre(s: str) -> str:
    """Minúsculas, sin acentos, espacios colapsados (para comparar con BD)."""
    if not s:
        return ""
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return " ".join(s.lower().strip().split())


def _mapa_nombres_a_ids(
    db: Session, empresa_id: Optional[int]
) -> Dict[str, Set[int]]:
    """
    Claves normalizadas: nombre completo y nombre + apellido paterno.
    Valores: ids de empleado (puede haber colisión → varios ids en el mismo set).
    """
    q = db.query(pm.Empleado)
    if empresa_id is not None:
        q = q.filter(pm.Empleado.empresa_id == empresa_id)
    inv: Dict[str, Set[int]] = defaultdict(set)
    for e in q.all():
        full = _norm_texto_nombre(
            f"{e.nombre or ''} {e.apellido_paterno or ''} {e.apellido_materno or ''}"
        )
        short = _norm_texto_nombre(f"{e.nombre or ''} {e.apellido_paterno or ''}")
        reverse_full = _norm_texto_nombre(
            f"{e.apellido_paterno or ''} {e.apellido_materno or ''} {e.nombre or ''}"
        )
        for k in {full, short, reverse_full}:
            if k:
                inv[k].add(e.id)
    return inv


def _fuzzy_match_unico_empleado(
    norm_csv: str,
    db: Session,
    empresa_id: Optional[int],
) -> Tuple[Optional[int], Optional[str]]:
    """
    Un único empleado con nombre completo muy parecido al del archivo
    (SequenceMatcher), para errores leves de escritura respecto a la BD.
    """
    if len(norm_csv) < FUZZY_NOMBRE_MIN_LEN:
        return None, None

    q = db.query(pm.Empleado)
    if empresa_id is not None:
        q = q.filter(pm.Empleado.empresa_id == empresa_id)

    scored: list[tuple[float, int]] = []
    for e in q.all():
        full = _norm_texto_nombre(
            f"{e.nombre or ''} {e.apellido_paterno or ''} {e.apellido_materno or ''}"
        )
        if not full or len(full) < 8:
            continue
        r = difflib.SequenceMatcher(a=norm_csv, b=full).ratio()
        scored.append((r, e.id))

    if not scored:
        return None, None
    scored.sort(key=lambda x: -x[0])
    best_r, best_id = scored[0]
    if best_r < FUZZY_NOMBRE_MIN_RATIO:
        return None, None
    if len(scored) > 1 and scored[0][0] - scored[1][0] < FUZZY_NOMBRE_GAP_SEGUNDO:
        return None, "Varios empleados con nombre muy parecido; indique número de empleado."
    return best_id, None


def _resolver_empleado_por_nombre_importacion(
    norm: str,
    inv: Dict[str, Set[int]],
    db: Session,
    empresa_id: Optional[int],
) -> Tuple[Optional[int], Optional[str]]:
    """Exacto en mapa de nombres; si falla, un solo candidato fuzzy claro."""
    ids = inv.get(norm, set())
    if len(ids) == 1:
        return next(iter(ids)), None
    if len(ids) > 1:
        return None, (
            "Varios empleados coinciden con ese nombre; use número de empleado "
            "o nombre completo (nombre y ambos apellidos)."
        )
    emp_id, amb = _fuzzy_match_unico_empleado(norm, db, empresa_id)
    if emp_id:
        return emp_id, None
    if amb:
        return None, amb
    return None, "No se encontró empleado con ese nombre."


def _parse_hora_hh_mm_seguro(hora_str: str) -> tuple[int, int]:
    """Hora 'HH:MM' o 'H:M' → (h, m) en rango válido."""
    s = str(hora_str).strip().replace(".", ":")
    parts = [p for p in s.split(":") if p != ""]
    if not parts:
        raise ValueError("hora vacía")
    h = int(float(parts[0]))
    m = int(float(parts[1])) if len(parts) > 1 else 0
    return h % 24, min(59, max(0, m))


def _requiere_cuatro_marcajes_jornada(db: Session, empleado_id: int, fecha: date) -> bool:
    """
    Mismo criterio que SyncService / procesar_dia: lun–vie con 4 checadas,
    salvo sáb/dom o checada especial de 2 marcas / jornada reducida L-V.
    """
    from app.modules.asistencia.checada_especial_resolver import obtener_checada_especial_vigente

    dia_semana = fecha.weekday()
    if dia_semana >= 5:
        return False
    emp = db.query(pm.Empleado).filter(pm.Empleado.id == empleado_id).first()
    if emp is None:
        return True
    ce = obtener_checada_especial_vigente(db, empleado_id, fecha)
    if ce and ce.checadas_requeridas == 2 and dia_semana < 5:
        return False
    if ce and ce.jornada_reducida_lv and dia_semana < 5:
        return False
    return True


def _eventos_desde_entrada_salida_csv(
    db: Session,
    empleado_id: int,
    fecha: date,
    hora_entrada: Optional[str],
    hora_salida: Optional[str],
) -> list[tuple[am.TipoChecada, datetime]]:
    """
    Secuencia (tipo, datetime naive México) a persistir.

    Los reportes legacy solo traen entrada y salida de jornada; el motor de asistencia
    cuenta 4 marcas en L–V. Se expande a ENTRADA, SALIDA_COMER, REGRESO_COMER, SALIDA
    interpolando la comida en el intervalo (misma entrada/salida reales que el archivo).
    """
    out: list[tuple[am.TipoChecada, datetime]] = []
    dt_in = dt_out = None
    try:
        if hora_entrada:
            hi, mi = _parse_hora_hh_mm_seguro(hora_entrada)
            dt_in = datetime(fecha.year, fecha.month, fecha.day, hi, mi)
    except (ValueError, IndexError, TypeError):
        dt_in = None
    try:
        if hora_salida:
            hs, ms = _parse_hora_hh_mm_seguro(hora_salida)
            dt_out = datetime(fecha.year, fecha.month, fecha.day, hs, ms)
    except (ValueError, IndexError, TypeError):
        dt_out = None

    if dt_in is None and dt_out is None:
        return []
    if dt_in is not None and dt_out is None:
        return [(am.TipoChecada.ENTRADA, dt_in)]
    if dt_in is None and dt_out is not None:
        return [(am.TipoChecada.SALIDA, dt_out)]

    assert dt_in is not None and dt_out is not None
    if dt_out <= dt_in:
        dt_out = dt_out + timedelta(days=1)

    if not _requiere_cuatro_marcajes_jornada(db, empleado_id, fecha):
        return [
            (am.TipoChecada.ENTRADA, dt_in),
            (am.TipoChecada.SALIDA, dt_out),
        ]

    total_sec = max((dt_out - dt_in).total_seconds(), 4.0)
    t0 = dt_in
    t3 = dt_out
    t1 = dt_in + timedelta(seconds=total_sec / 3.0)
    t2 = dt_in + timedelta(seconds=2.0 * total_sec / 3.0)
    if t1 <= t0:
        t1 = t0 + timedelta(seconds=1)
    if t2 <= t1:
        t2 = t1 + timedelta(seconds=1)
    if t3 <= t2:
        t3 = t2 + timedelta(seconds=1)
    return [
        (am.TipoChecada.ENTRADA, t0),
        (am.TipoChecada.SALIDA_COMER, t1),
        (am.TipoChecada.REGRESO_COMER, t2),
        (am.TipoChecada.SALIDA, t3),
    ]


def _columnas_biotime_desde_encabezado(header_row: tuple) -> Tuple[int, int, Optional[int]]:
    """Índices de columna: tiempo, id usuario, nombre (opcional)."""
    header = [str(v or "").strip().lower() for v in header_row]
    col_tiempo: Optional[int] = None
    col_id: Optional[int] = None
    col_nombre: Optional[int] = None
    for i, h in enumerate(header):
        if h and "tiempo" in h and col_tiempo is None:
            col_tiempo = i
        if h and (
            "id de usuario" in h
            or "id usuario" in h
            or h in ("user id", "userid", "id_usuario", "no. empleado", "no empleado")
        ) and col_id is None:
            col_id = i
        if h and h in (
            "nombre",
            "empleado",
            "nombre completo",
            "nombre_completo",
            "name",
            "colaborador",
            "trabajador",
        ) and col_nombre is None:
            col_nombre = i
    return (
        col_tiempo if col_tiempo is not None else COL_TIEMPO,
        col_id if col_id is not None else COL_ID_USUARIO,
        col_nombre,
    )


# ────────────────────────────────────────────────────────────────────────────
# Estructuras de resultado
# ────────────────────────────────────────────────────────────────────────────

@dataclass
class FilaError:
    fila: int
    id_usuario: str
    timestamp: str
    motivo: str


@dataclass(frozen=True)
class ColsReporteChecador:
    """Índices de columnas en export tipo «estadísticas de asistencia» (ZKTeco / legacy)."""

    idx_emp: int
    idx_fecha: int
    idx_nom: Optional[int]
    idx_combo: Optional[int]
    idx_entrada: Optional[int]
    idx_salida: Optional[int]

    def max_index(self) -> int:
        xs = [self.idx_emp, self.idx_fecha]
        if self.idx_nom is not None:
            xs.append(self.idx_nom)
        if self.idx_combo is not None:
            xs.append(self.idx_combo)
        if self.idx_entrada is not None:
            xs.append(self.idx_entrada)
        if self.idx_salida is not None:
            xs.append(self.idx_salida)
        return max(xs)


@dataclass
class ResultadoImportacion:
    insertadas: int = 0
    omitidas_duplicado: int = 0
    omitidas_sin_registro: int = 0
    errores: list[FilaError] = field(default_factory=list)
    """simple | biotime — rellenado por importar_historico_auto."""
    formato: Optional[str] = None
    """Filas para modal de confirmación (solo dry_run o cuando se solicite)."""
    vista_previa: list[Dict[str, Any]] = field(default_factory=list)

    def to_dict(self) -> dict:
        d: Dict[str, Any] = {
            "insertadas": self.insertadas,
            "omitidas_duplicado": self.omitidas_duplicado,
            "omitidas_sin_registro": self.omitidas_sin_registro,
            "total_errores": len(self.errores),
            "errores": [
                {
                    "fila": e.fila,
                    "id_usuario": e.id_usuario,
                    "timestamp": e.timestamp,
                    "motivo": e.motivo,
                }
                for e in self.errores
            ],
        }
        if self.formato is not None:
            d["formato"] = self.formato
        d["vista_previa"] = self.vista_previa
        return d


# ────────────────────────────────────────────────────────────────────────────
# Dispositivo virtual
# ────────────────────────────────────────────────────────────────────────────

def _obtener_o_crear_dispositivo(db: Session, dry_run: bool = False) -> Optional[am.Dispositivo]:
    if dry_run:
        return None
    disp = (
        db.query(am.Dispositivo)
        .filter(am.Dispositivo.nombre == DISPOSITIVO_IMPORTACION_NOMBRE)
        .first()
    )
    if disp:
        return disp

    disp = am.Dispositivo(
        nombre=DISPOSITIVO_IMPORTACION_NOMBRE,
        ubicacion=DISPOSITIVO_IMPORTACION_UBICACION,
        api_key=f"importacion-historica-{secrets.token_hex(16)}",
        activo=False,   # inactivo: no aparece en listas normales
    )
    db.add(disp)
    db.flush()
    return disp


def _meta_empleado(db: Session, emp_id: int, cache: Dict[int, Tuple[str, str]]) -> Tuple[str, str]:
    """(numero_empleado, nombre_completo) desde BD."""
    if emp_id in cache:
        return cache[emp_id]
    e = db.query(pm.Empleado).filter(pm.Empleado.id == emp_id).first()
    if not e:
        cache[emp_id] = ("", "")
        return cache[emp_id]
    nombre = " ".join(
        x for x in [e.nombre, e.apellido_paterno, e.apellido_materno] if x
    ).strip()
    cache[emp_id] = (e.numero_empleado or "", nombre)
    return cache[emp_id]


# ────────────────────────────────────────────────────────────────────────────
# Lógica de clasificación de tipo por posición en el día
# ────────────────────────────────────────────────────────────────────────────

_MAPA_TIPOS: dict[int, list[am.TipoChecada]] = {
    1: [am.TipoChecada.ENTRADA],
    2: [am.TipoChecada.ENTRADA, am.TipoChecada.SALIDA],
    3: [am.TipoChecada.ENTRADA, am.TipoChecada.SALIDA_COMER, am.TipoChecada.SALIDA],
    4: [am.TipoChecada.ENTRADA, am.TipoChecada.SALIDA_COMER, am.TipoChecada.REGRESO_COMER, am.TipoChecada.SALIDA],
}


def _deduplicar(eventos: list[datetime]) -> list[datetime]:
    """
    Elimina marcas consecutivas con menos de UMBRAL_DUPLICADO_SEG segundos
    entre sí (el dedo checó dos veces seguidas sin querer).
    """
    if not eventos:
        return []
    ordenados = sorted(eventos)
    resultado = [ordenados[0]]
    for evt in ordenados[1:]:
        if (evt - resultado[-1]).total_seconds() >= UMBRAL_DUPLICADO_SEG:
            resultado.append(evt)
    return resultado


def _asignar_tipos(eventos: list[datetime]) -> list[tuple[datetime, am.TipoChecada]]:
    """Devuelve la lista (timestamp, TipoChecada) para los eventos deduplicados del día."""
    limpios = _deduplicar(eventos)
    n = len(limpios)

    if n == 0:
        return []

    # Si después de deduplicar quedan más de 4, usar los 4 más representativos:
    # el primero, el último, y los dos intermedios equidistantes.
    if n > 4:
        indices = [0, n // 3, 2 * n // 3, n - 1]
        limpios = [limpios[i] for i in indices]
        n = 4

    tipos = _MAPA_TIPOS.get(n)
    if tipos is None:
        tipos = _MAPA_TIPOS[4]   # fallback

    return list(zip(limpios, tipos))


# ────────────────────────────────────────────────────────────────────────────
# Parseo del XLSX
# ────────────────────────────────────────────────────────────────────────────

def _parse_timestamp(valor) -> Optional[datetime]:
    """
    Convierte el valor de la celda 'Tiempo' a UTC.
    BioTime exporta en hora local México (naive); usamos to_utc() igual que
    el servicio biométrico para ser consistentes con el resto del sistema.
    """
    if valor is None:
        return None
    if isinstance(valor, datetime):
        dt_naive = valor.replace(tzinfo=None)   # asegurar naive antes de to_utc
    else:
        s = str(valor).strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%d/%m/%Y %H:%M:%S"):
            try:
                dt_naive = datetime.strptime(s, fmt)
                break
            except ValueError:
                pass
        else:
            return None

    # Igual que iclock_routes: "El dispositivo envía hora local (México); guardamos en UTC"
    return to_utc(dt_naive)


# ────────────────────────────────────────────────────────────────────────────
# Función principal
# ────────────────────────────────────────────────────────────────────────────

def importar_checadas_desde_xlsx_biotime(
    contenido: bytes,
    db: Session,
    empresa_id: Optional[int] = None,
    dry_run: bool = False,
    filename: Optional[str] = None,
) -> ResultadoImportacion:
    """
    Procesa el XLSX de ZKTeco BioTime y crea registros en la tabla `asistencias`.

    Parámetros:
        contenido  : bytes del archivo XLSX.
        db         : sesión de base de datos (sin commit; el caller hace commit).
        empresa_id : si se indica, valida que el empleado pertenezca a esa empresa.
        dry_run    : si True, no escribe en BD; llena `vista_previa` y hace rollback en la ruta.
        filename   : nombre del archivo subido (para distinguir .xls vs .xlsx).
    """
    resultado = ResultadoImportacion()

    try:
        filas = _cargar_filas_hoja_activa(contenido, filename)
    except ValueError:
        raise
    if not filas:
        raise ValueError("El archivo no tiene filas")

    dispositivo = _obtener_o_crear_dispositivo(db, dry_run=dry_run)

    header_row = filas[0]
    col_tiempo, col_id, col_nombre = _columnas_biotime_desde_encabezado(header_row)

    # ── Leer y agrupar eventos: clave ("id"|"name", identificador, fecha) ────
    # Cada evento guarda (fila, timestamp, texto_nombre_celda) para fallback por nombre.
    GroupKey = Tuple[str, str, object]
    grupos: dict[GroupKey, list[tuple[int, datetime, str]]] = defaultdict(list)
    errores_lectura: list[FilaError] = []

    for fila_idx, row in enumerate(filas[1:], start=2):
        if row is None or all(v is None or str(v).strip() == "" for v in row):
            continue

        ts = _parse_timestamp(row[col_tiempo] if len(row) > col_tiempo else None)
        if ts is None:
            errores_lectura.append(FilaError(
                fila=fila_idx,
                id_usuario=str(row[col_id] if len(row) > col_id else "?"),
                timestamp=str(row[col_tiempo] if len(row) > col_tiempo else ""),
                motivo="Timestamp inválido o vacío",
            ))
            continue

        nombre_celda = ""
        if col_nombre is not None and len(row) > col_nombre and row[col_nombre]:
            nombre_celda = str(row[col_nombre]).strip()

        id_raw = row[col_id] if len(row) > col_id else None
        fecha = ts.astimezone(ZONE_MEXICO).date()

        if id_raw is not None and str(id_raw).strip() != "":
            id_usuario = str(int(id_raw)) if isinstance(id_raw, float) else str(id_raw).strip()
            key: GroupKey = ("id", id_usuario, fecha)
        elif nombre_celda:
            key = ("name", _norm_texto_nombre(nombre_celda), fecha)
        else:
            errores_lectura.append(FilaError(
                fila=fila_idx,
                id_usuario="",
                timestamp=ts.isoformat(),
                motivo="Falta ID de usuario y nombre (indique al menos uno)",
            ))
            continue

        grupos[key].append((fila_idx, ts, nombre_celda))

    resultado.errores.extend(errores_lectura)

    nombre_map = _mapa_nombres_a_ids(db, empresa_id)
    nombre_map_global: Optional[Dict[str, Set[int]]] = None
    meta_cache: Dict[int, Tuple[str, str]] = {}

    # ── Cache empleados: numero_empleado → empleado_id ────────────────────────
    emp_cache: dict[str, Optional[int]] = {}

    def _buscar_empleado(num: str) -> Optional[int]:
        if num in emp_cache:
            return emp_cache[num]
        q = db.query(pm.Empleado.id).filter(pm.Empleado.numero_empleado == num)
        if empresa_id:
            q = q.filter(pm.Empleado.empresa_id == empresa_id)
        row_db = q.first()
        emp_cache[num] = row_db[0] if row_db else None
        return emp_cache[num]

    # ── Cache de existentes: (empleado_id, date_iso, tipo_value) ─────────────
    existing_cache: set[tuple[int, str, str]] = set()
    precargados: set[int] = set()

    def _precargar_existentes(emp_id: int):
        if emp_id in precargados:
            return
        rows = db.query(am.Asistencia.timestamp, am.Asistencia.tipo).filter(
            am.Asistencia.empleado_id == emp_id
        ).all()
        for r in rows:
            ts_utc = r.timestamp if r.timestamp.tzinfo else r.timestamp.replace(tzinfo=timezone.utc)
            d = ts_utc.astimezone(ZONE_MEXICO).date().isoformat()
            existing_cache.add((emp_id, d, r.tipo.value))
        precargados.add(emp_id)

    def _ya_existe(emp_id: int, fecha_iso: str, tipo: am.TipoChecada) -> bool:
        return (emp_id, fecha_iso, tipo.value) in existing_cache

    def _marcar(emp_id: int, fecha_iso: str, tipo: am.TipoChecada):
        existing_cache.add((emp_id, fecha_iso, tipo.value))

    # ── Procesar cada grupo (empleado, día) ───────────────────────────────────
    for key, eventos_fila in grupos.items():
        timestamps = [ts for _, ts, _ in eventos_fila]
        primera_fila = eventos_fila[0][0]
        nombre_hint = next((n for _, _, n in eventos_fila if n), "")
        fecha = key[2]
        fecha_iso = fecha.isoformat()

        if not timestamps:
            resultado.omitidas_sin_registro += 1
            continue

        emp_id: Optional[int] = None
        err_nom: Optional[str] = None
        if key[0] == "id":
            emp_id = _buscar_empleado(key[1])
            if emp_id is None and nombre_hint:
                emp_id, err_nom = _resolver_empleado_por_nombre_importacion(
                    _norm_texto_nombre(nombre_hint), nombre_map, db, empresa_id
                )
        else:
            emp_id, err_nom = _resolver_empleado_por_nombre_importacion(
                key[1], nombre_map, db, empresa_id
            )

        if emp_id is None:
            if key[0] == "id":
                if nombre_hint and err_nom:
                    motivo = (
                        f"Número '{key[1]}' no encontrado; por nombre ({nombre_hint}): {err_nom}"
                    )
                elif nombre_hint:
                    motivo = (
                        f"Número '{key[1]}' no encontrado; el nombre de la fila no coincide "
                        "con ningún empleado."
                    )
                else:
                    motivo = f"Empleado con número '{key[1]}' no encontrado en la base de datos"
                label = key[1]
            else:
                motivo = err_nom or "No se encontró empleado con ese nombre."
                label = nombre_hint or key[1]
            resultado.errores.append(FilaError(
                fila=primera_fila,
                id_usuario=label,
                timestamp=fecha_iso,
                motivo=motivo,
            ))
            continue

        _precargar_existentes(emp_id)

        pares = _asignar_tipos(timestamps)
        numero_arch = key[1] if key[0] == "id" else None

        for ts_evento, tipo in pares:
            ts_mx = ts_evento.astimezone(ZONE_MEXICO)
            hora_str = ts_mx.strftime("%H:%M")
            num_bd, nom_bd = _meta_empleado(db, emp_id, meta_cache)

            if _ya_existe(emp_id, fecha_iso, tipo):
                resultado.omitidas_duplicado += 1
                if dry_run:
                    resultado.vista_previa.append(
                        {
                            "fila": primera_fila,
                            "numero_archivo": numero_arch,
                            "nombre_archivo": nombre_hint or None,
                            "empleado_id": emp_id,
                            "numero_bd": num_bd,
                            "nombre_bd": nom_bd,
                            "fecha": fecha_iso,
                            "tipo": tipo.value,
                            "hora": hora_str,
                            "fecha_hora_local": f"{fecha_iso} {hora_str}",
                            "estado": "duplicado",
                            "motivo": None,
                        }
                    )
                continue

            if dry_run:
                resultado.vista_previa.append(
                    {
                        "fila": primera_fila,
                        "numero_archivo": numero_arch,
                        "nombre_archivo": nombre_hint or None,
                        "empleado_id": emp_id,
                        "numero_bd": num_bd,
                        "nombre_bd": nom_bd,
                        "fecha": fecha_iso,
                        "tipo": tipo.value,
                        "hora": hora_str,
                        "fecha_hora_local": f"{fecha_iso} {hora_str}",
                        "estado": "insertar",
                        "motivo": None,
                    }
                )
                _marcar(emp_id, fecha_iso, tipo)
                resultado.insertadas += 1
            else:
                assert dispositivo is not None
                nueva = am.Asistencia(
                    empleado_id=emp_id,
                    dispositivo_id=dispositivo.id,
                    timestamp=ts_evento,
                    tipo=tipo,
                    es_tiempo_extra=False,
                    sincronizado=True,
                )
                db.add(nueva)
                _marcar(emp_id, fecha_iso, tipo)
                resultado.insertadas += 1

    return resultado


# ════════════════════════════════════════════════════════════════════════════
# FORMATO SIMPLE — Plantilla manual
# ════════════════════════════════════════════════════════════════════════════
# Columnas requeridas (primera fila = encabezado):
#   Numero_Empleado (opcional si hay Nombre) | Nombre (opcional si hay número)
#   | Fecha | Hora_Entrada | Hora_Salida
#
# Reglas:
#   • Numero_Empleado: texto tal como está en BD (ej. "013", "134")
#   • Nombre: nombre completo como en el sistema (nombre y apellidos); sin acentos no importa
#   • Fecha: DD/MM/YYYY  o  YYYY-MM-DD
#   • Hora_Entrada / Hora_Salida: HH:MM  (vacío = sin ese evento)
# ════════════════════════════════════════════════════════════════════════════

def _parse_fecha_simple(valor) -> Optional[date]:
    from datetime import date as _date
    if valor is None:
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, _date):
        return valor
    s = str(valor).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def _parse_hora_simple(valor) -> Optional[str]:
    """
    Acepta:
      • objetos datetime.time y datetime.datetime (Excel devuelve estos cuando
        la celda está formateada como hora).
      • número float de Excel (fracción de día, 0.0 a 1.0) cuando openpyxl
        no reconoce el formato como hora.
      • strings "HH:MM" o "HH:MM:SS" (con o sin punto en vez de dos puntos).
    Devuelve "HH:MM" o None si no aplica (vacío, cero, "días libres").
    """
    import re as _re
    from datetime import time as _time
    if valor is None:
        return None

    if isinstance(valor, datetime):
        h, mi = valor.hour, valor.minute
        if h == 0 and mi == 0:
            return None
        return f"{h:02d}:{mi:02d}"

    if isinstance(valor, _time):
        h, mi = valor.hour, valor.minute
        if h == 0 and mi == 0:
            return None
        return f"{h:02d}:{mi:02d}"

    if isinstance(valor, (int, float)) and not isinstance(valor, bool):
        if valor == 0:
            return None
        if 0 < valor < 1:
            total_min = int(round(valor * 24 * 60))
            h, mi = total_min // 60, total_min % 60
            if h == 0 and mi == 0:
                return None
            return f"{h:02d}:{mi:02d}"
        return None

    s = str(valor).strip().replace('.', ':')
    if not s or s.lower() in ('none', 'nan', '-', '', '0', '00:00', '00:00:00'):
        return None
    m = _re.match(r'^(\d{1,2}):(\d{2})(?::(\d{2}))?$', s)
    if not m:
        return None
    h, mi = int(m.group(1)), int(m.group(2))
    if not (0 <= h <= 23 and 0 <= mi <= 59):
        return None
    if h == 0 and mi == 0:
        return None
    return f"{h:02d}:{mi:02d}"


def _parse_fecha_reporte_checador(valor) -> Optional[date]:
    """Fecha de reportes legacy del checador (serial Excel, datetime o string)."""
    from datetime import date as _date, timedelta as _timedelta
    if valor is None:
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, _date):
        return valor
    if isinstance(valor, (int, float)) and not isinstance(valor, bool):
        # Serial Excel típico (ej. 45870 -> 2025-08-01)
        serial = float(valor)
        if serial <= 0:
            return None
        try:
            base = datetime(1899, 12, 30)
            return (base + _timedelta(days=serial)).date()
        except Exception:
            return None
    s = str(valor).strip()
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def _parse_ingreso_salida_reporte_checador(valor) -> tuple[Optional[str], Optional[str]]:
    """
    Parsea celdas tipo '08:46-18:46' o '0.00-0.00' (sin registro).
    Devuelve (hora_entrada, hora_salida) en formato HH:MM.
    """
    from datetime import time as _time

    if valor is None:
        return None, None
    if isinstance(valor, datetime):
        if valor.hour == 0 and valor.minute == 0:
            return None, None
        return f"{valor.hour:02d}:{valor.minute:02d}", None
    if isinstance(valor, _time):
        if valor.hour == 0 and valor.minute == 0:
            return None, None
        return f"{valor.hour:02d}:{valor.minute:02d}", None
    if isinstance(valor, (int, float)) and not isinstance(valor, bool):
        h = _parse_hora_simple(valor)
        return h, None

    s = str(valor).strip()
    if not s or s.lower() in ("none", "nan", "-", ""):
        return None, None
    s_norm = s.replace("–", "-").replace("—", "-").replace("−", "-")
    low = s_norm.lower()

    split_at: Optional[tuple[str, int]] = None
    for sep in (" a ", " al ", " hasta ", " to ", " - ", "-"):
        j = low.find(sep)
        if j >= 0:
            split_at = (sep, j)
            break
    if split_at is None:
        h = _parse_hora_simple(s_norm)
        return h, None

    sep, j = split_at
    p1 = s_norm[:j].strip()
    p2 = s_norm[j + len(sep) :].strip()
    h1 = _parse_hora_simple(p1)
    h2 = _parse_hora_simple(p2)
    return h1, h2


def _indices_columnas_reporte_checador(filas: list[tuple]) -> ColsReporteChecador:
    """
    Localiza columnas a partir de la fila de encabezado del reporte.
    Soporta «Ingreso - Salida» en una sola columna o Entrada / Salida en columnas aparte.
    """
    default = ColsReporteChecador(0, 1, None, 2, None, None)
    lim = min(len(filas), 50)
    for row in filas[:lim]:
        if not row or len(row) < 3:
            continue
        lower = [str(v or "").strip().lower() for v in row]
        tiene_emp = any(
            ("nº de empleado" in v)
            or ("n° de empleado" in v)
            or ("no de empleado" in v)
            or ("no. de empleado" in v)
            for v in lower
        )
        if not tiene_emp:
            continue
        tiene_fecha = any("fecha" in v for v in lower)
        if not tiene_fecha:
            continue

        idx_emp = idx_fecha = idx_nom = None
        idx_combo = idx_combo_fallback = idx_entrada = idx_salida = None
        for i, v in enumerate(lower):
            if v in ("nº de empleado", "n° de empleado", "no de empleado", "no. de empleado"):
                idx_emp = i
            elif ("nº" in v or "n°" in v) and "empleado" in v:
                idx_emp = i
            elif v == "fecha":
                idx_fecha = i
            elif idx_fecha is None and "fecha" in v and len(v) < 36:
                idx_fecha = i
            elif v in ("nombre completo", "nombre", "colaborador"):
                idx_nom = i
            elif ("ingreso" in v and "salida" in v) or ("entrada" in v and "salida" in v):
                # Evitar tomar primero columnas de horas extra.
                if "hora extra" in v or "horas extra" in v:
                    if idx_combo_fallback is None:
                        idx_combo_fallback = i
                elif idx_combo is None:
                    idx_combo = i
            elif v in (
                "entrada",
                "ingreso",
                "hora entrada",
                "hora_entrada",
                "inicio",
            ) or (v.startswith("hora") and "entrada" in v):
                idx_entrada = i
            elif v in ("salida", "hora salida", "hora_salida", "fin") or (
                v.startswith("hora") and "salida" in v
            ):
                idx_salida = i

        if idx_combo is None:
            idx_combo = idx_combo_fallback
        if idx_combo is not None:
            idx_entrada = idx_salida = None

        tiene_horario = (
            idx_combo is not None or idx_entrada is not None or idx_salida is not None
        )
        if not tiene_horario or idx_emp is None or idx_fecha is None:
            continue
        return ColsReporteChecador(
            idx_emp, idx_fecha, idx_nom, idx_combo, idx_entrada, idx_salida
        )
    return default


def _horas_fila_reporte_checador(
    row: tuple, cols: ColsReporteChecador
) -> tuple[Optional[str], Optional[str]]:
    if cols.idx_combo is not None and cols.idx_combo < len(row):
        h1, h2 = _parse_ingreso_salida_reporte_checador(row[cols.idx_combo])
        if h1 or h2:
            return h1, h2
    h1 = h2 = None
    if cols.idx_entrada is not None and cols.idx_entrada < len(row):
        h1 = _parse_hora_simple(row[cols.idx_entrada])
    if cols.idx_salida is not None and cols.idx_salida < len(row):
        h2 = _parse_hora_simple(row[cols.idx_salida])
    if h1 or h2:
        return h1, h2
    if cols.idx_combo is not None and cols.idx_combo < len(row):
        return _parse_ingreso_salida_reporte_checador(row[cols.idx_combo])
    return None, None


def _extraer_nombre_fila_nombre_completo(row: tuple) -> Optional[str]:
    """Texto de persona en filas tipo 'Nombre Completo' | … | 'JUAN PÉREZ …'."""
    for c in range(1, len(row)):
        t = str(row[c] or "").strip()
        if len(t) < 3:
            continue
        if _parse_fecha_reporte_checador(row[c]) is not None:
            continue
        try:
            float(t)
            continue
        except ValueError:
            pass
        return t
    return None


def importar_checadas_desde_reporte_checador(
    contenido: bytes,
    db: Session,
    empresa_id: Optional[int] = None,
    dry_run: bool = False,
    filename: Optional[str] = None,
) -> ResultadoImportacion:
    """
    Importa desde reportes legacy por empleado:
      Nº de Empleado | Fecha | Ingreso - Salida | ...

    Si el archivo trae fila «Nombre Completo» antes del bloque o columna «Nombre», se usa
    **primero la coincidencia por nombre** con la BD (número del reporte puede no coincidir).
    Si no hay nombre o no resuelve, se usa el número de empleado del archivo.
    """
    resultado = ResultadoImportacion()
    try:
        filas = _cargar_filas_hoja_activa(contenido, filename)
    except ValueError:
        raise
    if not filas:
        raise ValueError("El archivo no tiene filas")

    cols = _indices_columnas_reporte_checador(filas)
    max_idx = cols.max_index()
    nombre_map = _mapa_nombres_a_ids(db, empresa_id)
    nombre_map_global: Optional[Dict[str, Set[int]]] = None
    nombre_contexto: Optional[str] = None

    dispositivo = _obtener_o_crear_dispositivo(db, dry_run=dry_run)
    meta_cache: Dict[int, Tuple[str, str]] = {}
    emp_cache: dict[str, Optional[int]] = {}
    existing_cache: set[tuple[int, str, str]] = set()
    precargados: set[int] = set()

    def _buscar_empleado(num: str) -> Optional[int]:
        if num in emp_cache:
            return emp_cache[num]
        q = db.query(pm.Empleado.id).filter(pm.Empleado.numero_empleado == num)
        if empresa_id:
            q = q.filter(pm.Empleado.empresa_id == empresa_id)
        r = q.first()
        emp_cache[num] = r[0] if r else None
        return emp_cache[num]

    def _buscar_empleado_global(num: str) -> Optional[int]:
        k = f"__global__:{num}"
        if k in emp_cache:
            return emp_cache[k]
        r = db.query(pm.Empleado.id).filter(pm.Empleado.numero_empleado == num).first()
        emp_cache[k] = r[0] if r else None
        return emp_cache[k]

    def _precargar(emp_id: int):
        if emp_id in precargados:
            return
        rows = db.query(am.Asistencia.timestamp, am.Asistencia.tipo).filter(
            am.Asistencia.empleado_id == emp_id
        ).all()
        for r in rows:
            ts_utc = r.timestamp if r.timestamp.tzinfo else r.timestamp.replace(tzinfo=timezone.utc)
            d = ts_utc.astimezone(ZONE_MEXICO).date().isoformat()
            existing_cache.add((emp_id, d, r.tipo.value))
        precargados.add(emp_id)

    def _ya_existe(emp_id: int, fecha_iso: str, tipo: am.TipoChecada) -> bool:
        return (emp_id, fecha_iso, tipo.value) in existing_cache

    def _marcar(emp_id: int, fecha_iso: str, tipo: am.TipoChecada):
        existing_cache.add((emp_id, fecha_iso, tipo.value))

    for row_idx, row in enumerate(filas, start=1):
        if not row or len(row) < 3:
            continue
        if all(v is None or str(v).strip() == "" for v in row):
            continue

        col0 = str(row[0] or "").strip().lower()
        if col0 in ("nombre completo",) or col0.startswith("nombre completo"):
            ext = _extraer_nombre_fila_nombre_completo(row)
            if ext:
                nombre_contexto = ext
            continue

        if max_idx >= len(row):
            continue

        num_raw = (
            str(row[cols.idx_emp] or "").strip() if cols.idx_emp < len(row) else ""
        )
        low_emp = num_raw.lower()
        if low_emp in ("nº de empleado", "n° de empleado", "no de empleado", "no. de empleado"):
            continue

        fecha = _parse_fecha_reporte_checador(
            row[cols.idx_fecha] if cols.idx_fecha < len(row) else None
        )
        if fecha is None:
            continue

        hora_entrada, hora_salida = _horas_fila_reporte_checador(row, cols)
        if not hora_entrada and not hora_salida:
            resultado.omitidas_sin_registro += 1
            if dry_run:
                nom_hint = nombre_contexto
                if cols.idx_nom is not None and cols.idx_nom < len(row):
                    t = str(row[cols.idx_nom] or "").strip()
                    if t:
                        nom_hint = t
                celdas_h = []
                if cols.idx_combo is not None and cols.idx_combo < len(row):
                    celdas_h.append(row[cols.idx_combo])
                if cols.idx_entrada is not None and cols.idx_entrada < len(row):
                    celdas_h.append(row[cols.idx_entrada])
                if cols.idx_salida is not None and cols.idx_salida < len(row):
                    celdas_h.append(row[cols.idx_salida])
                hint = " · ".join(str(c) for c in celdas_h if c is not None and str(c).strip())
                resultado.vista_previa.append(
                    {
                        "fila": row_idx,
                        "numero_archivo": num_raw or None,
                        "nombre_archivo": nom_hint,
                        "empleado_id": None,
                        "numero_bd": None,
                        "nombre_bd": None,
                        "fecha": fecha.isoformat(),
                        "tipo": None,
                        "hora": None,
                        "fecha_hora_local": None,
                        "estado": "sin_registro_dia_libre",
                        "motivo": (hint[:120] if hint else "Celdas de horario vacías o no reconocidas"),
                    }
                )
            continue

        nombre_celda: Optional[str] = None
        if cols.idx_nom is not None and cols.idx_nom < len(row):
            tnom = str(row[cols.idx_nom] or "").strip()
            if tnom:
                nombre_celda = tnom
        nombre_candidato = nombre_celda or nombre_contexto

        try:
            num_emp = str(int(float(num_raw))).zfill(3) if num_raw else ""
        except ValueError:
            num_emp = num_raw

        emp_id: Optional[int] = None
        nombre_motivo: Optional[str] = None
        if nombre_candidato:
            norm = _norm_texto_nombre(nombre_candidato)
            if norm:
                emp_id, nombre_motivo = _resolver_empleado_por_nombre_importacion(
                    norm, nombre_map, db, empresa_id
                )
                # Si viene filtrado por empresa y no resolvió, intentar global.
                if emp_id is None and empresa_id is not None:
                    if nombre_map_global is None:
                        nombre_map_global = _mapa_nombres_a_ids(db, None)
                    emp_id_g, _ = _resolver_empleado_por_nombre_importacion(
                        norm, nombre_map_global, db, None
                    )
                    if emp_id_g is not None:
                        emp_id = emp_id_g
                        nombre_motivo = None

        if emp_id is None and num_raw:
            emp_id = _buscar_empleado(num_emp)
            if emp_id is None and num_raw.replace(".", "", 1).isdigit():
                emp_id = _buscar_empleado(str(int(float(num_raw))))
            if emp_id is None and empresa_id is not None:
                emp_id = _buscar_empleado_global(num_emp)
                if (
                    emp_id is None
                    and num_raw.replace(".", "", 1).isdigit()
                ):
                    emp_id = _buscar_empleado_global(str(int(float(num_raw))))

        if emp_id is None:
            partes: list[str] = []
            if nombre_candidato and nombre_motivo:
                partes.append(f"Nombre ({nombre_candidato}): {nombre_motivo}")
            if num_raw:
                partes.append(
                    f"Número '{num_emp or num_raw}' no encontrado en la base de datos"
                )
            if not partes:
                partes.append("Sin número ni nombre para identificar al empleado")
            resultado.errores.append(FilaError(
                fila=row_idx,
                id_usuario=num_emp or nombre_candidato or "?",
                timestamp=fecha.isoformat(),
                motivo="; ".join(partes),
            ))
            continue

        _precargar(emp_id)
        fecha_iso = fecha.isoformat()
        num_bd, nom_bd = _meta_empleado(db, emp_id, meta_cache)

        eventos_dia = _eventos_desde_entrada_salida_csv(
            db, emp_id, fecha, hora_entrada, hora_salida
        )
        for tipo, dt_naive in eventos_dia:
            hora_str = dt_naive.strftime("%H:%M")
            if _ya_existe(emp_id, fecha_iso, tipo):
                resultado.omitidas_duplicado += 1
                if dry_run:
                    resultado.vista_previa.append(
                        {
                            "fila": row_idx,
                            "numero_archivo": num_emp or num_raw or None,
                            "nombre_archivo": nombre_candidato,
                            "empleado_id": emp_id,
                            "numero_bd": num_bd,
                            "nombre_bd": nom_bd,
                            "fecha": fecha_iso,
                            "tipo": tipo.value,
                            "hora": hora_str,
                            "fecha_hora_local": f"{fecha_iso} {hora_str}",
                            "estado": "duplicado",
                            "motivo": None,
                        }
                    )
                continue

            ts = to_utc(dt_naive)

            if dry_run:
                resultado.vista_previa.append(
                    {
                        "fila": row_idx,
                        "numero_archivo": num_emp or num_raw or None,
                        "nombre_archivo": nombre_candidato,
                        "empleado_id": emp_id,
                        "numero_bd": num_bd,
                        "nombre_bd": nom_bd,
                        "fecha": fecha_iso,
                        "tipo": tipo.value,
                        "hora": hora_str,
                        "fecha_hora_local": f"{fecha_iso} {hora_str}",
                        "estado": "insertar",
                        "motivo": None,
                    }
                )
                _marcar(emp_id, fecha_iso, tipo)
                resultado.insertadas += 1
            else:
                assert dispositivo is not None
                db.add(
                    am.Asistencia(
                        empleado_id=emp_id,
                        dispositivo_id=dispositivo.id,
                        timestamp=ts,
                        tipo=tipo,
                        es_tiempo_extra=False,
                        sincronizado=True,
                    )
                )
                _marcar(emp_id, fecha_iso, tipo)
                resultado.insertadas += 1
    return resultado


def importar_checadas_desde_plantilla_simple(
    contenido: bytes,
    db: Session,
    empresa_id: Optional[int] = None,
    dry_run: bool = False,
    filename: Optional[str] = None,
) -> ResultadoImportacion:
    """
    Importa desde la plantilla simple:
        Numero_Empleado | Fecha | Hora_Entrada | Hora_Salida
    """
    resultado = ResultadoImportacion()

    try:
        filas = _cargar_filas_hoja_activa(contenido, filename)
    except ValueError:
        raise
    if not filas:
        raise ValueError("El archivo no tiene filas")

    dispositivo = _obtener_o_crear_dispositivo(db, dry_run=dry_run)

    # ── Localizar columnas por nombre ─────────────────────────────────────
    col_map: dict[str, int] = {}
    header_row_idx = None

    for row_idx, row in enumerate(filas[:_FILAS_MUESTRA_DETECCION_FORMATO], start=1):
        if not row:
            continue
        nombres = [str(v or '').strip().lower() for v in row]
        has_num = any(
            'numero_empleado' in n or 'numero empleado' in n or '# emp' in n or n == 'num_emp'
            for n in nombres
        )
        has_nombre_header = any(
            n in (
                'nombre', 'empleado', 'nombre completo', 'nombre_completo', 'name',
                'colaborador', 'trabajador', 'nombre empleado', 'nombre_empleado',
            )
            for n in nombres
        )
        has_fecha = any(n in ('fecha', 'date') for n in nombres)
        if (has_num or has_nombre_header) and has_fecha:
            header_row_idx = row_idx
            for col_idx, name in enumerate(nombres):
                col_map[name] = col_idx
            break

    if header_row_idx is None:
        raise ValueError(
            "No se encontró el encabezado. Incluya la columna Fecha y al menos una de: "
            "Numero_Empleado o Nombre (más Hora_Entrada / Hora_Salida)."
        )

    def _col(*opciones) -> Optional[int]:
        for o in opciones:
            if o in col_map:
                return col_map[o]
        return None

    idx_num   = _col('numero_empleado', 'numero empleado', '# emp', 'num_emp')
    idx_nombre = _col(
        'nombre', 'empleado', 'nombre completo', 'nombre_completo', 'name',
        'colaborador', 'trabajador', 'nombre empleado', 'nombre_empleado',
    )
    idx_fecha = _col('fecha', 'date')
    idx_ent   = _col('hora_entrada', 'hora entrada', 'entrada', 'ingreso')
    idx_sal   = _col('hora_salida',  'hora salida',  'salida')

    idx_combo_plantilla: Optional[int] = None
    for k, ci in col_map.items():
        if "ingreso" in k and "salida" in k:
            idx_combo_plantilla = ci
            break

    if idx_fecha is None or (idx_num is None and idx_nombre is None):
        raise ValueError(
            "Columnas requeridas: Fecha y al menos una de Numero_Empleado o Nombre"
        )

    nombre_map = _mapa_nombres_a_ids(db, empresa_id)
    meta_cache: Dict[int, Tuple[str, str]] = {}

    # ── Caches ───────────────────────────────────────────────────────────
    emp_cache: dict[str, Optional[int]] = {}

    def _buscar_empleado(num: str) -> Optional[int]:
        if num in emp_cache:
            return emp_cache[num]
        q = db.query(pm.Empleado.id).filter(pm.Empleado.numero_empleado == num)
        if empresa_id:
            q = q.filter(pm.Empleado.empresa_id == empresa_id)
        r = q.first()
        emp_cache[num] = r[0] if r else None
        return emp_cache[num]

    existing_cache: set[tuple[int, str, str]] = set()
    precargados: set[int] = set()

    def _precargar(emp_id: int):
        if emp_id in precargados:
            return
        rows = db.query(am.Asistencia.timestamp, am.Asistencia.tipo).filter(
            am.Asistencia.empleado_id == emp_id
        ).all()
        for r in rows:
            ts_utc = r.timestamp if r.timestamp.tzinfo else r.timestamp.replace(tzinfo=timezone.utc)
            d = ts_utc.astimezone(ZONE_MEXICO).date().isoformat()
            existing_cache.add((emp_id, d, r.tipo.value))
        precargados.add(emp_id)

    def _ya_existe_s(emp_id: int, fecha_iso: str, tipo: am.TipoChecada) -> bool:
        return (emp_id, fecha_iso, tipo.value) in existing_cache

    def _marcar_s(emp_id: int, fecha_iso: str, tipo: am.TipoChecada):
        existing_cache.add((emp_id, fecha_iso, tipo.value))

    # ── Procesar filas ────────────────────────────────────────────────────
    all_rows = filas

    for row_offset, row in enumerate(all_rows[header_row_idx:], start=header_row_idx + 1):
        if not row or all(v is None or str(v).strip() == '' for v in row):
            continue

        def _get(idx):
            return row[idx] if idx is not None and idx < len(row) else None

        num_raw = str(_get(idx_num) or '').strip() if idx_num is not None else ''
        nombre_raw = str(_get(idx_nombre) or '').strip() if idx_nombre is not None else ''

        if (not num_raw or num_raw.lower() == 'none') and (
            not nombre_raw or nombre_raw.lower() == 'none'
        ):
            continue

        num_emp: Optional[str] = None
        if num_raw and num_raw.lower() != 'none':
            try:
                num_emp = str(int(float(num_raw))).zfill(3)
            except ValueError:
                num_emp = num_raw

        nombre_norm = (
            _norm_texto_nombre(nombre_raw)
            if nombre_raw and nombre_raw.lower() != 'none'
            else ""
        )

        fecha = _parse_fecha_simple(_get(idx_fecha))
        if fecha is None:
            resultado.errores.append(FilaError(
                fila=row_offset,
                id_usuario=num_emp or nombre_raw or "?",
                timestamp=str(_get(idx_fecha) or ''),
                motivo="Fecha inválida",
            ))
            continue

        fecha_iso = fecha.isoformat()
        hora_entrada: Optional[str] = None
        hora_salida: Optional[str] = None
        if idx_combo_plantilla is not None:
            hora_entrada, hora_salida = _parse_ingreso_salida_reporte_checador(
                _get(idx_combo_plantilla)
            )
        if not hora_entrada and not hora_salida:
            hora_entrada = _parse_hora_simple(_get(idx_ent))
            hora_salida = _parse_hora_simple(_get(idx_sal))

        if not hora_entrada and not hora_salida:
            resultado.omitidas_sin_registro += 1
            if dry_run:
                resultado.vista_previa.append(
                    {
                        "fila": row_offset,
                        "numero_archivo": num_emp,
                        "nombre_archivo": nombre_raw or None,
                        "empleado_id": None,
                        "numero_bd": None,
                        "nombre_bd": None,
                        "fecha": fecha_iso,
                        "tipo": None,
                        "hora": None,
                        "fecha_hora_local": None,
                        "estado": "sin_registro_dia_libre",
                        "motivo": None,
                    }
                )
            continue

        emp_id: Optional[int] = None
        err_nom: Optional[str] = None
        if num_emp:
            emp_id = _buscar_empleado(num_emp)
        if emp_id is None and nombre_norm:
            emp_id, err_nom = _resolver_empleado_por_nombre_importacion(
                nombre_norm, nombre_map, db, empresa_id
            )

        if emp_id is None:
            if num_emp and nombre_norm:
                motivo = (
                    f"Empleado '{num_emp}' no encontrado; por nombre ({nombre_raw}): "
                    f"{err_nom or 'sin coincidencia'}."
                )
            elif num_emp:
                motivo = f"Empleado '{num_emp}' no encontrado en la base de datos"
            else:
                motivo = err_nom or "No se encontró empleado con ese nombre."
            resultado.errores.append(FilaError(
                fila=row_offset,
                id_usuario=num_emp or nombre_raw or "?",
                timestamp=fecha_iso,
                motivo=motivo,
            ))
            continue

        _precargar(emp_id)

        eventos_dia = _eventos_desde_entrada_salida_csv(
            db, emp_id, fecha, hora_entrada, hora_salida
        )
        for tipo, dt_naive in eventos_dia:
            hora_str = dt_naive.strftime("%H:%M")
            ts = to_utc(dt_naive)
            num_bd, nom_bd = _meta_empleado(db, emp_id, meta_cache)

            if _ya_existe_s(emp_id, fecha_iso, tipo):
                resultado.omitidas_duplicado += 1
                if dry_run:
                    resultado.vista_previa.append(
                        {
                            "fila": row_offset,
                            "numero_archivo": num_emp,
                            "nombre_archivo": nombre_raw or None,
                            "empleado_id": emp_id,
                            "numero_bd": num_bd,
                            "nombre_bd": nom_bd,
                            "fecha": fecha_iso,
                            "tipo": tipo.value,
                            "hora": hora_str,
                            "fecha_hora_local": f"{fecha_iso} {hora_str}",
                            "estado": "duplicado",
                            "motivo": None,
                        }
                    )
                continue

            if dry_run:
                resultado.vista_previa.append(
                    {
                        "fila": row_offset,
                        "numero_archivo": num_emp,
                        "nombre_archivo": nombre_raw or None,
                        "empleado_id": emp_id,
                        "numero_bd": num_bd,
                        "nombre_bd": nom_bd,
                        "fecha": fecha_iso,
                        "tipo": tipo.value,
                        "hora": hora_str,
                        "fecha_hora_local": f"{fecha_iso} {hora_str}",
                        "estado": "insertar",
                        "motivo": None,
                    }
                )
                _marcar_s(emp_id, fecha_iso, tipo)
                resultado.insertadas += 1
            else:
                assert dispositivo is not None
                db.add(
                    am.Asistencia(
                        empleado_id=emp_id,
                        dispositivo_id=dispositivo.id,
                        timestamp=ts,
                        tipo=tipo,
                        es_tiempo_extra=False,
                        sincronizado=True,
                    )
                )
                _marcar_s(emp_id, fecha_iso, tipo)
                resultado.insertadas += 1

    return resultado


# ════════════════════════════════════════════════════════════════════════════
# AUTO-DETECCIÓN + GENERADOR DE PLANTILLA
# ════════════════════════════════════════════════════════════════════════════

def _detectar_formato(contenido: bytes, filename: Optional[str] = None) -> str:
    """Devuelve 'simple', 'biotime' o 'reporte_checador'."""

    def _escanea_muestra(muestra: list[tuple]) -> Optional[str]:
        for row in muestra:
            if not row:
                continue
            vals = [str(v or "").strip().lower() for v in row]
            tiene_ingreso_salida = any(
                (
                    ("ingreso" in v and "salida" in v)
                    or ("entrada" in v and "salida" in v)
                )
                for v in vals
            )
            tiene_col_empleado = any(
                (
                    ("empleado" in v and ("no" in v or "nº" in v or "n°" in v or "#" in v))
                    or ("numero empleado" in v)
                )
                for v in vals
            )
            if tiene_ingreso_salida and (tiene_col_empleado or any(v == "fecha" for v in vals)):
                return "reporte_checador"
            if any(
                "numero_empleado" in v or "numero empleado" in v or "# emp" in v for v in vals
            ):
                return "simple"
            if any(v in ("fecha", "date") for v in vals) and any(
                v in ("nombre", "empleado", "name", "colaborador", "trabajador") for v in vals
            ):
                return "simple"
            if any("tiempo" in v or "id de usuario" in v for v in vals):
                return "biotime"
        return None

    try:
        if _archivo_es_xls(filename, contenido):
            filas = _filas_desde_xls(contenido)
            hit = _escanea_muestra(filas[:_FILAS_MUESTRA_DETECCION_FORMATO])
            return hit if hit else "simple"
        import io

        wb = openpyxl.load_workbook(
            filename=io.BytesIO(contenido), data_only=True, read_only=True
        )
        ws = wb.active
        muestra: list[tuple] = []
        for row in ws.iter_rows(max_row=_FILAS_MUESTRA_DETECCION_FORMATO, values_only=True):
            if row:
                muestra.append(tuple(row))
        wb.close()
        hit = _escanea_muestra(muestra)
        return hit if hit else "simple"
    except Exception:
        pass
    return "simple"


def importar_historico_auto(
    contenido: bytes,
    db: Session,
    empresa_id: Optional[int] = None,
    dry_run: bool = False,
    filename: Optional[str] = None,
) -> ResultadoImportacion:
    """Detecta el formato automáticamente y llama al parser correspondiente."""
    fmt = _detectar_formato(contenido, filename)
    if fmt == "biotime":
        r = importar_checadas_desde_xlsx_biotime(
            contenido,
            db,
            empresa_id=empresa_id,
            dry_run=dry_run,
            filename=filename,
        )
    elif fmt == "reporte_checador":
        r = importar_checadas_desde_reporte_checador(
            contenido,
            db,
            empresa_id=empresa_id,
            dry_run=dry_run,
            filename=filename,
        )
    else:
        try:
            r = importar_checadas_desde_plantilla_simple(
                contenido,
                db,
                empresa_id=empresa_id,
                dry_run=dry_run,
                filename=filename,
            )
        except ValueError as e:
            # Fallback: algunos reportes legacy no se detectan por variaciones de encabezado.
            msg = str(e).lower()
            if "no se encontró el encabezado" in msg or "no se encontro el encabezado" in msg:
                r = importar_checadas_desde_reporte_checador(
                    contenido,
                    db,
                    empresa_id=empresa_id,
                    dry_run=dry_run,
                    filename=filename,
                )
                fmt = "reporte_checador"
            else:
                raise
    r.formato = fmt
    return r


def generar_plantilla_xlsx() -> bytes:
    """Genera un XLSX con solo los encabezados necesarios para importación histórica de checadas."""
    import io

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Checadas"

    headers = ["Numero_Empleado", "Nombre", "Fecha", "Hora_Entrada", "Hora_Salida"]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
